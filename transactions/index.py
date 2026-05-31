@login_required
def generer_rapport_admin(request):
    """
    RAPPORT ADMINISTRATEUR - Version Finale EXCEL
    ============================================
    Contient :
    1. Onglet "RECAP ADMIN" - Totaux généraux + Stats Opérateurs + Liste Agents + Liste Assistants
    2. Onglet "TRANSACTIONS" - Toutes les transactions (Agents + Assistants)
    3. Onglet "DEMANDES" - Toutes les demandes
    4. Onglets individuels pour chaque Agent (soldes à gauche, transactions à droite, demandes en bas)
    5. Onglets individuels pour chaque Assistant (soldes à gauche, transactions à droite, demandes en bas)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.utils.timezone import now
    from datetime import datetime, timedelta
    from django.db.models import Sum, Q
    from transactions.models import Transaction, Agent, Assistant, Admin, Caisse, DemandeApprovisionnement
    from django.contrib.auth.models import User
    
    # ==================== PARAMETRES ====================
    format_type = request.GET.get('format', 'excel')
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    today = now().date()
    
    if date_debut_str and date_fin_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            date_debut = today - timedelta(days=30)
            date_fin = today
    else:
        date_debut = today - timedelta(days=30)
        date_fin = today
    
    # ==================== STYLES EXCEL ====================
    COLOR_PRIMARY = "1a73e8"
    COLOR_DARK = "202124"
    COLOR_SUCCESS = "28a745"
    COLOR_DANGER = "dc3545"
    
    font_title = Font(bold=True, size=16, color="FFFFFF")
    font_subtitle = Font(bold=True, size=12, color=COLOR_PRIMARY)
    font_header = Font(bold=True, size=10, color="FFFFFF")
    font_bold = Font(bold=True)
    font_money = Font(bold=True, size=11, color=COLOR_SUCCESS)
    font_entree = Font(bold=True, size=14, color=COLOR_SUCCESS)
    font_sortie = Font(bold=True, size=14, color=COLOR_DANGER)
    
    fill_title = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
    fill_header = PatternFill(start_color=COLOR_DARK, end_color=COLOR_DARK, fill_type="solid")
    fill_success = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    fill_warning = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    fill_info = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
    fill_entree = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    fill_sortie = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    fill_admin = PatternFill(start_color="e2f0d9", end_color="e2f0d9", fill_type="solid")
    
    align_center = Alignment(horizontal='center', vertical='center')
    
    # ==================== COLLECTE DES DONNEES ====================
    all_transactions = Transaction.objects.filter(
        date__date__gte=date_debut,
        date__date__lte=date_fin
    ).select_related('user')
    
    # Totaux généraux
    total_entree = int(all_transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
    total_sortie = int(all_transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
    total_commission = int(all_transactions.aggregate(Sum('commission'))['commission__sum'] or 0)
    total_transactions = all_transactions.count()
    
    # Statistiques par opérateur
    stats = {
        'orange': {
            'depot': int(all_transactions.filter(operateur='orange', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
            'retrait': int(all_transactions.filter(operateur='orange', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
            'credit': int(all_transactions.filter(operateur='orange', type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0),
        },
        'malitel': {
            'depot': int(all_transactions.filter(operateur='malitel', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
            'retrait': int(all_transactions.filter(operateur='malitel', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
            'credit': int(all_transactions.filter(operateur='malitel', type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0),
        },
        'telecel': {
            'depot': int(all_transactions.filter(operateur='telecel', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
            'retrait': int(all_transactions.filter(operateur='telecel', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
            'credit': int(all_transactions.filter(operateur='telecel', type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0),
        },
        'wave': {
            'depot': int(all_transactions.filter(operateur='wave', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
            'retrait': int(all_transactions.filter(operateur='wave', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
        },
    }
    
    all_demandes = DemandeApprovisionnement.objects.filter(
        date_demande__date__gte=date_debut,
        date_demande__date__lte=date_fin
    ).order_by('-date_demande')
    
    # ==================== COLLECTE DES AGENTS ====================
    agents_data = []
    
    for agent in Agent.objects.all():
        agent_transactions = all_transactions.filter(user=agent.user)
        try:
            caisse = Caisse.objects.get(user=agent.user)
            solde_cash = int(caisse.solde_cash or 0)
            solde_uv = int(caisse.solde_uv or 0)
            solde_wave = int(caisse.solde_wave or 0)
        except Caisse.DoesNotExist:
            solde_cash = solde_uv = solde_wave = 0
        
        transactions_auj = agent_transactions.filter(date__date=today)
        cash_depot = int(transactions_auj.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        cash_retrait = int(transactions_auj.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        solde_cash_hier = solde_cash - (cash_depot - cash_retrait)
        
        uv_depot = int(transactions_auj.filter(operateur__in=['orange','malitel','telecel'], type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        uv_retrait = int(transactions_auj.filter(operateur__in=['orange','malitel','telecel'], type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        uv_credit = int(transactions_auj.filter(operateur__in=['orange','malitel','telecel'], type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0)
        solde_uv_hier = solde_uv - (uv_retrait - uv_depot - uv_credit)
        
        wave_depot = int(transactions_auj.filter(operateur='wave', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        wave_retrait = int(transactions_auj.filter(operateur='wave', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        solde_wave_hier = solde_wave - (wave_retrait - wave_depot)
        
        total_entree_agent = int(agent_transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        total_sortie_agent = int(agent_transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        commission_agent = int(agent_transactions.aggregate(Sum('commission'))['commission__sum'] or 0)
        
        agents_data.append({
            'nom': agent.nom,
            'telephone': agent.telephone or '-',
            'email': agent.user.email if agent.user else '-',
            'solde_cash': solde_cash,
            'solde_cash_hier': solde_cash_hier,
            'solde_uv': solde_uv,
            'solde_uv_hier': solde_uv_hier,
            'solde_wave': solde_wave,
            'solde_wave_hier': solde_wave_hier,
            'total_entree': total_entree_agent,
            'total_sortie': total_sortie_agent,
            'commission': commission_agent,
            'total_transactions': agent_transactions.count(),
            'transactions': agent_transactions,
            'demandes': all_demandes.filter(agent=agent),
        })
    
    # ==================== COLLECTE DES ASSISTANTS ====================
    assistants_data = []
    
    for assistant in Assistant.objects.all():
        assistant_transactions = all_transactions.filter(user=assistant.user)
        
        total_entree_assistant = int(assistant_transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        total_sortie_assistant = int(assistant_transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        commission_assistant = int(assistant_transactions.aggregate(Sum('commission'))['commission__sum'] or 0)
        
        assistants_data.append({
            'nom': assistant.nom,
            'telephone': assistant.telephone or '-',
            'email': assistant.user.email if assistant.user else '-',
            'total_entree': total_entree_assistant,
            'total_sortie': total_sortie_assistant,
            'commission': commission_assistant,
            'total_transactions': assistant_transactions.count(),
            'transactions': assistant_transactions,
            'demandes': all_demandes.filter(assistant_destinataire=assistant),
        })
    
    # ==================== EXPORT EXCEL ====================
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="rapport_admin_{date_debut}_{date_fin}.csv"'
        response.write('\ufeff')
        import csv
        writer = csv.writer(response)
        writer.writerow([f"RAPPORT ADMINISTRATEUR"])
        writer.writerow([f"Periode: {date_debut} au {date_fin}"])
        writer.writerow([])
        writer.writerow(["TOTAUX GENERAUX"])
        writer.writerow(["Total Entrees", f"{total_entree:,.0f} FCFA"])
        writer.writerow(["Total Sorties", f"{total_sortie:,.0f} FCFA"])
        writer.writerow(["Total Commission", f"{total_commission:,.0f} FCFA"])
        writer.writerow(["Nombre de transactions", total_transactions])
        return response
    
    # ==================== EXCEL PRINCIPAL ====================
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="rapport_admin_{date_debut}_{date_fin}.xlsx"'
    
    wb = Workbook()
    
    # ==================== ONGLET 1: RECAP ADMIN (RESTE IDENTIQUE) ====================
    ws = wb.active
    ws.title = "1. RECAP ADMIN"
    
    ws.merge_cells('A1:I1')
    ws['A1'] = "📊 RAPPORT ADMINISTRATEUR"
    ws['A1'].font = font_title
    ws['A1'].fill = fill_title
    ws['A1'].alignment = align_center
    
    ws['A2'] = f"📅 Periode: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    ws['A2'].font = font_bold
    ws['A3'] = f"⏰ Date d'export: {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}"
    
    current_row = 5
    
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "💰 TOTAUX GENERAUX"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_info
    current_row += 1
    
    ws[f'A{current_row}'] = "📈 TOTAL ENTREES"
    ws[f'A{current_row}'].font = font_entree
    ws[f'B{current_row}'] = f"{total_entree:,.0f} FCFA"
    ws[f'B{current_row}'].font = font_entree
    ws[f'B{current_row}'].fill = fill_entree
    ws[f'D{current_row}'] = "📉 TOTAL SORTIES"
    ws[f'D{current_row}'].font = font_sortie
    ws[f'E{current_row}'] = f"{total_sortie:,.0f} FCFA"
    ws[f'E{current_row}'].font = font_sortie
    ws[f'E{current_row}'].fill = fill_sortie
    current_row += 1
    
    ws[f'A{current_row}'] = "🎯 TOTAL COMMISSION"
    ws[f'A{current_row}'].font = font_bold
    ws[f'B{current_row}'] = f"{total_commission:,.0f} FCFA"
    ws[f'B{current_row}'].font = font_money
    ws[f'D{current_row}'] = "📋 NOMBRE DE TRANSACTIONS"
    ws[f'D{current_row}'].font = font_bold
    ws[f'E{current_row}'] = total_transactions
    current_row += 2
    
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "📱 STATISTIQUES PAR OPERATEUR"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_info
    current_row += 1
    
    headers = ['Opérateur', 'Dépôts', 'Retraits', 'Crédits', 'Total', 'Commission']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    current_row += 1
    
    ops_data = [('Orange', 'orange'), ('Malitel', 'malitel'), ('Telecel', 'telecel'), ('Wave', 'wave')]
    for op_name, op_key in ops_data:
        ws.cell(row=current_row, column=1, value=op_name)
        ws.cell(row=current_row, column=2, value=f"{stats[op_key]['depot']:,.0f} FCFA")
        ws.cell(row=current_row, column=3, value=f"{stats[op_key]['retrait']:,.0f} FCFA")
        ws.cell(row=current_row, column=4, value=f"{stats[op_key]['credit']:,.0f} FCFA" if op_key != 'wave' else "-")
        total = stats[op_key]['depot'] + stats[op_key]['retrait'] + (stats[op_key].get('credit', 0) if op_key != 'wave' else 0)
        ws.cell(row=current_row, column=5, value=f"{total:,.0f} FCFA")
        ws.cell(row=current_row, column=6, value=f"{int(total * 0.01):,.0f} FCFA")
        current_row += 1
    current_row += 2
    
    ws.merge_cells(f'A{current_row}:I{current_row}')
    ws[f'A{current_row}'] = "👥 LISTE DES AGENTS"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_info
    current_row += 1
    
    headers_agents = ['Agent', 'Téléphone', 'Email', 'Entrées', 'Sorties', 'Commission', 'Nb Ops']
    for col, header in enumerate(headers_agents, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    current_row += 1
    
    for a in agents_data:
        ws.cell(row=current_row, column=1, value=a['nom'])
        ws.cell(row=current_row, column=2, value=a['telephone'])
        ws.cell(row=current_row, column=3, value=a['email'])
        ws.cell(row=current_row, column=4, value=f"{a['total_entree']:,.0f} FCFA")
        ws.cell(row=current_row, column=5, value=f"{a['total_sortie']:,.0f} FCFA")
        ws.cell(row=current_row, column=6, value=f"{a['commission']:,.0f} FCFA")
        ws.cell(row=current_row, column=7, value=a['total_transactions'])
        current_row += 1
    
    current_row += 2
    
    ws.merge_cells(f'A{current_row}:I{current_row}')
    ws[f'A{current_row}'] = "👥 LISTE DES ASSISTANTS"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_info
    current_row += 1
    
    headers_assistants = ['Assistant', 'Téléphone', 'Email', 'Entrées', 'Sorties', 'Commission', 'Nb Ops']
    for col, header in enumerate(headers_assistants, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    current_row += 1
    
    for a in assistants_data:
        ws.cell(row=current_row, column=1, value=a['nom'])
        ws.cell(row=current_row, column=2, value=a['telephone'])
        ws.cell(row=current_row, column=3, value=a['email'])
        ws.cell(row=current_row, column=4, value=f"{a['total_entree']:,.0f} FCFA")
        ws.cell(row=current_row, column=5, value=f"{a['total_sortie']:,.0f} FCFA")
        ws.cell(row=current_row, column=6, value=f"{a['commission']:,.0f} FCFA")
        ws.cell(row=current_row, column=7, value=a['total_transactions'])
        current_row += 1
    
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # ==================== ONGLET 2: TRANSACTIONS ====================
    ws2 = wb.create_sheet("2. TRANSACTIONS")
    ws2.merge_cells('A1:H1')
    ws2['A1'] = "DETAIL DES TRANSACTIONS"
    ws2['A1'].font = font_title
    ws2['A1'].fill = fill_title
    ws2['A1'].alignment = align_center
    ws2['A2'] = f"📅 Periode: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    
    row = 4
    headers_trans = ['Date', 'Référence', 'Type', 'Opérateur', 'Client', 'Montant', 'Commission', 'Utilisateur']
    for col, header in enumerate(headers_trans, 1):
        cell = ws2.cell(row=row, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    row += 1
    
    for t in all_transactions.order_by('-date'):
        ws2.cell(row=row, column=1, value=t.date.strftime('%d/%m/%Y %H:%M'))
        ws2.cell(row=row, column=2, value=t.reference)
        ws2.cell(row=row, column=3, value=t.get_type_transaction_display())
        ws2.cell(row=row, column=4, value=t.get_operateur_display())
        ws2.cell(row=row, column=5, value=t.numero_client)
        ws2.cell(row=row, column=6, value=f"{int(t.montant):,.0f}")
        ws2.cell(row=row, column=6).font = font_money
        ws2.cell(row=row, column=7, value=f"{int(t.commission):,.0f}")
        ws2.cell(row=row, column=8, value=t.user.username if t.user else "-")
        row += 1
    
    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 16
    
    # ==================== ONGLET 3: DEMANDES ====================
    if all_demandes.exists():
        ws3 = wb.create_sheet("3. DEMANDES")
        ws3.merge_cells('A1:G1')
        ws3['A1'] = "DEMANDES D'APPROVISIONNEMENT"
        ws3['A1'].font = font_title
        ws3['A1'].fill = fill_title
        ws3['A1'].alignment = align_center
        ws3['A2'] = f"📅 Periode: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        
        row = 4
        headers_dem = ['Date', 'Type', 'Montant', 'Statut', 'Motif', 'Agent', 'Assistant']
        for col, header in enumerate(headers_dem, 1):
            cell = ws3.cell(row=row, column=col, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        row += 1
        
        for d in all_demandes:
            ws3.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y %H:%M'))
            ws3.cell(row=row, column=2, value=d.get_type_echange_display())
            ws3.cell(row=row, column=3, value=f"{int(d.montant):,.0f} FCFA")
            ws3.cell(row=row, column=4, value=d.get_statut_display())
            ws3.cell(row=row, column=5, value=d.motif or "-")
            ws3.cell(row=row, column=6, value=d.agent.nom if d.agent else "-")
            ws3.cell(row=row, column=7, value=d.assistant_destinataire.nom if d.assistant_destinataire else "-")
            row += 1
        
        for col in range(1, 8):
            ws3.column_dimensions[get_column_letter(col)].width = 18
    
    # ==================== ONGLETS INDIVIDUELS POUR CHAQUE AGENT ====================
    for a in agents_data:
        nom_feuille = f"Agent_{a['nom'][:20]}".replace(' ', '_').replace('-', '_')
        ws_agent = wb.create_sheet(nom_feuille[:25])
        
        # ========== COLONNE GAUCHE (A à D) : SOLDES ET RÉSUMÉ ==========
        ws_agent.merge_cells('A1:D1')
        ws_agent['A1'] = f"AGENT : {a['nom']}"
        ws_agent['A1'].font = font_title
        ws_agent['A1'].fill = fill_title
        ws_agent['A1'].alignment = align_center
        
        ws_agent['A2'] = f"📞 {a['telephone']}  |  ✉️ {a['email']}"
        ws_agent['A3'] = f"📅 Periode: {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        
        row = 5
        
        ws_agent.merge_cells(f'A{row}:D{row}')
        ws_agent[f'A{row}'] = "💰 SOLDES"
        ws_agent[f'A{row}'].font = font_subtitle
        ws_agent[f'A{row}'].fill = fill_info
        row += 1
        
        ws_agent[f'A{row}'] = "Compte"
        ws_agent[f'B{row}'] = "Aujourd'hui"
        ws_agent[f'C{row}'] = "Hier"
        ws_agent[f'D{row}'] = "Variation"
        for col in range(1, 5):
            ws_agent.cell(row=row, column=col).font = font_header
            ws_agent.cell(row=row, column=col).fill = fill_header
        row += 1
        
        ws_agent[f'A{row}'] = "💵 Cash"
        ws_agent[f'B{row}'] = f"{a['solde_cash']:,.0f} FCFA"
        ws_agent[f'C{row}'] = f"{a['solde_cash_hier']:,.0f} FCFA"
        ws_agent[f'D{row}'] = f"{a['solde_cash'] - a['solde_cash_hier']:+,.0f} FCFA"
        row += 1
        
        ws_agent[f'A{row}'] = "📱 UV Touchpiont"
        ws_agent[f'B{row}'] = f"{a['solde_uv']:,.0f} FCFA"
        ws_agent[f'C{row}'] = f"{a['solde_uv_hier']:,.0f} FCFA"
        ws_agent[f'D{row}'] = f"{a['solde_uv'] - a['solde_uv_hier']:+,.0f} FCFA"
        row += 1
        
        ws_agent[f'A{row}'] = "🌊 UV Wave"
        ws_agent[f'B{row}'] = f"{a['solde_wave']:,.0f} FCFA"
        ws_agent[f'C{row}'] = f"{a['solde_wave_hier']:,.0f} FCFA"
        ws_agent[f'D{row}'] = f"{a['solde_wave'] - a['solde_wave_hier']:+,.0f} FCFA"
        row += 2
        
        ws_agent[f'A{row}'] = "📊 RÉSUMÉ"
        ws_agent[f'A{row}'].font = font_bold
        ws_agent[f'B{row}'] = f"Entrées: {a['total_entree']:,.0f} FCFA"
        ws_agent[f'C{row}'] = f"Sorties: {a['total_sortie']:,.0f} FCFA"
        ws_agent[f'D{row}'] = f"Commission: {a['commission']:,.0f} FCFA"
        row += 2
        
        # ========== DEMANDES (EN BAS, COLONNE A à D) ==========
        ws_agent.merge_cells(f'A{row}:D{row}')
        ws_agent[f'A{row}'] = "📨 DEMANDES D'APPROVISIONNEMENT"
        ws_agent[f'A{row}'].font = font_subtitle
        ws_agent[f'A{row}'].fill = fill_info
        row += 1
        
        headers_dem = ['Date', 'Type', 'Montant', 'Statut']
        for col, header in enumerate(headers_dem, 1):
            cell = ws_agent.cell(row=row, column=col, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        row += 1
        
        for d in a['demandes']:
            ws_agent.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y'))
            ws_agent.cell(row=row, column=2, value=d.get_type_echange_display())
            ws_agent.cell(row=row, column=3, value=f"{int(d.montant):,.0f} FCFA")
            ws_agent.cell(row=row, column=4, value=d.get_statut_display())
            row += 1
        
        if len(a['demandes']) == 0:
            ws_agent.cell(row=row, column=1, value="Aucune demande")
        
        # ========== COLONNE DROITE (F à K) : TRANSACTIONS (remontées plus haut) ==========
        col_trans = 6  # Colonne F
        row_trans = 5  # Commence à la ligne 5 (juste après le titre)
        
        ws_agent.merge_cells(start_row=row_trans, start_column=col_trans, end_row=row_trans, end_column=col_trans+5)
        ws_agent.cell(row=row_trans, column=col_trans, value="📊 LISTE DES TRANSACTIONS")
        ws_agent.cell(row=row_trans, column=col_trans).font = font_subtitle
        ws_agent.cell(row=row_trans, column=col_trans).fill = fill_info
        ws_agent.cell(row=row_trans, column=col_trans).alignment = align_center
        row_trans += 1
        
        headers_trans_pers = ['Date', 'Type', 'Opérateur', 'Client', 'Montant', 'Commission']
        for col, header in enumerate(headers_trans_pers, col_trans):
            cell = ws_agent.cell(row=row_trans, column=col, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        row_trans += 1
        
        for t in a['transactions'].order_by('-date'):
            ws_agent.cell(row=row_trans, column=col_trans, value=t.date.strftime('%d/%m/%Y'))
            ws_agent.cell(row=row_trans, column=col_trans+1, value=t.get_type_transaction_display())
            ws_agent.cell(row=row_trans, column=col_trans+2, value=t.get_operateur_display())
            ws_agent.cell(row=row_trans, column=col_trans+3, value=t.numero_client)
            ws_agent.cell(row=row_trans, column=col_trans+4, value=f"{int(t.montant):,.0f}")
            ws_agent.cell(row=row_trans, column=col_trans+5, value=f"{int(t.commission):,.0f}")
            row_trans += 1
        
        if a['transactions'].count() == 0:
            ws_agent.cell(row=row_trans, column=col_trans, value="Aucune transaction")
        
        # Ajuster largeurs
        for col in range(1, 5):
            ws_agent.column_dimensions[get_column_letter(col)].width = 18
        for col in range(6, 12):
            ws_agent.column_dimensions[get_column_letter(col)].width = 16
    
    # ==================== ONGLETS INDIVIDUELS POUR CHAQUE ASSISTANT ====================
    for a in assistants_data:
        nom_feuille = f"Assistant_{a['nom'][:20]}".replace(' ', '_').replace('-', '_')
        ws_assistant = wb.create_sheet(nom_feuille[:25])
        
        # ========== COLONNE GAUCHE (A à D) : RÉSUMÉ ==========
        ws_assistant.merge_cells('A1:D1')
        ws_assistant['A1'] = f"ASSISTANT : {a['nom']}"
        ws_assistant['A1'].font = font_title
        ws_assistant['A1'].fill = fill_title
        ws_assistant['A1'].alignment = align_center
        
        ws_assistant['A2'] = f"📞 {a['telephone']}  |  ✉️ {a['email']}"
        ws_assistant['A3'] = f"📅 Periode: {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        ws_assistant['A4'] = "ℹ️ Les assistants partagent le solde de l'administrateur"
        ws_assistant['A4'].fill = fill_info
        
        row = 6
        
        ws_assistant.merge_cells(f'A{row}:D{row}')
        ws_assistant[f'A{row}'] = "📊 RÉSUMÉ DES OPÉRATIONS"
        ws_assistant[f'A{row}'].font = font_subtitle
        ws_assistant[f'A{row}'].fill = fill_info
        row += 1
        
        ws_assistant[f'A{row}'] = "💰 Total Entrées"
        ws_assistant[f'B{row}'] = f"{a['total_entree']:,.0f} FCFA"
        ws_assistant[f'C{row}'] = "💸 Total Sorties"
        ws_assistant[f'D{row}'] = f"{a['total_sortie']:,.0f} FCFA"
        row += 1
        
        ws_assistant[f'A{row}'] = "🎯 Commission totale"
        ws_assistant[f'B{row}'] = f"{a['commission']:,.0f} FCFA"
        ws_assistant[f'C{row}'] = "📋 Nombre de transactions"
        ws_assistant[f'D{row}'] = a['total_transactions']
        row += 2
        
        # ========== DEMANDES (EN BAS, COLONNE A à D) ==========
        ws_assistant.merge_cells(f'A{row}:D{row}')
        ws_assistant[f'A{row}'] = "📨 DEMANDES TRAITÉES"
        ws_assistant[f'A{row}'].font = font_subtitle
        ws_assistant[f'A{row}'].fill = fill_info
        row += 1
        
        headers_dem = ['Date', 'Type', 'Montant', 'Statut']
        for col, header in enumerate(headers_dem, 1):
            cell = ws_assistant.cell(row=row, column=col, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        row += 1
        
        for d in a['demandes']:
            ws_assistant.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y'))
            ws_assistant.cell(row=row, column=2, value=d.get_type_echange_display())
            ws_assistant.cell(row=row, column=3, value=f"{int(d.montant):,.0f} FCFA")
            ws_assistant.cell(row=row, column=4, value=d.get_statut_display())
            row += 1
        
        if len(a['demandes']) == 0:
            ws_assistant.cell(row=row, column=1, value="Aucune demande")
        
        # ========== COLONNE DROITE (F à K) : TRANSACTIONS (remontées plus haut) ==========
        col_trans = 6  # Colonne F
        row_trans = 6  # Commence à la ligne 6 (juste après le titre)
        
        ws_assistant.merge_cells(start_row=row_trans, start_column=col_trans, end_row=row_trans, end_column=col_trans+5)
        ws_assistant.cell(row=row_trans, column=col_trans, value="📊 LISTE DES TRANSACTIONS")
        ws_assistant.cell(row=row_trans, column=col_trans).font = font_subtitle
        ws_assistant.cell(row=row_trans, column=col_trans).fill = fill_info
        ws_assistant.cell(row=row_trans, column=col_trans).alignment = align_center
        row_trans += 1
        
        headers_trans_pers = ['Date', 'Type', 'Opérateur', 'Client', 'Montant', 'Commission']
        for col, header in enumerate(headers_trans_pers, col_trans):
            cell = ws_assistant.cell(row=row_trans, column=col, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        row_trans += 1
        
        for t in a['transactions'].order_by('-date'):
            ws_assistant.cell(row=row_trans, column=col_trans, value=t.date.strftime('%d/%m/%Y'))
            ws_assistant.cell(row=row_trans, column=col_trans+1, value=t.get_type_transaction_display())
            ws_assistant.cell(row=row_trans, column=col_trans+2, value=t.get_operateur_display())
            ws_assistant.cell(row=row_trans, column=col_trans+3, value=t.numero_client)
            ws_assistant.cell(row=row_trans, column=col_trans+4, value=f"{int(t.montant):,.0f}")
            ws_assistant.cell(row=row_trans, column=col_trans+5, value=f"{int(t.commission):,.0f}")
            row_trans += 1
        
        if a['transactions'].count() == 0:
            ws_assistant.cell(row=row_trans, column=col_trans, value="Aucune transaction")
        
        # Ajuster largeurs
        for col in range(1, 5):
            ws_assistant.column_dimensions[get_column_letter(col)].width = 18
        for col in range(6, 12):
            ws_assistant.column_dimensions[get_column_letter(col)].width = 16
    
    wb.save(response)
    return response