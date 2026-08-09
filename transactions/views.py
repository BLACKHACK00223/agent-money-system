# transactions/views.py
import base64
from pathlib import Path
import django
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db.models import Sum, Q
from datetime import datetime, timedelta
from decimal import Decimal
import json
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.contrib.auth import logout
from django.views.decorators.http import require_http_methods
from django.utils import timezone
import csv
from .models import Admin, Agent, Assistant, Caisse, CompteEpargneAdmin, OperationCaisse, OperationEpargne, OperationUv, Transaction, DemandeApprovisionnement, ApprovisionnementDirect
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import io

from .models import (
    Admin, Agent, Assistant, Caisse, Transaction, DemandeApprovisionnement, 
    ApprovisionnementDirect, Facture, PaiementFacture, Dette, RemboursementDette,
    CompteEpargne, OperationCompte
)
from .forms import (OrangeTransactionForm, WaveTransactionForm, 
                   MalitelTransactionForm, TelecelTransactionForm)


# ==================== AUTHENTIFICATION ====================

def logout_view(request):
    """Vue personnalisée pour la déconnexion"""
    logout(request)
    messages.success(request, 'Vous avez été déconnecté avec succès.')
    return redirect('login')


@login_required
def dashboard_redirect(request):
    """
    Redirige vers le bon tableau de bord selon le rôle
    """
    # Vérifier si l'utilisateur est un ADMIN
    try:
        admin = Admin.objects.get(user=request.user)
        return redirect('dashboard_admin')
    except Admin.DoesNotExist:
        pass
    
    # Vérifier si l'utilisateur est un AGENT
    try:
        agent = Agent.objects.get(user=request.user)
        return redirect('dashboard_agent')
    except Agent.DoesNotExist:
        pass
    
    # Vérifier si l'utilisateur est un ASSISTANT
    try:
        assistant = Assistant.objects.get(user=request.user)
        return redirect('dashboard_assistant')
    except Assistant.DoesNotExist:
        pass
    
    # Sinon, rediriger vers login
    messages.error(request, 'Vous n\'avez pas de profil configuré.')
    return redirect('login')


@login_required
def dashboard_admin(request):
    """
    Tableau de bord pour l'ADMIN
    - Voit son propre compte
    - Voit tous les agents
    - Voit les demandes en attente
    - Statistiques du jour et d'hier
    - Top agents
    """
    try:
        admin = Admin.objects.get(user=request.user)
        caisse_admin = admin.user.caisse
        
        today = datetime.now().date()
        yesterday = today - timedelta(days=1)
        
        # ========== AGENTS ET ASSISTANTS ==========
        agents = Agent.objects.filter(est_actif=True)
        assistants = Assistant.objects.filter(est_actif=True)
        total_agents = agents.count()
        total_assistants = assistants.count()
        
        # ========== TRANSACTIONS ==========
        transactions_today = Transaction.objects.filter(date__date=today)
        transactions_yesterday = Transaction.objects.filter(date__date=yesterday)
        
        # Stats du jour
        stats_today = {
            'depots': transactions_today.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0,
            'retraits': transactions_today.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0,
            'credits': transactions_today.filter(type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0,
            'commission': transactions_today.aggregate(Sum('commission'))['commission__sum'] or 0,
            'nombre': transactions_today.count(),
        }
        
        # Stats d'hier
        stats_yesterday = {
            'depots': transactions_yesterday.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0,
            'retraits': transactions_yesterday.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0,
            'credits': transactions_yesterday.filter(type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0,
            'commission': transactions_yesterday.aggregate(Sum('commission'))['commission__sum'] or 0,
            'nombre': transactions_yesterday.count(),
        }
        
        # Évolutions
        evolution = {
            'depots': ((stats_today['depots'] - stats_yesterday['depots']) / stats_yesterday['depots'] * 100) if stats_yesterday['depots'] > 0 else 0,
            'retraits': ((stats_today['retraits'] - stats_yesterday['retraits']) / stats_yesterday['retraits'] * 100) if stats_yesterday['retraits'] > 0 else 0,
            'credits': ((stats_today['credits'] - stats_yesterday['credits']) / stats_yesterday['credits'] * 100) if stats_yesterday['credits'] > 0 else 0,
            'commission': ((stats_today['commission'] - stats_yesterday['commission']) / stats_yesterday['commission'] * 100) if stats_yesterday['commission'] > 0 else 0,
            'nombre': ((stats_today['nombre'] - stats_yesterday['nombre']) / stats_yesterday['nombre'] * 100) if stats_yesterday['nombre'] > 0 else 0,
        }
        
        # ========== DEMANDES ==========
        demandes_attente = DemandeApprovisionnement.objects.filter(
            statut='en_attente'
        ).order_by('-date_demande')
        
        demandes_today = DemandeApprovisionnement.objects.filter(date_demande__date=today)
        demandes_yesterday = DemandeApprovisionnement.objects.filter(date_demande__date=yesterday)
        
        demandes_stats = {
            'aujourdhui': demandes_today.count(),
            'hier': demandes_yesterday.count(),
            'total_attente': demandes_attente.count(),
            'total_validees': DemandeApprovisionnement.objects.filter(statut='valide').count(),
            'total_refusees': DemandeApprovisionnement.objects.filter(statut='refuse').count(),
        }
        
        # ========== NOUVEAUX AGENTS ==========
        nouveaux_agents_today = Agent.objects.filter(created_at__date=today).count()
        nouveaux_agents_yesterday = Agent.objects.filter(created_at__date=yesterday).count()
        
        # ========== STATISTIQUES PAR OPÉRATEUR ==========
        stats_par_operateur = {}
        for operateur in ['orange', 'wave', 'malitel', 'telecel']:
            ops = transactions_today.filter(operateur=operateur)
            stats_par_operateur[operateur] = {
                'nombre': ops.count(),
                'montant': ops.aggregate(Sum('montant'))['montant__sum'] or 0,
            }
        
        # ========== STATISTIQUES PAR AGENT ==========
        stats_par_agent = []
        for agent in agents:
            transactions_agent = Transaction.objects.filter(
                user=agent.user,
                date__date=today
            )
            stats_par_agent.append({
                'agent': agent,
                'caisse': agent.user.caisse,
                'nombre': transactions_agent.count(),
                'montant': transactions_agent.aggregate(Sum('montant'))['montant__sum'] or 0,
            })
        
        # ========== TOP AGENTS ==========
        top_agents = []
        for agent in agents:
            nb_trans = Transaction.objects.filter(user=agent.user, date__date=today).count()
            if nb_trans > 0:
                top_agents.append({
                    'agent': agent,
                    'transactions': nb_trans,
                    'montant': Transaction.objects.filter(user=agent.user, date__date=today).aggregate(Sum('montant'))['montant__sum'] or 0,
                })
        top_agents = sorted(top_agents, key=lambda x: x['transactions'], reverse=True)[:5]
        
        # ========== TRANSACTIONS ADMIN ==========
        transactions_admin = Transaction.objects.filter(
            user=request.user,
            date__date=today
        )
        
        stats_admin = {
            'depots': transactions_admin.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0,
            'retraits': transactions_admin.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0,
            'credits': transactions_admin.filter(type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0,
            'commission': transactions_admin.aggregate(Sum('commission'))['commission__sum'] or 0,
            'nombre': transactions_admin.count(),
        }
        
        # ========== DERNIÈRES TRANSACTIONS ==========
        dernieres_transactions = Transaction.objects.all().order_by('-date')[:30]
        
        # ========== TOTAL COMMISSION ==========
        total_commission = Transaction.objects.aggregate(Sum('commission'))['commission__sum'] or 0
        
        context = {
            'title': 'Tableau de bord - ADMIN',
            'admin': admin,
            'caisse': caisse_admin,
            'agents': agents,
            'assistants': assistants,
            'total_agents': total_agents,
            'total_assistants': total_assistants,
            'stats_today': stats_today,
            'stats_yesterday': stats_yesterday,
            'evolution': evolution,
            'demandes_attente': demandes_attente,
            'demandes_stats': demandes_stats,
            'nouveaux_agents_today': nouveaux_agents_today,
            'nouveaux_agents_yesterday': nouveaux_agents_yesterday,
            'stats_par_operateur': stats_par_operateur,
            'stats_par_agent': stats_par_agent,
            'top_agents': top_agents,
            'stats_admin': stats_admin,
            'dernieres_transactions': dernieres_transactions,
            'total_commission': total_commission,
            'transactions_jour': transactions_today,
        }
        return render(request, 'transactions/dashboard_admin.html', context)
        
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas configuré comme administrateur.')
        return redirect('login')


@login_required
def dashboard_agent(request):
    """
    Tableau de bord pour l'AGENT
    """
    try:
        agent = Agent.objects.get(user=request.user)
        caisse = agent.user.caisse
        
        # Vérifier si l'agent a une caisse
        if not caisse:
            messages.error(request, 'Votre caisse n\'est pas configurée.')
            return redirect('login')
        
        # ========== RÉCUPÉRER LES ASSISTANTS ==========
        assistants = Assistant.objects.filter(est_actif=True)
        
        # Transactions du jour
        today = datetime.now().date()
        transactions_jour = Transaction.objects.filter(
            user=request.user,
            date__date=today
        )
        
        # Statistiques du jour
        stats_jour = {
            'depots': transactions_jour.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0,
            'retraits': transactions_jour.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0,
            'credits': transactions_jour.filter(type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0,
            'commission': transactions_jour.aggregate(Sum('commission'))['commission__sum'] or 0,
            'nombre': transactions_jour.count(),
        }
        
        # Statistiques par opérateur
        stats_par_operateur = {}
        for operateur in ['orange', 'wave', 'malitel', 'telecel']:
            ops = transactions_jour.filter(operateur=operateur)
            stats_par_operateur[operateur] = {
                'nombre': ops.count(),
                'montant': ops.aggregate(Sum('montant'))['montant__sum'] or 0,
            }
        
        # Dernières transactions
        dernieres_transactions = Transaction.objects.filter(
            user=request.user
        ).order_by('-date')[:20]
        
        # Demandes en cours
        demandes_en_cours = DemandeApprovisionnement.objects.filter(
            agent=agent,
            statut='en_attente'
        )
        
        # Historique des demandes
        historique_demandes = DemandeApprovisionnement.objects.filter(
            agent=agent
        ).order_by('-date_demande')[:10]
        
        context = {
            'title': 'Tableau de bord - Agent',
            'agent': agent,
            'caisse': caisse,
            'assistants': assistants,  # ← TRÈS IMPORTANT : AJOUTER CETTE LIGNE
            'stats_jour': stats_jour,
            'stats_par_operateur': stats_par_operateur,
            'transactions_jour': transactions_jour[:20],
            'dernieres_transactions': dernieres_transactions,
            'demandes_en_cours': demandes_en_cours,
            'historique_demandes': historique_demandes,
        }
        return render(request, 'transactions/dashboard_agent.html', context)
        
    except Agent.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas configuré comme agent.')
        return redirect('login')

@login_required
def dashboard_assistant(request):
    """
    Tableau de bord pour l'ASSISTANT
    - Partage la caisse de l'admin
    - Voit les demandes d'approvisionnement des agents qui lui sont destinées
    - Peut faire des transactions (dépôt, retrait, crédit)
    """
    try:
        assistant = Assistant.objects.get(user=request.user)
    except Assistant.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas configuré comme assistant.')
        return redirect('login')
    
    # L'assistant partage la caisse de son admin
    caisse = assistant.admin.user.caisse
    
    today = datetime.now().date()
    
    # Transactions du jour
    transactions_jour = Transaction.objects.filter(
        user=request.user,
        date__date=today
    )
    
    # Statistiques du jour
    stats_jour = {
        'depots': transactions_jour.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0,
        'retraits': transactions_jour.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0,
        'credits': transactions_jour.filter(type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0,
        'commission': transactions_jour.aggregate(Sum('commission'))['commission__sum'] or 0,
        'nombre': transactions_jour.count(),
    }
    
    # Dernières transactions
    dernieres_transactions = Transaction.objects.filter(
        user=request.user
    ).order_by('-date')[:20]
    
    # ========== DEMANDES REÇUES DES AGENTS ==========
    # L'assistant reçoit les demandes des agents qui l'ont choisi comme destinataire
    demandes_recues = DemandeApprovisionnement.objects.filter(
        destinataire_type='assistant',
        assistant_destinataire=assistant
    ).order_by('-date_demande')
    
    # Statistiques des demandes
    demandes_stats = {
        'en_attente': demandes_recues.filter(statut='en_attente').count(),
        'validees': demandes_recues.filter(statut='valide').count(),
        'refusees': demandes_recues.filter(statut='refuse').count(),
    }
    
    context = {
        'title': 'Tableau de bord - Assistant',
        'assistant': assistant,
        'caisse': caisse,
        'stats_jour': stats_jour,
        'transactions_jour': transactions_jour[:20],
        'dernieres_transactions': dernieres_transactions,
        'demandes_recues': demandes_recues[:10],
        'demandes_stats': demandes_stats,
    }
    return render(request, 'transactions/dashboard_assistant.html', context)

@login_required
def historique_demandes_agent(request):
    """
    Historique des demandes d'approvisionnement pour l'AGENT ou ASSISTANT
    - Pour un agent: ses demandes envoyées
    - Pour un assistant: les demandes reçues des agents
    """
    # Vérifier si c'est un agent ou un assistant
    try:
        agent = Agent.objects.get(user=request.user)
        # C'est un agent: voir ses demandes envoyées
        demandes = DemandeApprovisionnement.objects.filter(agent=agent).order_by('-date_demande')
        title = 'Mes demandes'
    except Agent.DoesNotExist:
        try:
            assistant = Assistant.objects.get(user=request.user)
            # C'est un assistant: voir les demandes reçues
            demandes = DemandeApprovisionnement.objects.filter(
                destinataire_type='assistant',
                assistant_destinataire=assistant
            ).order_by('-date_demande')
            title = 'Demandes reçues'
        except Assistant.DoesNotExist:
            messages.error(request, 'Vous n\'êtes pas autorisé.')
            return redirect('login')
    
    # Date d'aujourd'hui
    today = timezone.now().date()
    
    # ========== FILTRES ==========
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    statut = request.GET.get('statut')
    
    # Si aucun filtre de date n'est appliqué, afficher uniquement les demandes du jour
    if not date_debut and not date_fin and not statut:
        demandes = demandes.filter(date_demande__date=today)
        date_debut_display = today.strftime('%Y-%m-%d')
        date_fin_display = today.strftime('%Y-%m-%d')
    else:
        date_debut_display = date_debut
        date_fin_display = date_fin
        
        # Filtre par date début
        if date_debut:
            try:
                date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
                demandes = demandes.filter(date_demande__date__gte=date_debut_obj)
            except ValueError:
                pass
        
        # Filtre par date fin
        if date_fin:
            try:
                date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
                demandes = demandes.filter(date_demande__date__lte=date_fin_obj)
            except ValueError:
                pass
        
        # Filtre par statut
        if statut:
            demandes = demandes.filter(statut=statut)
    
    # Calcul des statistiques
    stats = {
        'attente': demandes.filter(statut='attente').count(),
        'valide': demandes.filter(statut='valide').count(),
        'refuse': demandes.filter(statut='refuse').count(),
    }
    
    context = {
        'title': title,
        'demandes': demandes,
        'stats': stats,
        'date_debut': date_debut_display,
        'date_fin': date_fin_display,
    }
    return render(request, 'transactions/historique_demandes.html', context)
# ==================== TRANSACTIONS ====================

@login_required
def transaction_user(request, operateur, type_transaction):
    """
    Transaction pour l'utilisateur connecté (ADMIN, AGENT ou ASSISTANT)
    - ADMIN: utilise sa propre caisse
    - AGENT: utilise sa propre caisse
    - ASSISTANT: utilise la caisse de son ADMIN (impacte le solde admin)
    Vérifie les soldes avant d'effectuer la transaction
    Supporte les requêtes AJAX pour le modal de confirmation
    """
    # Vérifier le rôle
    is_admin = hasattr(request.user, 'admin_profile')
    is_agent = hasattr(request.user, 'agent_profile')
    is_assistant = hasattr(request.user, 'assistant_profile')
    
    if not (is_admin or is_agent or is_assistant):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Vous n\'êtes pas configuré.'})
        messages.error(request, 'Vous n\'êtes pas configuré.')
        return redirect('login')
    
    # Récupérer la caisse
    # - ADMIN: sa propre caisse
    # - AGENT: sa propre caisse
    # - ASSISTANT: la caisse de son ADMIN (impacte le solde admin)
    if is_assistant:
        assistant = request.user.assistant_profile
        caisse = assistant.admin.user.caisse  # ← Caisse de l'ADMIN
    else:
        caisse = request.user.caisse  # ← Propre caisse
    
    # Déterminer le formulaire
    forms = {
        'orange': OrangeTransactionForm,
        'wave': WaveTransactionForm,
        'malitel': MalitelTransactionForm,
        'telecel': TelecelTransactionForm,
    }
    
    form_class = forms.get(operateur)
    if not form_class:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Opérateur invalide.'})
        messages.error(request, 'Opérateur invalide.')
        return redirect('dashboard_redirect')
    
    # Vérifier si l'opération est disponible pour cet opérateur
    if operateur == 'wave' and type_transaction == 'credit':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Le crédit Wave n\'est pas disponible'})
        messages.error(request, '❌ Le crédit Wave n\'est pas disponible')
        return redirect('dashboard_redirect')
    
    if request.method == 'POST':
        form = form_class(request.POST)
        if form.is_valid():
            montant = form.cleaned_data['montant']
            
            # ========== VÉRIFICATIONS DES SOLDES ==========
            solde_ok = True
            message_erreur = ""
            
            # ORANGE, MALITEL, TELECEL (via UV Touspiont)
            if operateur in ['orange', 'malitel', 'telecel']:
                if type_transaction == 'depot':
                    # DÉPÔT: client donne cash → agent donne ses UV (il faut assez d'UV)
                    if caisse.solde_uv < montant:
                        solde_ok = False
                        message_erreur = f"❌ Solde UV Touspiont insuffisant. Solde actuel: {caisse.solde_uv:,.0f} FCFA"
                    
                elif type_transaction == 'retrait':
                    # RETRAIT: client prend cash → agent donne son cash (il faut assez de cash)
                    if caisse.solde_cash < montant:
                        solde_ok = False
                        message_erreur = f"❌ Solde Cash insuffisant. Solde actuel: {caisse.solde_cash:,.0f} FCFA"
                    
                elif type_transaction == 'credit':
                    # CRÉDIT: client recharge → agent donne ses UV (il faut assez d'UV)
                    if caisse.solde_uv < montant:
                        solde_ok = False
                        message_erreur = f"❌ Solde UV Touspiont insuffisant pour le crédit. Solde actuel: {caisse.solde_uv:,.0f} FCFA"
            
            # WAVE
            elif operateur == 'wave':
                if type_transaction == 'depot':
                    # DÉPÔT WAVE: client donne cash → agent donne ses Wave (il faut assez de Wave)
                    if caisse.solde_wave < montant:
                        solde_ok = False
                        message_erreur = f"❌ Solde Wave insuffisant. Solde actuel: {caisse.solde_wave:,.0f} FCFA"
                    
                elif type_transaction == 'retrait':
                    # RETRAIT WAVE: client prend cash → agent donne son cash (il faut assez de cash)
                    if caisse.solde_cash < montant:
                        solde_ok = False
                        message_erreur = f"❌ Solde Cash insuffisant. Solde actuel: {caisse.solde_cash:,.0f} FCFA"
            
            # Si solde insuffisant
            if not solde_ok:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': message_erreur})
                messages.error(request, message_erreur)
                return redirect('dashboard_redirect')
            
            # Déterminer le rôle et l'admin associé
            if is_admin:
                role = 'admin'
                assistant_admin = None
            elif is_agent:
                role = 'agent'
                assistant_admin = None
            else:
                role = 'assistant'
                assistant_admin = request.user.assistant_profile.admin
            
            # Créer et sauvegarder la transaction
            transaction = form.save(commit=False)
            transaction.user = request.user
            transaction.type_transaction = type_transaction
            transaction.operateur = operateur
            transaction.role = role
            transaction.assistant_admin = assistant_admin
            
            try:
                transaction.save()  # ← Ici la logique du modèle Transaction met à jour la caisse
                
                # Réponse JSON pour AJAX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'reference': transaction.reference,
                        'message': f'Transaction {operateur.capitalize()} effectuée avec succès!'
                    })
                
                messages.success(request, f'✅ Transaction {operateur.capitalize()} effectuée avec succès! Réf: {transaction.reference}')
                
                # Redirection selon le rôle
                if is_admin:
                    return redirect('dashboard_admin')
                elif is_agent:
                    return redirect('dashboard_agent')
                else:
                    return redirect('dashboard_assistant')
                    
            except Exception as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': str(e)})
                messages.error(request, f'Erreur: {str(e)}')
        else:
            # Erreurs du formulaire
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': list(form.errors.values())[0][0]})
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = form_class(initial={'type_transaction': type_transaction})
    
    # Déterminer l'URL de redirection pour le formulaire
    if is_admin:
        dashboard_url = 'dashboard_admin'
    elif is_agent:
        dashboard_url = 'dashboard_agent'
    else:
        dashboard_url = 'dashboard_assistant'
    
    context = {
        'title': f'{operateur.capitalize()} - {type_transaction.capitalize()}',
        'form': form,
        'type_transaction': type_transaction,
        'operateur': operateur,
        'is_admin': is_admin,
        'is_agent': is_agent,
        'is_assistant': is_assistant,
        'caisse': caisse,
        'dashboard_url': dashboard_url,
    }
    return render(request, 'transactions/transaction_form.html', context)
# ==================== DEMANDES D'APPROVISIONNEMENT ====================

@login_required
@require_http_methods(["POST"])
def demander_approvisionnement_api(request):
    """
    API pour les demandes d'approvisionnement (AJAX seulement)
    L'agent peut choisir entre Admin ou Assistant comme destinataire
    """
    try:
        agent = Agent.objects.get(user=request.user)
        caisse = agent.user.caisse
    except Agent.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Vous n\'êtes pas configuré comme agent.'})
    
    type_echange = request.POST.get('type_echange')
    montant = request.POST.get('montant')
    destinataire_type = request.POST.get('destinataire_type', 'admin')  # 'admin' ou 'assistant'
    assistant_id = request.POST.get('assistant_id')
    
    if not type_echange or not montant:
        return JsonResponse({'success': False, 'error': 'Veuillez remplir tous les champs'})
    
    if not destinataire_type:
        return JsonResponse({'success': False, 'error': 'Veuillez choisir un destinataire (Admin ou Assistant)'})
    
    # Si destinataire est assistant, vérifier que l'assistant_id est fourni
    if destinataire_type == 'assistant' and not assistant_id:
        return JsonResponse({'success': False, 'error': 'Veuillez sélectionner un assistant'})
    
    try:
        montant = Decimal(montant)
    except:
        return JsonResponse({'success': False, 'error': 'Montant invalide'})
    
    if montant < 1000:
        return JsonResponse({'success': False, 'error': 'Le montant minimum est de 1000 FCFA'})
    
    # Si destinataire est assistant, vérifier que l'assistant existe et est actif
    assistant_destinataire = None
    if destinataire_type == 'assistant':
        try:
            assistant_destinataire = Assistant.objects.get(id=assistant_id, est_actif=True)
        except Assistant.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Assistant non trouvé ou inactif'})
    
    # Vérifier le solde selon le type d'échange
    if type_echange == 'uv_to_cash':
        if montant > caisse.solde_uv:
            return JsonResponse({'success': False, 'error': f"Solde UV insuffisant. Solde actuel: {caisse.solde_uv:,.0f} FCFA"})
    elif type_echange == 'wave_to_cash':
        if montant > caisse.solde_wave:
            return JsonResponse({'success': False, 'error': f"Solde Wave insuffisant. Solde actuel: {caisse.solde_wave:,.0f} FCFA"})
    elif type_echange == 'cash_to_uv':
        if montant > caisse.solde_cash:
            return JsonResponse({'success': False, 'error': f"Solde Cash insuffisant. Solde actuel: {caisse.solde_cash:,.0f} FCFA"})
    elif type_echange == 'cash_to_wave':
        if montant > caisse.solde_cash:
            return JsonResponse({'success': False, 'error': f"Solde Cash insuffisant. Solde actuel: {caisse.solde_cash:,.0f} FCFA"})
    else:
        return JsonResponse({'success': False, 'error': 'Type d\'échange invalide'})
    
    # Créer la demande avec le destinataire choisi (sans motif)
    try:
        demande = DemandeApprovisionnement.objects.create(
            agent=agent,
            type_echange=type_echange,
            montant=montant,
            motif='',  # Motif vide
            destinataire_type=destinataire_type,
            assistant_destinataire=assistant_destinataire,
            statut='en_attente'
        )
        
        # Message personnalisé selon le destinataire
        if destinataire_type == 'admin':
            message = f'Demande envoyée à l\'Administrateur'
        else:
            message = f'Demande envoyée à l\'Assistant {assistant_destinataire.nom}'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'demande_id': demande.id
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def demander_approvisionnement(request):
    """
    L'AGENT fait une demande d'approvisionnement
    Peut choisir entre ADMIN ou ASSISTANT comme destinataire
    Types d'échanges possibles:
    - uv_to_cash: Échanger UV contre Cash
    - wave_to_cash: Échanger Wave contre Cash
    - cash_to_uv: Échanger Cash contre UV
    - cash_to_wave: Échanger Cash contre Wave
    """
    try:
        agent = Agent.objects.get(user=request.user)
        caisse = agent.user.caisse
    except Agent.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas configuré comme agent.')
        return redirect('login')
    
    # Récupérer la liste des assistants disponibles
    assistants = Assistant.objects.filter(est_actif=True)
    
    # Détecter si c'est une requête AJAX
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    # Pré-sélection du type depuis l'URL
    type_preset = request.GET.get('type', '')
    
    if request.method == 'POST':
        type_echange = request.POST.get('type_echange')
        montant = request.POST.get('montant')
        destinataire_type = request.POST.get('destinataire_type', 'admin')
        assistant_id = request.POST.get('assistant_id')
        
        if not type_echange or not montant or not destinataire_type:
            error_msg = 'Veuillez remplir tous les champs correctement.'
            if is_ajax:
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('demander_approvisionnement')
        
        # Si destinataire est assistant, vérifier l'assistant_id
        if destinataire_type == 'assistant' and not assistant_id:
            error_msg = 'Veuillez sélectionner un assistant.'
            if is_ajax:
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('demander_approvisionnement')
        
        try:
            montant = Decimal(montant)
        except:
            error_msg = 'Montant invalide'
            if is_ajax:
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('demander_approvisionnement')
        
        if montant < 1000:
            error_msg = 'Le montant minimum est de 1 000 FCFA'
            if is_ajax:
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('demander_approvisionnement')
        
        # Vérifier le solde selon le type d'échange
        solde_ok = True
        message_erreur = ""
        
        if type_echange == 'uv_to_cash':
            if montant > caisse.solde_uv:
                solde_ok = False
                message_erreur = f"Solde UV insuffisant. Solde actuel: {caisse.solde_uv:,.0f} FCFA"
        elif type_echange == 'wave_to_cash':
            if montant > caisse.solde_wave:
                solde_ok = False
                message_erreur = f"Solde Wave insuffisant. Solde actuel: {caisse.solde_wave:,.0f} FCFA"
        elif type_echange == 'cash_to_uv':
            if montant > caisse.solde_cash:
                solde_ok = False
                message_erreur = f"Solde Cash insuffisant. Solde actuel: {caisse.solde_cash:,.0f} FCFA"
        elif type_echange == 'cash_to_wave':
            if montant > caisse.solde_cash:
                solde_ok = False
                message_erreur = f"Solde Cash insuffisant. Solde actuel: {caisse.solde_cash:,.0f} FCFA"
        else:
            solde_ok = False
            message_erreur = "Type d'échange invalide"
        
        if not solde_ok:
            if is_ajax:
                return JsonResponse({'success': False, 'error': message_erreur})
            messages.error(request, message_erreur)
            return redirect('demander_approvisionnement')
        
        # Récupérer l'assistant si nécessaire
        assistant_destinataire = None
        if destinataire_type == 'assistant' and assistant_id:
            try:
                assistant_destinataire = Assistant.objects.get(id=assistant_id, est_actif=True)
            except Assistant.DoesNotExist:
                error_msg = 'Assistant non trouvé.'
                if is_ajax:
                    return JsonResponse({'success': False, 'error': error_msg})
                messages.error(request, error_msg)
                return redirect('demander_approvisionnement')
        
        # Créer la demande (sans motif)
        try:
            demande = DemandeApprovisionnement.objects.create(
                agent=agent,
                type_echange=type_echange,
                montant=montant,
                motif='',  # Motif vide
                destinataire_type=destinataire_type,
                assistant_destinataire=assistant_destinataire,
                statut='en_attente'
            )
            
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'message': 'Demande envoyée avec succès',
                    'demande_id': demande.id
                })
            
            if destinataire_type == 'admin':
                messages.success(request, f'✅ Demande envoyée à l\'Administrateur! {montant:,.0f} FCFA')
            else:
                messages.success(request, f'✅ Demande envoyée à l\'Assistant {assistant_destinataire.nom}! {montant:,.0f} FCFA')
            return redirect('dashboard_agent')
            
        except Exception as e:
            if is_ajax:
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(request, f'Erreur: {str(e)}')
            return redirect('demander_approvisionnement')
    
    context = {
        'title': 'Demander un approvisionnement',
        'agent': agent,
        'caisse': caisse,
        'type_preset': type_preset,
        'assistants': assistants,
    }
    return render(request, 'transactions/demande_approvisionnement.html', context)


@login_required
def valider_demande(request, demande_id):
    """
    L'ADMIN valide ou refuse une demande d'approvisionnement
    """
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    demande = django.shortcuts.get_object_or_404(DemandeApprovisionnement, id=demande_id)
    
    print(f"=== VALIDATION DEMANDE ===")
    print(f"Demande ID: {demande_id}")
    print(f"Statut actuel: {demande.statut}")
    
    # Vérifier que la demande est en attente
    if demande.statut != 'en_attente':
        messages.error(request, f'Cette demande a déjà été {demande.get_statut_display().lower()}.')
        return redirect('dashboard_admin')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        print(f"Action: {action}")
        
        if action == 'valider':
            print("Tentative de validation...")
            if demande.valider_par_admin(admin):
                messages.success(request, f'✅ Demande validée! {demande.montant:,.0f} FCFA échangés.')
                print(f"Validation réussie - Nouveau statut: {demande.statut}")
            else:
                messages.error(request, '❌ Solde insuffisant pour valider cette demande.')
                print("Échec de la validation - Solde insuffisant")
        
        elif action == 'refuser':
            demande.statut = 'refuse'
            demande.traite_par_admin = admin
            demande.date_traitement = datetime.now()
            demande.save()
            messages.info(request, 'Demande refusée.')
            print(f"Demande refusée - Nouveau statut: {demande.statut}")
        
        return redirect('dashboard_admin')
    
    context = {
        'title': 'Valider une demande',
        'demande': demande,
    }
    return render(request, 'transactions/valider_demande.html', context)

# views.py
# views.py
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.template.loader import render_to_string
from django.utils import timezone

@login_required
def impression_recu(request, transaction_id):
    """
    Vue pour l'impression des reçus
    Envoie uniquement les données brutes, le JS fait le formatage
    """
    transaction = get_object_or_404(Transaction, reference=transaction_id)
    
    # Déterminer le rôle de l'utilisateur
    user_role = "Admin"
    if hasattr(transaction.user, 'agent'):
        user_role = "Agent"
    elif hasattr(transaction.user, 'assistant'):
        user_role = "Assistant"
    elif transaction.user.is_superuser or transaction.user.is_staff:
        user_role = "Admin"
    
    # Données brutes (sans formatage)
    data = {
        "operateur": transaction.user.username if transaction.user else "-",
        "role_operateur": user_role,
        "type": transaction.type_transaction if transaction.type_transaction else "-",
        "operateur_money": transaction.get_operateur_display() if transaction.operateur else "-",
        "client": transaction.numero_client if transaction.numero_client else "-",
        "nom_client": transaction.nom_client if transaction.nom_client else "-",
        "montant": int(transaction.montant) if transaction.montant else 0,
        "reference": transaction.reference if transaction.reference else "-",
        "date": transaction.date.strftime('%d/%m/%Y') if transaction.date else "-",
        "heure": transaction.date.strftime('%H:%M:%S') if transaction.date else "-",
    }
    
    # Retourner les données en JSON pour QZ Tray
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        return JsonResponse(data)
    
    # Pour l'impression navigateur (fallback)
    context = {
        'transaction': transaction,
        'data': data,
        'user_role': user_role,
        'entreprise': {
            'nom': 'KONE SERVICES',
            'telephone': '76 89 77 31',
            'adresse': 'Services Transfert'
        }
    }
    return render(request, 'transactions/recu.html', context)
# ==================== HISTORIQUES ====================

@login_required
def historique_admin(request):
    """
    Historique des transactions pour l'ADMIN (toutes)
    Supporte l'AJAX pour le chargement dynamique
    """
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    # ========== FILTRES ==========
    type_filtre = request.GET.get('type')
    operateur_filtre = request.GET.get('operateur')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    user_id = request.GET.get('user_id')
    
    transactions = Transaction.objects.all()
    
    # Filtre par utilisateur (par ID)
    if user_id:
        transactions = transactions.filter(user_id=user_id)
    
    if type_filtre:
        transactions = transactions.filter(type_transaction=type_filtre)
    if operateur_filtre:
        transactions = transactions.filter(operateur=operateur_filtre)
    if date_debut:
        try:
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            transactions = transactions.filter(date__date__gte=date_debut_obj)
        except ValueError:
            pass
    if date_fin:
        try:
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
            transactions = transactions.filter(date__date__lte=date_fin_obj)
        except ValueError:
            pass
    
    # ========== TOTAUX (avant pagination) ==========
    total_entree = transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    total_sortie = transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    total_commission = transactions.aggregate(Sum('commission'))['commission__sum'] or 0
    
    # ========== PAGINATION ==========
    page = request.GET.get('page', 1)
    paginator = Paginator(transactions.order_by('-date'), 10)
    
    try:
        transactions_page = paginator.page(page)
    except:
        transactions_page = paginator.page(1)
    
    # ========== DÉTECTION AJAX ==========
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        transactions_data = []
        for t in transactions_page:
            if hasattr(t.user, 'admin_profile'):
                user_type = 'admin'
                user_name = t.user.admin_profile.nom
            elif hasattr(t.user, 'agent_profile'):
                user_type = 'agent'
                user_name = t.user.agent_profile.nom
            elif hasattr(t.user, 'assistant_profile'):
                user_type = 'assistant'
                user_name = t.user.assistant_profile.nom
            else:
                user_type = 'unknown'
                user_name = t.user.username
            
            transactions_data.append({
                'reference': t.reference,
                'user_type': user_type,
                'user_name': user_name,
                'type': t.type_transaction,
                'operateur': t.operateur,
                'operateur_label': t.get_operateur_display(),
                'numero_client': t.numero_client,
                'montant': float(t.montant),
                'commission': float(t.commission),
                'date': t.date.strftime('%d/%m/%Y %H:%M'),
                'est_annule': t.est_annule,
            })
        
        current = transactions_page.number
        total_pages = paginator.num_pages
        start_page = max(1, current - 2)
        end_page = min(total_pages, current + 2)
        
        return JsonResponse({
            'success': True,
            'transactions': transactions_data,
            'stats': {
                'count': transactions.count(),
                'total_entree': float(total_entree),
                'total_sortie': float(total_sortie),
                'total_commission': float(total_commission),
            },
            'pagination': {
                'current_page': current,
                'total_pages': total_pages,
                'has_next': transactions_page.has_next(),
                'has_previous': transactions_page.has_previous(),
                'next_page': transactions_page.next_page_number() if transactions_page.has_next() else None,
                'previous_page': transactions_page.previous_page_number() if transactions_page.has_previous() else None,
                'start_page': start_page,
                'end_page': end_page,
            }
        })
    
    # ========== RENDU NORMAL ==========
    admins = Admin.objects.all()
    agents = Agent.objects.filter(est_actif=True)
    assistants = Assistant.objects.filter(est_actif=True)
    
    demandes_attente = DemandeApprovisionnement.objects.filter(statut='en_attente')
    context = {
        'title': 'Historique des transactions',
        'transactions': transactions_page,
        'total_entree': total_entree,
        'total_sortie': total_sortie,
        'total_commission': total_commission,
        'admins': admins,
        'agents': agents,
        'assistants': assistants,
        'demandes_attente': demandes_attente,
    }
    return render(request, 'transactions/historique_admin.html', context)

@login_required
@login_required
def historique_agent(request):
    """
    Historique des transactions pour l'AGENT ou ASSISTANT (ses propres transactions)
    Avec filtres par date, opérateur et type
    Affiche par défaut les transactions du jour
    """
    # Vérifier si c'est un agent ou un assistant
    try:
        agent = Agent.objects.get(user=request.user)
        is_agent = True
    except Agent.DoesNotExist:
        try:
            assistant = Assistant.objects.get(user=request.user)
            is_agent = False
            agent = assistant  # Pour garder la compatibilité avec le template
        except Assistant.DoesNotExist:
            messages.error(request, 'Vous n\'êtes pas autorisé.')
            return redirect('login')
    
    # Récupérer toutes les transactions de l'utilisateur
    transactions = Transaction.objects.filter(user=request.user).order_by('-date')
    
    # Date d'aujourd'hui avec timezone
    today = timezone.now().date()
    
    # ========== FILTRES ==========
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    operateur = request.GET.get('operateur')
    type_transaction = request.GET.get('type')
    
    # Si aucun filtre de date n'est appliqué, afficher uniquement les transactions du jour
    if not date_debut and not date_fin:
        transactions = transactions.filter(date__date=today)
        date_debut_display = today.strftime('%Y-%m-%d')
        date_fin_display = today.strftime('%Y-%m-%d')
    else:
        date_debut_display = date_debut
        date_fin_display = date_fin
        
        # Filtre par date début
        if date_debut:
            try:
                date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
                transactions = transactions.filter(date__date__gte=date_debut_obj)
            except ValueError:
                pass
        
        # Filtre par date fin
        if date_fin:
            try:
                date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
                transactions = transactions.filter(date__date__lte=date_fin_obj)
            except ValueError:
                pass
    
    # Filtre par opérateur
    if operateur:
        transactions = transactions.filter(operateur=operateur)
    
    # Filtre par type
    if type_transaction:
        transactions = transactions.filter(type_transaction=type_transaction)
    
    # ========== CALCUL DES TOTAUX (APRES FILTRES) ==========
    total_entree = transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    total_sortie = transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    total_commission = transactions.aggregate(Sum('commission'))['commission__sum'] or 0
    
    # ========== PAGINATION ==========
    page = request.GET.get('page', 1)
    paginator = Paginator(transactions, 10)
    
    try:
        transactions_page = paginator.page(page)
    except PageNotAnInteger:
        transactions_page = paginator.page(1)
    except EmptyPage:
        transactions_page = paginator.page(paginator.num_pages)
    
    context = {
        'title': 'Mes transactions',
        'agent': agent,  # Garde le nom agent pour la compatibilité avec le template
        'transactions': transactions_page,
        'total_entree': total_entree,
        'total_sortie': total_sortie,
        'total_commission': total_commission,
        'date_debut': date_debut_display,
        'date_fin': date_fin_display,
    }
    return render(request, 'transactions/historique_agent.html', context)

@login_required
def historique_demandes_agent(request):
    """
    Historique des demandes d'approvisionnement ET des opérations de caisse
    - Pour un AGENT: ses demandes envoyées + ses opérations de caisse
    - Pour un ASSISTANT: les demandes qu'il a reçues + les opérations de sa caisse
    Avec filtres par date, statut et type d'échange
    """
    from decimal import Decimal
    
    # Vérifier si c'est un agent ou un assistant
    try:
        agent = Agent.objects.get(user=request.user)
        type_utilisateur = 'agent'
        # Demandes de l'agent
        demandes = DemandeApprovisionnement.objects.filter(agent=agent).order_by('-date_demande')
        # Opérations de caisse de l'agent
        operations = OperationCaisse.objects.filter(caisse__user=request.user).order_by('-date_operation')

    except Agent.DoesNotExist:
        try:
            assistant = Assistant.objects.get(user=request.user)
            type_utilisateur = 'assistant'
            # Demandes reçues par l'assistant
            demandes = DemandeApprovisionnement.objects.filter(
                destinataire_type='assistant',
                assistant_destinataire=assistant
            ).order_by('-date_demande')
            # Opérations de caisse de l'assistant (même caisse que l'admin)
            operations = OperationCaisse.objects.filter(caisse__user=request.user).order_by('-date_operation')
        except Assistant.DoesNotExist:
            messages.error(request, 'Vous n\'êtes pas autorisé.')
            return redirect('login')
    
    today = timezone.now().date()
    
    # ========== FILTRES ==========
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    statut = request.GET.get('statut')
    type_echange = request.GET.get('type_echange')
    type_operation_filtre = request.GET.get('type_operation')
    
    # Appliquer les filtres de date sur les demandes et opérations
    if date_debut:
        try:
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            demandes = demandes.filter(date_demande__date__gte=date_debut_obj)
            operations = operations.filter(date_operation__date__gte=date_debut_obj)
        except ValueError:
            pass
    
    if date_fin:
        try:
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
            demandes = demandes.filter(date_demande__date__lte=date_fin_obj)
            operations = operations.filter(date_operation__date__lte=date_fin_obj)
        except ValueError:
            pass
    
    # Si aucun filtre de date n'est appliqué, afficher les demandes et opérations du jour
    if not date_debut and not date_fin:
        demandes = demandes.filter(date_demande__date=today)
        operations = operations.filter(date_operation__date=today)
        date_debut_display = today.strftime('%Y-%m-%d')
        date_fin_display = today.strftime('%Y-%m-%d')
    else:
        date_debut_display = date_debut or today.strftime('%Y-%m-%d')
        date_fin_display = date_fin or today.strftime('%Y-%m-%d')
    
    # Filtre par statut sur les demandes
    if statut:
        demandes = demandes.filter(statut=statut)
    
    # Filtre par type d'échange sur les demandes
    if type_echange:
        demandes = demandes.filter(type_echange=type_echange)
    
    # Filtre par type d'opération sur les opérations de caisse
    if type_operation_filtre:
        operations = operations.filter(type_operation=type_operation_filtre)
    
    # ========== STATISTIQUES ==========
    if type_utilisateur == 'agent':
        stats = {
            'demandes_attente': DemandeApprovisionnement.objects.filter(agent=agent, statut='en_attente').count(),
            'demandes_valide': DemandeApprovisionnement.objects.filter(agent=agent, statut='valide').count(),
            'demandes_refuse': DemandeApprovisionnement.objects.filter(agent=agent, statut='refuse').count(),
            'operations_encaissement': OperationCaisse.objects.filter(caisse__user=request.user, type_operation='encaissement').count(),
            'operations_decaissement': OperationCaisse.objects.filter(caisse__user=request.user, type_operation='decaissement').count(),
        }
    else:
        stats = {
            'demandes_attente': DemandeApprovisionnement.objects.filter(
                destinataire_type='assistant',
                assistant_destinataire=assistant,
                statut='en_attente'
            ).count(),
            'demandes_valide': DemandeApprovisionnement.objects.filter(
                destinataire_type='assistant',
                assistant_destinataire=assistant,
                statut='valide'
            ).count(),
            'demandes_refuse': DemandeApprovisionnement.objects.filter(
                destinataire_type='assistant',
                assistant_destinataire=assistant,
                statut='refuse'
            ).count(),
            'operations_encaissement': OperationCaisse.objects.filter(caisse__user=request.user, type_operation='encaissement').count(),
            'operations_decaissement': OperationCaisse.objects.filter(caisse__user=request.user, type_operation='decaissement').count(),
        }
    
    # Montants totaux des demandes par statut
    stats['montant_attente'] = demandes.filter(statut='en_attente').aggregate(Sum('montant'))['montant__sum'] or 0
    stats['montant_valide'] = demandes.filter(statut='valide').aggregate(Sum('montant'))['montant__sum'] or 0
    
    # Montants totaux des opérations de caisse
    stats['total_encaissements'] = operations.filter(type_operation='encaissement').aggregate(Sum('montant'))['montant__sum'] or 0
    stats['total_decaissements'] = operations.filter(type_operation='decaissement').aggregate(Sum('montant'))['montant__sum'] or 0
    
    # ========== PAGINATION ==========
    page = request.GET.get('page', 1)
    paginator = Paginator(demandes, 10)
    
    try:
        demandes_page = paginator.page(page)
    except PageNotAnInteger:
        demandes_page = paginator.page(1)
    except EmptyPage:
        demandes_page = paginator.page(paginator.num_pages)
    
    # Pagination pour les opérations
    page_ops = request.GET.get('page_ops', 1)
    paginator_ops = Paginator(operations, 10)
    
    try:
        operations_page = paginator_ops.page(page_ops)
    except PageNotAnInteger:
        operations_page = paginator_ops.page(1)
    except EmptyPage:
        operations_page = paginator_ops.page(paginator_ops.num_pages)
    
    context = {
        'title': 'Mon historique',
        'type_utilisateur': type_utilisateur,
        'demandes': demandes_page,
        'operations': operations_page,
        'stats': stats,
        'date_debut': date_debut_display,
        'date_fin': date_fin_display,
        'statut_filtre': statut,
        'type_echange_filtre': type_echange,
        'type_operation_filtre': type_operation_filtre,
    }
    return render(request, 'transactions/historique_demandes.html', context)
    
@login_required
def traiter_demande_assistant(request, demande_id):
    """
    L'ASSISTANT traite (valide ou refuse) une demande d'approvisionnement
    L'assistant utilise la caisse de son ADMIN
    """
    try:
        assistant = Assistant.objects.get(user=request.user)
    except Assistant.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Vous n\'êtes pas autorisé.'})
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    demande = get_object_or_404(
        DemandeApprovisionnement, 
        id=demande_id,
        destinataire_type='assistant',
        assistant_destinataire=assistant,
        statut='en_attente'
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if action == 'valider':
            # Récupérer la caisse de l'ADMIN
            caisse_admin = assistant.admin.user.caisse
            caisse_agent = demande.agent.user.caisse
            
            # Vérifier les soldes selon le type d'échange
            solde_ok = True
            message_erreur = ""
            
            print(f"=== DEBUG VALIDATION ===")
            print(f"Type échange: {demande.type_echange}")
            print(f"Montant: {demande.montant}")
            print(f"Solde Admin Cash: {caisse_admin.solde_cash}")
            print(f"Solde Admin UV: {caisse_admin.solde_uv}")
            print(f"Solde Admin Wave: {caisse_admin.solde_wave}")
            print(f"Solde Agent Cash: {caisse_agent.solde_cash}")
            print(f"Solde Agent UV: {caisse_agent.solde_uv}")
            print(f"Solde Agent Wave: {caisse_agent.solde_wave}")
            
            if demande.type_echange == 'uv_to_cash':
                if caisse_agent.solde_uv < demande.montant:
                    solde_ok = False
                    message_erreur = f"Solde UV de l'agent insuffisant. Solde actuel: {caisse_agent.solde_uv:,.0f} FCFA"
                elif caisse_admin.solde_cash < demande.montant:
                    solde_ok = False
                    message_erreur = f"Solde Cash de l'administrateur insuffisant. Solde actuel: {caisse_admin.solde_cash:,.0f} FCFA"
                    
            elif demande.type_echange == 'wave_to_cash':
                if caisse_agent.solde_wave < demande.montant:
                    solde_ok = False
                    message_erreur = f"Solde Wave de l'agent insuffisant. Solde actuel: {caisse_agent.solde_wave:,.0f} FCFA"
                elif caisse_admin.solde_cash < demande.montant:
                    solde_ok = False
                    message_erreur = f"Solde Cash de l'administrateur insuffisant. Solde actuel: {caisse_admin.solde_cash:,.0f} FCFA"
                    
            elif demande.type_echange == 'cash_to_uv':
                if caisse_agent.solde_cash < demande.montant:
                    solde_ok = False
                    message_erreur = f"Solde Cash de l'agent insuffisant. Solde actuel: {caisse_agent.solde_cash:,.0f} FCFA"
                elif caisse_admin.solde_uv < demande.montant:
                    solde_ok = False
                    message_erreur = f"Solde UV de l'administrateur insuffisant. Solde actuel: {caisse_admin.solde_uv:,.0f} FCFA"
                    
            elif demande.type_echange == 'cash_to_wave':
                if caisse_agent.solde_cash < demande.montant:
                    solde_ok = False
                    message_erreur = f"Solde Cash de l'agent insuffisant. Solde actuel: {caisse_agent.solde_cash:,.0f} FCFA"
                elif caisse_admin.solde_wave < demande.montant:
                    solde_ok = False
                    message_erreur = f"Solde Wave de l'administrateur insuffisant. Solde actuel: {caisse_admin.solde_wave:,.0f} FCFA"
            
            if not solde_ok:
                if is_ajax:
                    return JsonResponse({'success': False, 'error': message_erreur})
                messages.error(request, message_erreur)
                return redirect('dashboard_assistant')
            
            # Effectuer la transaction manuellement
            try:
                # Mise à jour des soldes
                if demande.type_echange == 'uv_to_cash':
                    caisse_agent.solde_uv -= demande.montant
                    caisse_agent.solde_cash += demande.montant
                    caisse_admin.solde_cash -= demande.montant
                    caisse_admin.solde_uv += demande.montant
                    
                elif demande.type_echange == 'wave_to_cash':
                    caisse_agent.solde_wave -= demande.montant
                    caisse_agent.solde_cash += demande.montant
                    caisse_admin.solde_cash -= demande.montant
                    caisse_admin.solde_wave += demande.montant
                    
                elif demande.type_echange == 'cash_to_uv':
                    caisse_agent.solde_cash -= demande.montant
                    caisse_agent.solde_uv += demande.montant
                    caisse_admin.solde_uv -= demande.montant
                    caisse_admin.solde_cash += demande.montant
                    
                elif demande.type_echange == 'cash_to_wave':
                    caisse_agent.solde_cash -= demande.montant
                    caisse_agent.solde_wave += demande.montant
                    caisse_admin.solde_wave -= demande.montant
                    caisse_admin.solde_cash += demande.montant
                
                # Sauvegarde
                caisse_agent.save()
                caisse_admin.save()
                
                # Mise à jour de la demande
                demande.statut = 'valide'
                demande.traite_par_assistant = assistant
                demande.date_traitement = timezone.now()
                demande.save()
                
                if is_ajax:
                    return JsonResponse({'success': True, 'message': 'Demande validée avec succès'})
                messages.success(request, f'✅ Demande validée ! {demande.montant:,.0f} FCFA échangés.')
                
            except Exception as e:
                if is_ajax:
                    return JsonResponse({'success': False, 'error': str(e)})
                messages.error(request, f'Erreur: {str(e)}')
        
        elif action == 'refuser':
            demande.statut = 'refuse'
            demande.traite_par_assistant = assistant
            demande.date_traitement = timezone.now()
            demande.save()
            if is_ajax:
                return JsonResponse({'success': True, 'message': 'Demande refusée'})
            messages.info(request, 'Demande refusée.')
        
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Action invalide'})
        return redirect('dashboard_assistant')
    
    return redirect('dashboard_assistant')
# ==================== GESTION DES AGENTS ====================

@login_required
def gestion_agents(request):
    """
    Page de gestion des agents ET assistants (vue unifiée)
    """
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    # ========== AGENTS ==========
    agents = Agent.objects.all().order_by('-created_at')
    agents_actifs = agents.filter(est_actif=True).count()
    agents_inactifs = agents.filter(est_actif=False).count()
    
    # ========== ASSISTANTS ==========
    assistants = Assistant.objects.filter(admin=admin).order_by('-created_at')
    assistants_actifs = assistants.filter(est_actif=True).count()
    assistants_inactifs = assistants.filter(est_actif=False).count()
    
    demandes_attente = DemandeApprovisionnement.objects.filter(statut='en_attente')
    context = {
        'title': 'Gestion des utilisateurs',
        'agents': agents,
        'agents_actifs': agents_actifs,
        'agents_inactifs': agents_inactifs,
        'assistants': assistants,
        'assistants_actifs': assistants_actifs,
        'assistants_inactifs': assistants_inactifs,
        'demandes_attente': demandes_attente,
    }
    return render(request, 'transactions/gestion_agents.html', context)

@login_required
def ajouter_agent(request):
    """
    Ajouter ou modifier un agent
    """
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    if request.method == 'POST':
        agent_id = request.POST.get('agent_id')
        nom = request.POST.get('nom')
        telephone = request.POST.get('telephone')
        email = request.POST.get('email')
        est_actif = request.POST.get('est_actif') == 'true'
        password = request.POST.get('password')
        
        if not nom or not telephone:
            messages.error(request, 'Le nom et le téléphone sont obligatoires.')
            return redirect('gestion_agents')
        
        if agent_id:
            # Modification d'un agent existant
            try:
                agent = Agent.objects.get(id=agent_id)
                agent.nom = nom
                agent.telephone = telephone
                agent.email = email
                agent.est_actif = est_actif
                agent.save()
                
                # Mettre à jour l'utilisateur associé
                user = agent.user
                user.email = email
                if password:
                    user.password = make_password(password)
                user.save()
                
                messages.success(request, f'✅ Agent "{nom}" modifié avec succès.')
            except Agent.DoesNotExist:
                messages.error(request, 'Agent non trouvé.')
        else:
            # Création d'un nouvel agent
            if not password:
                messages.error(request, 'Le mot de passe est obligatoire pour un nouvel agent.')
                return redirect('gestion_agents')
            
            # Créer l'utilisateur
            username = telephone
            base_username = username
            counter = 1
            while User.objects.filter(username=username).exists():
                username = f"{base_username}_{counter}"
                counter += 1
            
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password
            )
            
            # Créer l'agent
            agent = Agent.objects.create(
                user=user,
                nom=nom,
                telephone=telephone,
                email=email,
                est_actif=est_actif
            )
            
            messages.success(request, f'✅ Agent "{nom}" créé avec succès. Identifiant: {username}')
        
        return redirect('gestion_agents')
    
    return redirect('gestion_agents')


from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib import messages
from django.db.models import Sum
from transactions.models import Agent, Assistant, Admin, Caisse, HistoriqueAgent


@login_required
def operation_agent(request):
    """
    Opération sur agent : Admin ou Assistant peut ajouter ou retirer des fonds
    Admin et Assistant partagent le même solde (caisse de l'admin)
    """
    if request.method != 'POST':
        return redirect('gestion_agents')
    
    agent_id = request.POST.get('agent_id')
    operation_type = request.POST.get('operation_type')  # 'ajout' ou 'retrait'
    type_fonds = request.POST.get('type_fonds')  # 'cash', 'uv', 'wave'
    montant = request.POST.get('montant')
    description = request.POST.get('description', '')
    
    if not agent_id:
        messages.error(request, "Veuillez sélectionner un agent")
        return redirect('gestion_agents')
    
    try:
        montant = Decimal(str(montant))
        if montant <= 0:
            messages.error(request, "Le montant doit être supérieur à 0")
            return redirect('gestion_agents')
        
        if montant < 100:
            messages.error(request, "Le montant minimum est de 100 FCFA")
            return redirect('gestion_agents')
        
        # Récupérer l'agent
        agent = Agent.objects.get(id=agent_id)
        
        # Déterminer qui est l'opérateur et la caisse partagée
        try:
            admin_obj = Admin.objects.get(user=request.user)
            operateur_type = 'admin'
            operateur_nom = admin_obj.nom
            caisse_partagee = admin_obj.user.caisse
        except Admin.DoesNotExist:
            assistant = Assistant.objects.get(user=request.user)
            operateur_type = 'assistant'
            operateur_nom = assistant.nom
            caisse_partagee = assistant.admin.user.caisse
        
        # Déterminer le champ et les soldes
        if type_fonds == 'cash':
            champ = 'solde_cash'
            solde_partage = getattr(caisse_partagee, champ, 0)
            solde_agent = getattr(agent.user.caisse, champ, 0)
            type_label = "Cash"
        elif type_fonds == 'uv':
            champ = 'solde_uv'
            solde_partage = getattr(caisse_partagee, champ, 0)
            solde_agent = getattr(agent.user.caisse, champ, 0)
            type_label = "UV Touchpoint"
        else:
            champ = 'solde_wave'
            solde_partage = getattr(caisse_partagee, champ, 0)
            solde_agent = getattr(agent.user.caisse, champ, 0)
            type_label = "UV Wave"
        
        if operation_type == 'ajout':
            # ➕ AJOUT : Donner à l'agent
            if solde_partage < montant:
                messages.error(request, f"Solde {type_label} insuffisant. Disponible: {solde_partage:,.0f} FCFA")
                return redirect('gestion_agents')
            
            # Débiter la caisse partagée
            setattr(caisse_partagee, champ, solde_partage - montant)
            caisse_partagee.save()
            
            # Créditer l'agent
            setattr(agent.user.caisse, champ, solde_agent + montant)
            agent.user.caisse.save()
            
            # Enregistrer l'historique
            HistoriqueAgent.objects.create(
                agent=agent,
                type_operation='decaissement',
                operateur_type=operateur_type,
                operateur_nom=operateur_nom,
                montant_cash=montant if type_fonds == 'cash' else 0,
                montant_uv=montant if type_fonds == 'uv' else 0,
                montant_wave=montant if type_fonds == 'wave' else 0,
                description=f"Ajout de {montant:,.0f} FCFA en {type_label} - {description}" if description else f"Ajout de {montant:,.0f} FCFA en {type_label}"
            )
            
            messages.success(request, f"✅ {montant:,.0f} FCFA ajoutés à {agent.nom} en {type_label}")
            
        elif operation_type == 'retrait':
            # ➖ RETRAIT : Reprendre à l'agent
            if solde_agent < montant:
                messages.error(request, f"Solde {type_label} de {agent.nom} insuffisant. Disponible: {solde_agent:,.0f} FCFA")
                return redirect('gestion_agents')
            
            # Débiter l'agent
            setattr(agent.user.caisse, champ, solde_agent - montant)
            agent.user.caisse.save()
            
            # Créditer la caisse partagée
            setattr(caisse_partagee, champ, solde_partage + montant)
            caisse_partagee.save()
            
            # Enregistrer l'historique
            HistoriqueAgent.objects.create(
                agent=agent,
                type_operation='encaissement',
                operateur_type=operateur_type,
                operateur_nom=operateur_nom,
                montant_cash=montant if type_fonds == 'cash' else 0,
                montant_uv=montant if type_fonds == 'uv' else 0,
                montant_wave=montant if type_fonds == 'wave' else 0,
                description=f"Retrait de {montant:,.0f} FCFA en {type_label} - {description}" if description else f"Retrait de {montant:,.0f} FCFA en {type_label}"
            )
            
            messages.success(request, f"✅ {montant:,.0f} FCFA retirés de {agent.nom} en {type_label}")
        
        else:
            messages.error(request, "Type d'opération invalide")
        
    except Agent.DoesNotExist:
        messages.error(request, "Agent introuvable")
    except Assistant.DoesNotExist:
        messages.error(request, "Assistant non trouvé")
    except Admin.DoesNotExist:
        messages.error(request, "Admin non trouvé")
    except Exception as e:
        messages.error(request, f"Erreur: {str(e)}")
    
    return redirect('gestion_agents')

from django.core.paginator import Paginator
from datetime import datetime
from django.db.models import Sum, Q
from transactions.models import HistoriqueAgent, Agent

@login_required
def historique_operations(request):
    """Page d'historique des opérations (Admin/Assistant → Agents)"""
    agents = Agent.objects.all()
    return render(request, 'transactions/historique_operations.html', {'agents': agents})


@login_required
def api_historique_agent_page(request):
    """API pour récupérer l'historique des opérations avec pagination et filtres"""
    
    # Récupérer les paramètres
    page = request.GET.get('page', 1)
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    agent_id = request.GET.get('agent_id')
    type_operation = request.GET.get('type')  # 'decaissement' ou 'encaissement'
    
    # Base des opérations
    historique = HistoriqueAgent.objects.all().order_by('-date_operation')
    
    # Appliquer les filtres
    if date_debut_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            historique = historique.filter(date_operation__date__gte=date_debut)
        except:
            pass
    
    if date_fin_str:
        try:
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
            historique = historique.filter(date_operation__date__lte=date_fin)
        except:
            pass
    
    if agent_id:
        try:
            historique = historique.filter(agent_id=int(agent_id))
        except:
            pass
    
    if type_operation:
        historique = historique.filter(type_operation=type_operation)
    
    # Calcul des stats (sans pagination)
    total_ajouts = 0
    total_retraits = 0
    
    for h in historique:
        montant_total = float(h.get_montant_total())
        if h.type_operation == 'decaissement':
            total_ajouts += montant_total
        else:
            total_retraits += montant_total
    
    # Pagination
    paginator = Paginator(historique, 20)
    page_obj = paginator.get_page(page)
    
    # Formater les données
    data = []
    for h in page_obj:
        isAjout = h.type_operation == 'decaissement'
        data.append({
            'id': h.id,
            'agent': h.agent.nom,
            'agent_id': h.agent.id,
            'type_operation': h.type_operation,
            'type_label': 'Ajout (Donné)' if isAjout else 'Retrait (Récupéré)',
            'operateur_type': h.operateur_type,
            'operateur_label': 'Admin' if h.operateur_type == 'admin' else 'Assistant',
            'operateur_nom': h.operateur_nom,
            'montant_total': float(h.get_montant_total()),
            'description': h.description or '',
            'date': h.date_operation.strftime('%d/%m/%Y %H:%M:%S')
        })
    
    # Pagination info
    start_page = max(1, page_obj.number - 2)
    end_page = min(paginator.num_pages, page_obj.number + 2)
    
    return JsonResponse({
        'success': True,
        'historique': data,
        'stats': {
            'total_ajouts': total_ajouts,
            'total_retraits': total_retraits,
            'total_operations': historique.count()
        },
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
            'start_page': start_page,
            'end_page': end_page,
        }
    })
    
@login_required
def api_historique_agent(request, agent_id=None):
    """API pour récupérer l'historique des opérations sur les agents"""
    historique = HistoriqueAgent.objects.all().order_by('-date_operation')[:200]
    
    if agent_id:
        historique = historique.filter(agent_id=agent_id)
    
    data = {
        'success': True,
        'historique': []
    }
    
    for h in historique:
        if h.type_operation == 'decaissement':
            type_label = '📤 Ajout (Admin/Assistant donne)'
        else:
            type_label = '📥 Retrait (Admin/Assistant reprend)'
        
        operateur_label = '👨‍💼 Admin' if h.operateur_type == 'admin' else '🤝 Assistant'
        
        data['historique'].append({
            'id': h.id,
            'agent': h.agent.nom,
            'agent_id': h.agent.id,
            'type_operation': h.type_operation,
            'type_label': type_label,
            'operateur_type': h.operateur_type,
            'operateur_label': operateur_label,
            'operateur_nom': h.operateur_nom,
            'montant_cash': float(h.montant_cash),
            'montant_uv': float(h.montant_uv),
            'montant_wave': float(h.montant_wave),
            'montant_total': float(h.get_montant_total()),
            'description': h.description or '',
            'date': h.date_operation.strftime('%d/%m/%Y %H:%M:%S')
        })
    
    return JsonResponse(data)

@login_required
def modifier_mot_de_passe_agent(request, user_id):
    if request.method == 'POST':
        nouveau_password = request.POST.get('nouveau_password')
        
        if not nouveau_password or len(nouveau_password) < 4:
            messages.error(request, "Le mot de passe doit contenir au moins 4 caractères")
            return redirect('gestion_agents')
        
        try:
            user = User.objects.get(id=user_id)
            agent = Agent.objects.get(user=user)
            
            # Modifier le mot de passe hashé
            user.password = make_password(nouveau_password)
            user.save()
            
            # Modifier le mot de passe en clair
            agent.mot_de_passe_clair = nouveau_password
            agent.save()
            
            messages.success(request, f"Mot de passe de {agent.nom} modifié avec succès")
            
        except User.DoesNotExist:
            messages.error(request, "Utilisateur introuvable")
        except Agent.DoesNotExist:
            messages.error(request, "Agent introuvable")
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
        
        return redirect('gestion_agents')
    
    return redirect('gestion_agents')


@login_required
def api_assistant_infos(request, assistant_id):
    """API pour récupérer les informations d'un assistant"""
    from transactions.models import Assistant
    
    try:
        assistant = Assistant.objects.get(id=assistant_id)
        return JsonResponse({
            'success': True,
            'username': assistant.user.username if assistant.user else ''
        })
    except Assistant.DoesNotExist:
        return JsonResponse({
            'success': False,
            'username': ''
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'username': '',
            'error': str(e)
        })

@login_required
def modifier_caisse(request):
    """Modifier la caisse d'un agent"""
    if request.method == 'POST':
        agent_id = request.POST.get('agent_id')  # ← c'est l'ID de l'agent
        nouveau_cash = request.POST.get('solde_cash')
        nouveau_uv = request.POST.get('solde_uv')
        nouveau_wave = request.POST.get('solde_wave')
        
        if not agent_id:
            messages.error(request, "Agent non spécifié")
            return redirect('gestion_agents')
        
        try:
            # Récupérer L'AGENT (pas l'admin)
            from transactions.models import Agent
            agent = Agent.objects.get(id=agent_id)
            user_agent = agent.user  # ← l'utilisateur agent
            
            # Récupérer la caisse de l'AGENT
            caisse = Caisse.objects.get(user=user_agent)
            
            # Sauvegarder les anciennes valeurs
            ancien_cash = caisse.solde_cash
            ancien_uv = caisse.solde_uv
            ancien_wave = caisse.solde_wave
            
            # Mettre à jour la caisse de l'AGENT
            caisse.solde_cash = Decimal(str(nouveau_cash)) if nouveau_cash else Decimal('0')
            caisse.solde_uv = Decimal(str(nouveau_uv)) if nouveau_uv else Decimal('0')
            caisse.solde_wave = Decimal(str(nouveau_wave)) if nouveau_wave else Decimal('0')
            caisse.save()
            
            messages.success(request, f"Caisse de l'agent {agent.nom} mise à jour avec succès")
            
        except Agent.DoesNotExist:
            messages.error(request, "Agent introuvable")
        except Caisse.DoesNotExist:
            messages.error(request, "Caisse de l'agent introuvable")
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
        
        return redirect('gestion_agents')
    
    return redirect('gestion_agents')

@login_required
def api_agent_caisse(request, agent_id):
    """
    API pour récupérer les soldes de la caisse d'un agent
    """
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Non autorisé'})
    
    try:
        agent = Agent.objects.get(id=agent_id)
        caisse = agent.user.caisse
        
        return JsonResponse({
            'success': True,
            'solde_cash': float(caisse.solde_cash),
            'solde_uv': float(caisse.solde_uv),
            'solde_wave': float(caisse.solde_wave),
        })
    except Agent.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Agent non trouvé'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def supprimer_agent(request):
    """
    Supprimer un agent (désactivation ou suppression définitive)
    """
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    if request.method == 'POST':
        agent_id = request.POST.get('agent_id')
        action = request.POST.get('action', 'desactiver')
        
        try:
            agent = Agent.objects.get(id=agent_id)
            
            if action == 'supprimer':
                # Suppression définitive
                user = agent.user
                agent.delete()
                user.delete()
                messages.success(request, f'✅ Agent "{agent.nom}" supprimé définitivement.')
            else:
                # Désactivation simple
                agent.est_actif = False
                agent.save()
                messages.success(request, f'✅ Agent "{agent.nom}" désactivé.')
                
        except Agent.DoesNotExist:
            messages.error(request, 'Agent non trouvé.')
    
    return redirect('gestion_agents')


@login_required
def activer_agent(request, agent_id):
    """
    Réactiver un agent désactivé
    """
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    try:
        agent = Agent.objects.get(id=agent_id)
        agent.est_actif = True
        agent.save()
        messages.success(request, f'✅ Agent "{agent.nom}" réactivé avec succès.')
    except Agent.DoesNotExist:
        messages.error(request, 'Agent non trouvé.')
    
    return redirect('gestion_agents')

@login_required
def detail_agent(request, agent_id):
    """
    Page dédiée à un agent avec tous ses détails
    """
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    agent = get_object_or_404(Agent, id=agent_id)
    user = agent.user
    caisse = user.caisse
    
    today = datetime.now().date()
    
    # ========== TRANSACTIONS D'AUJOURD'HUI ==========
    transactions_today = Transaction.objects.filter(user=user, date__date=today)
    
    # Calcul des variations des transactions
    cash_depot_today = transactions_today.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    cash_retrait_today = transactions_today.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    variation_cash_transactions = cash_depot_today - cash_retrait_today
    
    uv_depot_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_retrait_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_credit_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='credit'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_uv_transactions = uv_retrait_today - uv_depot_today - uv_credit_today
    
    wave_depot_today = transactions_today.filter(
        operateur='wave',
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    wave_retrait_today = transactions_today.filter(
        operateur='wave',
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_wave_transactions = wave_retrait_today - wave_depot_today
    
    # ========== DEMANDES VALIDÉES D'AUJOURD'HUI ==========
    demandes_validees_today = DemandeApprovisionnement.objects.filter(
        agent=agent,
        statut='valide',
        date_traitement__date=today
    )
    
    # Calcul des variations des demandes
    variation_cash_demandes = 0
    variation_uv_demandes = 0
    variation_wave_demandes = 0
    
    for demande in demandes_validees_today:
        if demande.type_echange == 'uv_to_cash':
            variation_uv_demandes -= demande.montant
            variation_cash_demandes += demande.montant
        elif demande.type_echange == 'wave_to_cash':
            variation_wave_demandes -= demande.montant
            variation_cash_demandes += demande.montant
        elif demande.type_echange == 'cash_to_uv':
            variation_cash_demandes -= demande.montant
            variation_uv_demandes += demande.montant
        elif demande.type_echange == 'cash_to_wave':
            variation_cash_demandes -= demande.montant
            variation_wave_demandes += demande.montant
    
    # ========== VARIATION TOTALE (transactions + demandes) ==========
    variation_cash_today = variation_cash_transactions + variation_cash_demandes
    variation_uv_today = variation_uv_transactions + variation_uv_demandes
    variation_wave_today = variation_wave_transactions + variation_wave_demandes
    
    # ========== SOLDES D'HIER ==========
    if caisse.last_balance_update == today:
        solde_cash_hier = caisse.solde_cash_hier
        solde_uv_hier = caisse.solde_uv_hier
        solde_wave_hier = caisse.solde_wave_hier
    else:
        solde_cash_hier = caisse.solde_cash - variation_cash_today
        solde_uv_hier = caisse.solde_uv - variation_uv_today
        solde_wave_hier = caisse.solde_wave - variation_wave_today
        
        caisse.solde_cash_hier = solde_cash_hier
        caisse.solde_uv_hier = solde_uv_hier
        caisse.solde_wave_hier = solde_wave_hier
        caisse.last_balance_update = today
        caisse.save()
    
    evolution_cash = caisse.solde_cash - solde_cash_hier
    evolution_uv = caisse.solde_uv - solde_uv_hier
    evolution_wave = caisse.solde_wave - solde_wave_hier
    
    # ========== FILTRES TRANSACTIONS ==========
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    type_filtre = request.GET.get('type')
    operateur_filtre = request.GET.get('operateur')
    show_all = request.GET.get('show_all')
    
    # ========== TRANSACTIONS ==========
    transactions = Transaction.objects.filter(user=user).order_by('-date')
    
    if not show_all and not date_debut and not date_fin and not type_filtre and not operateur_filtre:
        transactions = transactions.filter(date__date=today)
    else:
        if date_debut:
            try:
                date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
                transactions = transactions.filter(date__date__gte=date_debut_obj)
            except:
                pass
        if date_fin:
            try:
                date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
                transactions = transactions.filter(date__date__lte=date_fin_obj)
            except:
                pass
        if type_filtre:
            transactions = transactions.filter(type_transaction=type_filtre)
        if operateur_filtre:
            transactions = transactions.filter(operateur=operateur_filtre)
    
    # ========== STATS TRANSACTIONS ==========
    total_entree = transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    total_sortie = transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    total_commission = transactions.aggregate(Sum('commission'))['commission__sum'] or 0
    nombre_transactions = transactions.count()
    
    # ========== DEMANDES DE L'AGENT ==========
    show_all_demandes = request.GET.get('show_all_demandes')
    demande_statut = request.GET.get('demande_statut')
    demande_type = request.GET.get('demande_type')
    demande_date_debut = request.GET.get('demande_date_debut')
    demande_date_fin = request.GET.get('demande_date_fin')
    
    # Demandes ENVOYÉES par l'agent
    demandes = DemandeApprovisionnement.objects.filter(agent=agent).order_by('-date_demande')
    
    # Filtres des demandes
    if demande_date_debut:
        try:
            date_debut_obj = datetime.strptime(demande_date_debut, '%Y-%m-%d').date()
            demandes = demandes.filter(date_demande__date__gte=date_debut_obj)
        except:
            pass
    if demande_date_fin:
        try:
            date_fin_obj = datetime.strptime(demande_date_fin, '%Y-%m-%d').date()
            demandes = demandes.filter(date_demande__date__lte=date_fin_obj)
        except:
            pass
    if demande_statut:
        demandes = demandes.filter(statut=demande_statut)
    if demande_type:
        demandes = demandes.filter(type_echange=demande_type)
    
    if not show_all_demandes:
        demandes = demandes.filter(date_demande__date=today)
    
    total_demandes = demandes.count()
    
    # Stats des demandes (toutes, sans filtre pour les stats globales)
    demandes_all = DemandeApprovisionnement.objects.filter(agent=agent)
    demande_stats = {
        'attente': demandes_all.filter(statut='en_attente').count(),
        'valide': demandes_all.filter(statut='valide').count(),
        'refuse': demandes_all.filter(statut='refuse').count(),
        'total': demandes_all.count(),
    }
    
    # ========== EXPORT ==========
    export_format = request.GET.get('export')
    if export_format in ['csv', 'excel']:
        return export_transactions(transactions, agent, caisse, total_entree, total_sortie, total_commission, demandes, export_format)
    
    # ========== PAGINATION ==========
    page = request.GET.get('page', 1)
    paginator = Paginator(transactions, 15)
    transactions_page = paginator.get_page(page)
    
    context = {
        'title': f'Agent - {agent.nom}',
        'agent': agent,
        'caisse': caisse,
        'transactions': transactions_page,
        'total_entree': total_entree,
        'total_sortie': total_sortie,
        'total_commission': total_commission,
        'nombre_transactions': nombre_transactions,
        'demandes': demandes[:20],  # Limite à 20 pour l'affichage
        'demande_stats': demande_stats,
        'total_demandes': total_demandes,
        'solde_cash_hier': solde_cash_hier,
        'solde_uv_hier': solde_uv_hier,
        'solde_wave_hier': solde_wave_hier,
        'evolution_cash': evolution_cash,
        'evolution_uv': evolution_uv,
        'evolution_wave': evolution_wave,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'type_filtre': type_filtre,
        'operateur_filtre': operateur_filtre,
        'show_all': show_all,
        'show_all_demandes': show_all_demandes,
        'demande_statut': demande_statut,
        'demande_type': demande_type,
        'demande_date_debut': demande_date_debut,
        'demande_date_fin': demande_date_fin,
    }
    return render(request, 'transactions/detail_agent.html', context)


def export_transactions(transactions, agent, caisse, total_entree, total_sortie, total_commission, demandes, format_type):
    """
    Exporte les transactions, demandes et soldes au format CSV ou Excel
    """
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    
    # Calcul des soldes d'hier
    transactions_today = Transaction.objects.filter(user=agent.user, date__date=today)
    
    cash_depot_today = transactions_today.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    cash_retrait_today = transactions_today.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    variation_cash_today = cash_depot_today - cash_retrait_today
    
    uv_depot_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_retrait_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_credit_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='credit'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_uv_today = uv_retrait_today - uv_depot_today - uv_credit_today
    
    wave_depot_today = transactions_today.filter(
        operateur='wave',
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    wave_retrait_today = transactions_today.filter(
        operateur='wave',
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_wave_today = wave_retrait_today - wave_depot_today
    
    solde_cash_hier = caisse.solde_cash - variation_cash_today
    solde_uv_hier = caisse.solde_uv - variation_uv_today
    solde_wave_hier = caisse.solde_wave - variation_wave_today
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="rapport_{agent.nom}_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
        
        writer = csv.writer(response)
        
        # En-tête principal
        writer.writerow([f"RAPPORT DETAILLE - {agent.nom}"])
        writer.writerow([f"Date d'export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        writer.writerow([])
        
        # SOLDES
        writer.writerow(["=== SOLDES ==="])
        writer.writerow(["Compte", "Solde actuel", "Solde hier", "Variation"])
        writer.writerow(["Argent Cash", f"{caisse.solde_cash:,.0f} FCFA", f"{solde_cash_hier:,.0f} FCFA", f"{caisse.solde_cash - solde_cash_hier:+,.0f} FCFA"])
        writer.writerow(["UV Touspiont", f"{caisse.solde_uv:,.0f} FCFA", f"{solde_uv_hier:,.0f} FCFA", f"{caisse.solde_uv - solde_uv_hier:+,.0f} FCFA"])
        writer.writerow(["UV Wave", f"{caisse.solde_wave:,.0f} FCFA", f"{solde_wave_hier:,.0f} FCFA", f"{caisse.solde_wave - solde_wave_hier:+,.0f} FCFA"])
        writer.writerow([])
        
        # TOTAUX TRANSACTIONS
        writer.writerow(["=== TOTAUX DES TRANSACTIONS ==="])
        writer.writerow(["Total Entrées", f"{total_entree:,.0f} FCFA"])
        writer.writerow(["Total Sorties", f"{total_sortie:,.0f} FCFA"])
        writer.writerow(["Total Commission", f"{total_commission:,.0f} FCFA"])
        writer.writerow(["Nombre de transactions", transactions.count()])
        writer.writerow([])
        
        # DEMANDES
        writer.writerow(["=== DEMANDES D'APPROVISIONNEMENT ==="])
        writer.writerow(["Date", "Type", "Montant", "Statut", "Motif"])
        for d in demandes:
            writer.writerow([
                d.date_demande.strftime('%d/%m/%Y %H:%M'),
                d.get_type_echange_display(),
                f"{d.montant:,.0f} FCFA",
                d.get_statut_display(),
                d.motif or ""
            ])
        writer.writerow([])
        
        # DETAIL DES TRANSACTIONS
        writer.writerow(["=== DETAIL DES TRANSACTIONS ==="])
        writer.writerow(['Référence', 'Type', 'Opérateur', 'Client', 'Montant (FCFA)', 'Commission (FCFA)', 'Date'])
        
        for t in transactions:
            writer.writerow([
                t.reference,
                t.get_type_transaction_display(),
                t.get_operateur_display(),
                t.numero_client,
                f"{t.montant:,.0f}",
                f"{t.commission:,.0f}",
                t.date.strftime('%d/%m/%Y %H:%M:%S')
            ])
        
        return response
    
    elif format_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="rapport_{agent.nom}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        wb = Workbook()
        
        # Styles
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center')
        
        # ========== FEUILLE 1: RÉCAPITULATIF ==========
        ws_summary = wb.active
        ws_summary.title = "Récapitulatif"
        
        ws_summary['A1'] = f"RAPPORT - {agent.nom}"
        ws_summary['A1'].font = title_font
        ws_summary['A2'] = f"Date d'export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        # Soldes
        ws_summary['A4'] = "SOLDES"
        ws_summary['A4'].font = header_font
        ws_summary['A5'] = "Compte"
        ws_summary['B5'] = "Solde actuel"
        ws_summary['C5'] = "Solde hier"
        ws_summary['D5'] = "Variation"
        
        for col in range(1, 5):
            ws_summary.cell(row=5, column=col).font = header_font
        
        soldes_data = [
            ["Argent Cash", f"{caisse.solde_cash:,.0f} FCFA", f"{solde_cash_hier:,.0f} FCFA", f"{caisse.solde_cash - solde_cash_hier:+,.0f} FCFA"],
            ["UV Touspiont", f"{caisse.solde_uv:,.0f} FCFA", f"{solde_uv_hier:,.0f} FCFA", f"{caisse.solde_uv - solde_uv_hier:+,.0f} FCFA"],
            ["UV Wave", f"{caisse.solde_wave:,.0f} FCFA", f"{solde_wave_hier:,.0f} FCFA", f"{caisse.solde_wave - solde_wave_hier:+,.0f} FCFA"],
        ]
        
        for row, data in enumerate(soldes_data, 6):
            for col, val in enumerate(data, 1):
                ws_summary.cell(row=row, column=col, value=val)
        
        # Totaux transactions
        ws_summary['A10'] = "TOTAUX DES TRANSACTIONS"
        ws_summary['A10'].font = header_font
        ws_summary['A11'] = "Total Entrées"
        ws_summary['B11'] = f"{total_entree:,.0f} FCFA"
        ws_summary['A12'] = "Total Sorties"
        ws_summary['B12'] = f"{total_sortie:,.0f} FCFA"
        ws_summary['A13'] = "Total Commission"
        ws_summary['B13'] = f"{total_commission:,.0f} FCFA"
        ws_summary['A14'] = "Nombre de transactions"
        ws_summary['B14'] = transactions.count()
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 25
        ws_summary.column_dimensions['C'].width = 25
        ws_summary.column_dimensions['D'].width = 20
        
        # ========== FEUILLE 2: DEMANDES ==========
        ws_demandes = wb.create_sheet("Demandes")
        
        headers_demandes = ['Date', 'Type', 'Montant', 'Statut', 'Motif']
        for col, header in enumerate(headers_demandes, 1):
            cell = ws_demandes.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
        
        for row, d in enumerate(demandes, 2):
            ws_demandes.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y %H:%M'))
            ws_demandes.cell(row=row, column=2, value=d.get_type_echange_display())
            ws_demandes.cell(row=row, column=3, value=f"{d.montant:,.0f} FCFA")
            ws_demandes.cell(row=row, column=4, value=d.get_statut_display())
            ws_demandes.cell(row=row, column=5, value=d.motif or "")
        
        for col in range(1, 6):
            ws_demandes.column_dimensions[chr(64 + col)].width = 20
        
        # ========== FEUILLE 3: TRANSACTIONS ==========
        ws_trans = wb.create_sheet("Transactions")
        
        headers_trans = ['Référence', 'Type', 'Opérateur', 'Client', 'Montant (FCFA)', 'Commission (FCFA)', 'Date']
        for col, header in enumerate(headers_trans, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
        
        for row, t in enumerate(transactions, 2):
            ws_trans.cell(row=row, column=1, value=t.reference)
            ws_trans.cell(row=row, column=2, value=t.get_type_transaction_display())
            ws_trans.cell(row=row, column=3, value=t.get_operateur_display())
            ws_trans.cell(row=row, column=4, value=t.numero_client)
            ws_trans.cell(row=row, column=5, value=float(t.montant))
            ws_trans.cell(row=row, column=6, value=float(t.commission))
            ws_trans.cell(row=row, column=7, value=t.date.strftime('%d/%m/%Y %H:%M:%S'))
        
        for col in range(1, 8):
            ws_trans.column_dimensions[chr(64 + col)].width = 18
        
        wb.save(response)
        return response
    
    return None


# ==================== EXPORTS ====================

@login_required
def exporter_historique_agent(request, format_type):
    """
    Exporte les transactions de l'agent avec les filtres appliqués
    Inclut les soldes, variations et demandes
    format_type: 'csv' ou 'excel'
    """
    try:
        agent = Agent.objects.get(user=request.user)
    except Agent.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    # Récupérer les dates du filtre
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    # Définir la période à analyser
    if date_debut_str and date_fin_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            date_debut = timezone.now().date()
            date_fin = timezone.now().date()
    else:
        # Par défaut: aujourd'hui
        date_debut = timezone.now().date()
        date_fin = timezone.now().date()
    
    # Vérifier que date_debut <= date_fin
    if date_debut > date_fin:
        date_debut, date_fin = date_fin, date_debut
    
    # Date du jour pour le nom du fichier
    today = timezone.now().date()
    
    # Date de la veille pour calculer solde hier
    date_hier = date_debut - timedelta(days=1)
    
    # Récupérer la caisse de l'agent
    caisse = agent.user.caisse
    
    # Récupérer les transactions pour la période (date__date__range)
    transactions = Transaction.objects.filter(
        user=request.user,
        date__date__range=[date_debut, date_fin]
    ).order_by('-date')
    
    # Récupérer les demandes pour la période
    demandes = DemandeApprovisionnement.objects.filter(
        agent=agent,
        date_demande__date__range=[date_debut, date_fin]
    ).order_by('-date_demande')
    
    # Récupérer les transactions avant la période (jusqu'à la veille inclus)
    transactions_avant = Transaction.objects.filter(
        user=request.user,
        date__date__lte=date_hier
    )
    
    # ========== CALCUL DES TOTAUX POUR LA PÉRIODE ==========
    total_entree = transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    total_sortie = transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    
    # ========== CALCUL DES SOLDES ==========
    
            # 1. Calculer les soldes à HIER (avant la période)
    cash_depot_avant = transactions_avant.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    cash_retrait_avant = transactions_avant.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    variation_cash_avant = cash_depot_avant - cash_retrait_avant
    
    uv_depot_avant = transactions_avant.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_retrait_avant = transactions_avant.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_credit_avant = transactions_avant.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='credit'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_uv_avant = uv_retrait_avant - uv_depot_avant - uv_credit_avant
    
    wave_depot_avant = transactions_avant.filter(
        operateur='wave',
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    wave_retrait_avant = transactions_avant.filter(
        operateur='wave',
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_wave_avant = wave_retrait_avant - wave_depot_avant
    
    # Solde à HIER (avant la période)
    solde_cash_hier = caisse.solde_cash - variation_cash_avant
    solde_uv_hier = caisse.solde_uv - variation_uv_avant
    solde_wave_hier = caisse.solde_wave - variation_wave_avant
    
    # 2. Calculer les variations PENDANT la période
    cash_depot_periode = transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    cash_retrait_periode = transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    variation_cash_periode = cash_depot_periode - cash_retrait_periode
    
    uv_depot_periode = transactions.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_retrait_periode = transactions.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_credit_periode = transactions.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='credit'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_uv_periode = uv_retrait_periode - uv_depot_periode - uv_credit_periode
    
    wave_depot_periode = transactions.filter(
        operateur='wave',
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    wave_retrait_periode = transactions.filter(
        operateur='wave',
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_wave_periode = wave_retrait_periode - wave_depot_periode
    
    # Solde à la FIN de la période
    solde_cash_fin = solde_cash_hier + variation_cash_periode
    solde_uv_fin = solde_uv_hier + variation_uv_periode
    solde_wave_fin = solde_wave_hier + variation_wave_periode
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="historique_{today.strftime("%Y%m%d")}.csv"'
        
        response.write('\ufeff')
        writer = csv.writer(response)
        
        # En-tête principal
        writer.writerow([f"HISTORIQUE COMPLET - {agent.nom}"])
        writer.writerow([f"Période: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"])
        writer.writerow([f"Date d'export: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        writer.writerow([])
        
        # SOLDES
        writer.writerow(["=== SOLDES ==="])
        writer.writerow(["Compte", f"Solde au {date_hier.strftime('%d/%m/%Y')}", f"Solde au {date_fin.strftime('%d/%m/%Y')}", "Variation"])
        writer.writerow(["Argent Cash", f"{solde_cash_hier:,.0f} FCFA", f"{solde_cash_fin:,.0f} FCFA", f"{variation_cash_periode:+,.0f} FCFA"])
        writer.writerow(["UV Touspiont", f"{solde_uv_hier:,.0f} FCFA", f"{solde_uv_fin:,.0f} FCFA", f"{variation_uv_periode:+,.0f} FCFA"])
        writer.writerow(["UV Wave", f"{solde_wave_hier:,.0f} FCFA", f"{solde_wave_fin:,.0f} FCFA", f"{variation_wave_periode:+,.0f} FCFA"])
        writer.writerow([])
        
        # TOTAUX TRANSACTIONS
        writer.writerow(["=== TOTAUX DES TRANSACTIONS ==="])
        writer.writerow([f"Total Entrées (Dépôts)", f"{total_entree:,.0f} FCFA"])
        writer.writerow([f"Total Sorties (Retraits)", f"{total_sortie:,.0f} FCFA"])
        writer.writerow(["Nombre de transactions", transactions.count()])
        writer.writerow([])
        
        # DEMANDES
        writer.writerow(["=== DEMANDES D'APPROVISIONNEMENT ==="])
        writer.writerow(["Date", "Type", "Montant", "Statut", "Motif"])
        for d in demandes:
            writer.writerow([
                d.date_demande.strftime('%d/%m/%Y %H:%M'),
                d.get_type_echange_display(),
                f"{d.montant:,.0f} FCFA",
                d.get_statut_display(),
                d.motif or ""
            ])
        writer.writerow([])
        
        # DETAIL DES TRANSACTIONS
        writer.writerow(["=== DETAIL DES TRANSACTIONS ==="])
        writer.writerow(['Référence', 'Type', 'Opérateur', 'Client', 'Montant (FCFA)', 'Date'])
        
        for t in transactions:
            writer.writerow([
                t.reference,
                t.get_type_transaction_display(),
                t.get_operateur_display(),
                t.numero_client,
                f"{t.montant:,.0f}",
                t.date.strftime('%d/%m/%Y %H:%M:%S')
            ])
        
        return response
    
    elif format_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="historique_{today.strftime("%Y%m%d")}.xlsx"'
        
        wb = Workbook()
        
        # Styles
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center')
        
        # ========== FEUILLE 1: RÉCAPITULATIF ==========
        ws_summary = wb.active
        ws_summary.title = "Récapitulatif"
        
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = f"HISTORIQUE COMPLET - {agent.nom}"
        ws_summary['A1'].font = title_font
        ws_summary['A1'].alignment = center_align
        
        ws_summary['A2'] = f"Période: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        ws_summary['A3'] = f"Date d'export: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        # Soldes
        ws_summary['A5'] = "SOLDES"
        ws_summary['A5'].font = header_font
        ws_summary['A6'] = "Compte"
        ws_summary['B6'] = f"Solde au {date_hier.strftime('%d/%m/%Y')}"
        ws_summary['C6'] = f"Solde au {date_fin.strftime('%d/%m/%Y')}"
        ws_summary['D6'] = "Variation"
        
        for col in range(1, 5):
            ws_summary.cell(row=6, column=col).font = header_font
            ws_summary.cell(row=6, column=col).alignment = center_align
        
        soldes_data = [
            ["Argent Cash", f"{solde_cash_hier:,.0f} FCFA", f"{solde_cash_fin:,.0f} FCFA", f"{variation_cash_periode:+,.0f} FCFA"],
            ["UV Touspiont", f"{solde_uv_hier:,.0f} FCFA", f"{solde_uv_fin:,.0f} FCFA", f"{variation_uv_periode:+,.0f} FCFA"],
            ["UV Wave", f"{solde_wave_hier:,.0f} FCFA", f"{solde_wave_fin:,.0f} FCFA", f"{variation_wave_periode:+,.0f} FCFA"],
        ]
        
        for row, data in enumerate(soldes_data, 7):
            for col, val in enumerate(data, 1):
                ws_summary.cell(row=row, column=col, value=val)
        
        # Totaux transactions
        ws_summary['A11'] = "TOTAUX DES TRANSACTIONS"
        ws_summary['A11'].font = header_font
        ws_summary['A12'] = "Total Entrées (Dépôts)"
        ws_summary['B12'] = f"{total_entree:,.0f} FCFA"
        ws_summary['A13'] = "Total Sorties (Retraits)"
        ws_summary['B13'] = f"{total_sortie:,.0f} FCFA"
        ws_summary['A14'] = "Nombre de transactions"
        ws_summary['B14'] = transactions.count()
        
        # Stats des demandes
        ws_summary['A16'] = "STATISTIQUES DES DEMANDES"
        ws_summary['A16'].font = header_font
        ws_summary['A17'] = "En attente"
        ws_summary['B17'] = demandes.filter(statut='attente').count()
        ws_summary['A18'] = "Validées"
        ws_summary['B18'] = demandes.filter(statut='valide').count()
        ws_summary['A19'] = "Refusées"
        ws_summary['B19'] = demandes.filter(statut='refuse').count()
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 25
        ws_summary.column_dimensions['C'].width = 25
        ws_summary.column_dimensions['D'].width = 20
        
        # ========== FEUILLE 2: DEMANDES ==========
        ws_demandes = wb.create_sheet("Demandes")
        
        headers_demandes = ['Date', 'Type', 'Montant', 'Statut', 'Motif']
        for col, header in enumerate(headers_demandes, 1):
            cell = ws_demandes.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
        
        for row, d in enumerate(demandes, 2):
            ws_demandes.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y %H:%M'))
            ws_demandes.cell(row=row, column=2, value=d.get_type_echange_display())
            ws_demandes.cell(row=row, column=3, value=f"{d.montant:,.0f} FCFA")
            ws_demandes.cell(row=row, column=4, value=d.get_statut_display())
            ws_demandes.cell(row=row, column=5, value=d.motif or "")
        
        for col in range(1, 6):
            ws_demandes.column_dimensions[chr(64 + col)].width = 20
        
        # ========== FEUILLE 3: TRANSACTIONS ==========
        ws_trans = wb.create_sheet("Transactions")
        
        headers_trans = ['Référence', 'Type', 'Opérateur', 'Client', 'Montant (FCFA)', 'Date']
        for col, header in enumerate(headers_trans, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
        
        for row, t in enumerate(transactions, 2):
            ws_trans.cell(row=row, column=1, value=t.reference)
            ws_trans.cell(row=row, column=2, value=t.get_type_transaction_display())
            ws_trans.cell(row=row, column=3, value=t.get_operateur_display())
            ws_trans.cell(row=row, column=4, value=t.numero_client)
            ws_trans.cell(row=row, column=5, value=float(t.montant))
            ws_trans.cell(row=row, column=6, value=t.date.strftime('%d/%m/%Y %H:%M:%S'))
        
        # Format des nombres
        for row in range(2, transactions.count() + 2):
            ws_trans.cell(row=row, column=5).number_format = '#,##0'
        
        for col in range(1, 7):
            ws_trans.column_dimensions[chr(64 + col)].width = 18
        
        wb.save(response)
        return response
    
    return None


@login_required
def exporter_rapport_complet_agent(request, format_type):
    """
    Exporte un rapport complet: soldes, transactions, demandes
    Pour agent ou admin
    format_type: 'csv', 'excel' ou 'pdf'
    Prend en compte les filtres de date pour calculer les soldes
    """
    # Récupérer les dates du filtre (si présentes)
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    # Définir la période à analyser
    if date_debut_str and date_fin_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            date_debut = timezone.now().date()
            date_fin = timezone.now().date()
    else:
        # Par défaut: aujourd'hui
        date_debut = timezone.now().date()
        date_fin = timezone.now().date()
    
    # Date du jour pour le nom du fichier
    today = timezone.now().date()
    
    # Date de la veille pour calculer solde hier
    date_hier = date_debut - timedelta(days=1)
    
    # Vérifier si l'utilisateur est un agent ou un admin
    if hasattr(request.user, 'agent_profile'):
        # C'est un agent
        agent = request.user.agent_profile
        caisse = agent.user.caisse
        user_type = "Agent"
        user_name = agent.nom
        
        # Récupérer les transactions de l'agent pour la période
        transactions = Transaction.objects.filter(
            user=request.user,
            date__date__gte=date_debut,
            date__date__lte=date_fin
        ).order_by('-date')
        
        # Récupérer les demandes de l'agent UNIQUEMENT pour la période
        demandes = DemandeApprovisionnement.objects.filter(
            agent=agent,
            date_demande__date__gte=date_debut,
            date_demande__date__lte=date_fin
        ).order_by('-date_demande')
        
        # Calculer les soldes à la date de début
        transactions_avant = Transaction.objects.filter(
            user=request.user,
            date__date__lt=date_debut
        )
        
    elif hasattr(request.user, 'admin_profile'):
        # C'est un admin
        admin = request.user.admin_profile
        caisse = request.user.caisse
        user_type = "Administrateur"
        user_name = admin.nom
        
        # Récupérer toutes les transactions pour la période
        transactions = Transaction.objects.filter(
            date__date__gte=date_debut,
            date__date__lte=date_fin
        ).order_by('-date')
        
        # Récupérer toutes les demandes pour la période
        demandes = DemandeApprovisionnement.objects.filter(
            date_demande__date__gte=date_debut,
            date_demande__date__lte=date_fin
        ).order_by('-date_demande')
        
        # Calculer les soldes à la date de début
        transactions_avant = Transaction.objects.filter(
            date__date__lt=date_debut
        )
        
    elif hasattr(request.user, 'assistant_profile'):
        # C'est un assistant
        assistant = request.user.assistant_profile
        caisse = assistant.admin.user.caisse
        user_type = "Assistant"
        user_name = assistant.nom
        
        # Récupérer les transactions de l'assistant pour la période
        transactions = Transaction.objects.filter(
            user=request.user,
            date__date__gte=date_debut,
            date__date__lte=date_fin
        ).order_by('-date')
        
        # Récupérer les demandes
        demandes = DemandeApprovisionnement.objects.filter(
            date_demande__date__gte=date_debut,
            date_demande__date__lte=date_fin
        ).order_by('-date_demande')
        
        # Calculer les soldes à la date de début
        transactions_avant = Transaction.objects.filter(
            user=request.user,
            date__date__lt=date_debut
        )
        
    else:
        # Superutilisateur ou autre
        try:
            caisse = Caisse.objects.get(user=request.user)
            user_type = "Utilisateur"
            user_name = request.user.username
            transactions = Transaction.objects.filter(
                user=request.user,
                date__date__gte=date_debut,
                date__date__lte=date_fin
            ).order_by('-date')
            demandes = DemandeApprovisionnement.objects.filter(
                agent__user=request.user,
                date_demande__date__gte=date_debut,
                date_demande__date__lte=date_fin
            ).order_by('-date_demande')
            transactions_avant = Transaction.objects.filter(
                user=request.user,
                date__date__lt=date_debut
            )
        except:
            return HttpResponse("Impossible de générer le rapport. Données manquantes.", status=400)
    
    # ========== CALCUL DES TOTAUX POUR LA PÉRIODE ==========
    total_entree = transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    total_sortie = transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    
    # ========== CALCUL DES SOLDES ==========
    
    # 1. Calculer les soldes à HIER (avant la période)
    cash_depot_avant = transactions_avant.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    cash_retrait_avant = transactions_avant.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    variation_cash_avant = cash_depot_avant - cash_retrait_avant
    
    uv_depot_avant = transactions_avant.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_retrait_avant = transactions_avant.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_credit_avant = transactions_avant.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='credit'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_uv_avant = uv_retrait_avant - uv_depot_avant - uv_credit_avant
    
    wave_depot_avant = transactions_avant.filter(
        operateur='wave',
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    wave_retrait_avant = transactions_avant.filter(
        operateur='wave',
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_wave_avant = wave_retrait_avant - wave_depot_avant
    
    # Solde à HIER (avant la période)
    solde_cash_hier = caisse.solde_cash - variation_cash_avant
    solde_uv_hier = caisse.solde_uv - variation_uv_avant
    solde_wave_hier = caisse.solde_wave - variation_wave_avant
    
    # 2. Calculer les variations PENDANT la période
    cash_depot_periode = transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    cash_retrait_periode = transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    variation_cash_periode = cash_depot_periode - cash_retrait_periode
    
    uv_depot_periode = transactions.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_retrait_periode = transactions.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_credit_periode = transactions.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='credit'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_uv_periode = uv_retrait_periode - uv_depot_periode - uv_credit_periode
    
    wave_depot_periode = transactions.filter(
        operateur='wave',
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    wave_retrait_periode = transactions.filter(
        operateur='wave',
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_wave_periode = wave_retrait_periode - wave_depot_periode
    
    # Solde à la FIN de la période
    solde_cash_fin = solde_cash_hier + variation_cash_periode
    solde_uv_fin = solde_uv_hier + variation_uv_periode
    solde_wave_fin = solde_wave_hier + variation_wave_periode
    
    # ========== EXPORT PDF ==========
    if format_type == 'pdf':
        # Créer le contexte pour le template PDF
        context = {
            'user_name': user_name,
            'user_type': user_type,
            'date_export': timezone.now(),
            'date_debut': date_debut,
            'date_fin': date_fin,
            'date_hier': date_hier,
            'solde_cash_hier': solde_cash_hier,
            'solde_uv_hier': solde_uv_hier,
            'solde_wave_hier': solde_wave_hier,
            'solde_cash_fin': solde_cash_fin,
            'solde_uv_fin': solde_uv_fin,
            'solde_wave_fin': solde_wave_fin,
            'variation_cash_periode': variation_cash_periode,
            'variation_uv_periode': variation_uv_periode,
            'variation_wave_periode': variation_wave_periode,
            'total_entree': total_entree,
            'total_sortie': total_sortie,
            'transactions': transactions,
            'demandes': demandes,
        }
        
        # Rendre le template HTML
        template = get_template('transactions/rapport_pdf.html')
        html = template.render(context)
        
        # Créer le PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_{today.strftime("%Y%m%d")}.pdf"'
        
        # Convertir HTML en PDF
        pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('UTF-8')), dest=response)
        
        if pisa_status.err:
            return HttpResponse('Erreur lors de la génération du PDF', status=400)
        
        return response
    
    # ========== EXPORT CSV ==========
    elif format_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="rapport_{today.strftime("%Y%m%d")}.csv"'
        
        response.write('\ufeff')
        writer = csv.writer(response)
        
        writer.writerow([f"RAPPORT COMPLET - {user_name} ({user_type})"])
        writer.writerow([f"Date d'export: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        writer.writerow([])
        
        writer.writerow(["=== SOLDES ==="])
        writer.writerow(["Compte", f"Solde au {date_hier.strftime('%d/%m/%Y')}", f"Solde au {date_fin.strftime('%d/%m/%Y')}", "Variation"])
        writer.writerow(["Argent Cash", f"{solde_cash_hier:,.0f} FCFA", f"{solde_cash_fin:,.0f} FCFA", f"{variation_cash_periode:+,.0f} FCFA"])
        writer.writerow(["UV Touspiont", f"{solde_uv_hier:,.0f} FCFA", f"{solde_uv_fin:,.0f} FCFA", f"{variation_uv_periode:+,.0f} FCFA"])
        writer.writerow(["UV Wave", f"{solde_wave_hier:,.0f} FCFA", f"{solde_wave_fin:,.0f} FCFA", f"{variation_wave_periode:+,.0f} FCFA"])
        writer.writerow([])
        
        writer.writerow(["=== TOTAUX DES TRANSACTIONS ==="])
        writer.writerow([f"Total Entrées (Dépôts)", f"{total_entree:,.0f} FCFA"])
        writer.writerow([f"Total Sorties (Retraits)", f"{total_sortie:,.0f} FCFA"])
        writer.writerow(["Nombre de transactions", transactions.count()])
        writer.writerow([])
        
        writer.writerow(["=== DEMANDES D'APPROVISIONNEMENT ==="])
        writer.writerow(["Date", "Type", "Montant", "Statut", "Motif"])
        for d in demandes:
            writer.writerow([
                d.date_demande.strftime('%d/%m/%Y %H:%M'),
                d.get_type_echange_display(),
                f"{d.montant:,.0f} FCFA",
                d.get_statut_display(),
                d.motif or ""
            ])
        writer.writerow([])
        
        writer.writerow(["=== DETAIL DES TRANSACTIONS ==="])
        writer.writerow(['Référence', 'Type', 'Opérateur', 'Client', 'Montant (FCFA)', 'Date'])
        
        for t in transactions:
            writer.writerow([
                t.reference,
                t.get_type_transaction_display(),
                t.get_operateur_display(),
                t.numero_client,
                f"{t.montant:,.0f}",
                t.date.strftime('%d/%m/%Y %H:%M:%S')
            ])
        
        return response
    
    # ========== EXPORT EXCEL ==========
    elif format_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="rapport_{today.strftime("%Y%m%d")}.xlsx"'
        
        wb = Workbook()
        
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center')
        
        ws_summary = wb.active
        ws_summary.title = "Récapitulatif"
        
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = f"RAPPORT COMPLET - {user_name} ({user_type})"
        ws_summary['A1'].font = title_font
        ws_summary['A1'].alignment = center_align
        
        ws_summary['A2'] = f"Date d'export: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        ws_summary['A4'] = "SOLDES"
        ws_summary['A4'].font = header_font
        ws_summary['A5'] = "Compte"
        ws_summary['B5'] = f"Solde au {date_hier.strftime('%d/%m/%Y')}"
        ws_summary['C5'] = f"Solde au {date_fin.strftime('%d/%m/%Y')}"
        ws_summary['D5'] = "Variation"
        
        for col in range(1, 5):
            ws_summary.cell(row=5, column=col).font = header_font
            ws_summary.cell(row=5, column=col).alignment = center_align
        
        soldes_data = [
            ["Argent Cash", f"{solde_cash_hier:,.0f} FCFA", f"{solde_cash_fin:,.0f} FCFA", f"{variation_cash_periode:+,.0f} FCFA"],
            ["UV Touspiont", f"{solde_uv_hier:,.0f} FCFA", f"{solde_uv_fin:,.0f} FCFA", f"{variation_uv_periode:+,.0f} FCFA"],
            ["UV Wave", f"{solde_wave_hier:,.0f} FCFA", f"{solde_wave_fin:,.0f} FCFA", f"{variation_wave_periode:+,.0f} FCFA"],
        ]
        
        for row, data in enumerate(soldes_data, 6):
            for col, val in enumerate(data, 1):
                ws_summary.cell(row=row, column=col, value=val)
        
        ws_summary['A10'] = "TOTAUX DES TRANSACTIONS"
        ws_summary['A10'].font = header_font
        ws_summary['A11'] = "Total Entrées (Dépôts)"
        ws_summary['B11'] = f"{total_entree:,.0f} FCFA"
        ws_summary['A12'] = "Total Sorties (Retraits)"
        ws_summary['B12'] = f"{total_sortie:,.0f} FCFA"
        ws_summary['A13'] = "Nombre de transactions"
        ws_summary['B13'] = transactions.count()
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 25
        ws_summary.column_dimensions['C'].width = 25
        ws_summary.column_dimensions['D'].width = 20
        
        ws_demandes = wb.create_sheet("Demandes")
        
        headers_demandes = ['Date', 'Type', 'Montant', 'Statut', 'Motif']
        for col, header in enumerate(headers_demandes, 1):
            cell = ws_demandes.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
        
        for row, d in enumerate(demandes, 2):
            ws_demandes.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y %H:%M'))
            ws_demandes.cell(row=row, column=2, value=d.get_type_echange_display())
            ws_demandes.cell(row=row, column=3, value=f"{d.montant:,.0f} FCFA")
            ws_demandes.cell(row=row, column=4, value=d.get_statut_display())
            ws_demandes.cell(row=row, column=5, value=d.motif or "")
        
        for col in range(1, 6):
            ws_demandes.column_dimensions[chr(64 + col)].width = 20
        
        ws_trans = wb.create_sheet("Transactions")
        
        headers_trans = ['Référence', 'Type', 'Opérateur', 'Client', 'Montant (FCFA)', 'Date']
        for col, header in enumerate(headers_trans, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
        
        for row, t in enumerate(transactions, 2):
            ws_trans.cell(row=row, column=1, value=t.reference)
            ws_trans.cell(row=row, column=2, value=t.get_type_transaction_display())
            ws_trans.cell(row=row, column=3, value=t.get_operateur_display())
            ws_trans.cell(row=row, column=4, value=t.numero_client)
            ws_trans.cell(row=row, column=5, value=float(t.montant))
            ws_trans.cell(row=row, column=6, value=t.date.strftime('%d/%m/%Y %H:%M:%S'))
        
        for row in range(2, transactions.count() + 2):
            ws_trans.cell(row=row, column=5).number_format = '#,##0'
        
        for col in range(1, 7):
            ws_trans.column_dimensions[chr(64 + col)].width = 18
        
        wb.save(response)
        return response
    
    return None


# ==================== AJAX ====================

@login_required
@require_POST
def ajax_calculer_frais(request):
    """
    API pour calculer les frais en temps réel (AJAX)
    """
    data = json.loads(request.body)
    operateur = data.get('operateur')
    type_transaction = data.get('type')
    montant = Decimal(data.get('montant', 0))
    
    temp_transaction = Transaction(
        operateur=operateur,
        type_transaction=type_transaction,
        montant=montant
    )
    
    commission = temp_transaction.calculer_commission()
    frais = temp_transaction.calculer_frais_operateur()
    
    return JsonResponse({
        'commission': str(commission),
        'frais': str(frais),
        'total_a_payer': str(montant + frais) if type_transaction == 'depot' else str(montant)
    })


# ==================== GESTION DES ASSISTANTS (Admin uniquement) ====================

@login_required
def gestion_assistants(request):
    """Page de gestion des assistants"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    assistants = Assistant.objects.filter(admin=admin).order_by('-created_at')
    assistants_actifs = assistants.filter(est_actif=True).count()
    assistants_inactifs = assistants.filter(est_actif=False).count()
    
    context = {
        'title': 'Gestion des assistants',
        'assistants': assistants,
        'assistants_actifs': assistants_actifs,
        'assistants_inactifs': assistants_inactifs,
    }
    return render(request, 'transactions/gestion_assistants.html', context)


@login_required
def ajouter_assistant(request):
    """Ajouter un assistant"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    if request.method == 'POST':
        nom = request.POST.get('nom')
        telephone = request.POST.get('telephone')
        email = request.POST.get('email', '')
        username = request.POST.get('username')
        password = request.POST.get('password')
        est_actif = request.POST.get('est_actif') == 'true'
        
        if not nom or not telephone or not username or not password:
            messages.error(request, 'Le nom, le téléphone, l\'identifiant et le mot de passe sont obligatoires.')
            return redirect('gestion_agents')  # ← Redirige vers gestion_agents
        
        if User.objects.filter(username=username).exists():
            messages.error(request, f"Le nom d'utilisateur '{username}' existe déjà.")
            return redirect('gestion_agents')
        
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password
        )
        
        assistant = Assistant.objects.create(
            user=user,
            nom=nom,
            telephone=telephone,
            email=email,
            admin=admin,
            est_actif=est_actif,
            created_by=admin
        )
        
        messages.success(request, f'✅ Assistant "{nom}" créé avec succès. Identifiant: {username}')
        return redirect('gestion_agents')
    
    return redirect('gestion_agents')


@login_required
def modifier_assistant(request, assistant_id):
    """Modifier un assistant"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    assistant = get_object_or_404(Assistant, id=assistant_id, admin=admin)
    
    if request.method == 'POST':
        assistant.nom = request.POST.get('nom', assistant.nom)
        assistant.telephone = request.POST.get('telephone', assistant.telephone)
        assistant.email = request.POST.get('email', assistant.email)
        assistant.est_actif = request.POST.get('est_actif') == 'true'
        assistant.save()
        
        messages.success(request, f'✅ Assistant "{assistant.nom}" modifié avec succès.')
        return redirect('gestion_assistants')
    
    return redirect('gestion_assistants')


@login_required
def modifier_mot_de_passe_assistant(request, assistant_id):
    """Modifier le mot de passe d'un assistant"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    assistant = get_object_or_404(Assistant, id=assistant_id, admin=admin)
    
    if request.method == 'POST':
        nouveau_password = request.POST.get('nouveau_password')
        
        if not nouveau_password or len(nouveau_password) < 4:
            messages.error(request, 'Le mot de passe doit contenir au moins 4 caractères.')
            return redirect('gestion_assistants')
        
        assistant.user.set_password(nouveau_password)
        assistant.user.save()
        
        messages.success(request, f'✅ Mot de passe modifié pour "{assistant.nom}".')
        return redirect('gestion_assistants')
    
    return redirect('gestion_assistants')


@login_required
def toggle_assistant_status(request, assistant_id):
    """Activer/Désactiver un assistant"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    assistant = get_object_or_404(Assistant, id=assistant_id, admin=admin)
    assistant.est_actif = not assistant.est_actif
    assistant.save()
    
    status = "activé" if assistant.est_actif else "désactivé"
    messages.success(request, f'✅ Assistant "{assistant.nom}" {status}.')
    return redirect('gestion_assistants')


@login_required
def supprimer_assistant(request, assistant_id):
    """Supprimer un assistant"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    assistant = get_object_or_404(Assistant, id=assistant_id, admin=admin)
    nom = assistant.nom
    user = assistant.user
    
    assistant.delete()
    user.delete()
    
    messages.success(request, f'✅ Assistant "{nom}" supprimé définitivement.')
    return redirect('gestion_assistants')

@login_required
def detail_assistant(request, assistant_id):
    """Page dédiée à un assistant avec ses demandes reçues"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    assistant = get_object_or_404(Assistant, id=assistant_id, admin=admin)
    caisse = assistant.admin.user.caisse  # L'assistant partage la caisse de l'admin
    
    today = datetime.now().date()
    
    # ========== TRANSACTIONS DE L'ASSISTANT ==========
    transactions = Transaction.objects.filter(user=assistant.user).order_by('-date')
    
    # ========== TRANSACTIONS D'AUJOURD'HUI ==========
    transactions_today = transactions.filter(date__date=today)
    
    # Calcul des variations des transactions
    cash_depot_today = transactions_today.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    cash_retrait_today = transactions_today.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    variation_cash_transactions = cash_depot_today - cash_retrait_today
    
    uv_depot_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_retrait_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_credit_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='credit'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_uv_transactions = uv_retrait_today - uv_depot_today - uv_credit_today
    
    wave_depot_today = transactions_today.filter(
        operateur='wave',
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    wave_retrait_today = transactions_today.filter(
        operateur='wave',
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_wave_transactions = wave_retrait_today - wave_depot_today
    
    # ========== DEMANDES VALIDÉES PAR L'ASSISTANT AUJOURD'HUI ==========
    demandes_validees_today = DemandeApprovisionnement.objects.filter(
        destinataire_type='assistant',
        assistant_destinataire=assistant,
        statut='valide',
        date_traitement__date=today
    )
    
    # Calcul des variations des demandes
    variation_cash_demandes = 0
    variation_uv_demandes = 0
    variation_wave_demandes = 0
    
    for demande in demandes_validees_today:
        if demande.type_echange == 'uv_to_cash':
            variation_uv_demandes -= demande.montant
            variation_cash_demandes += demande.montant
        elif demande.type_echange == 'wave_to_cash':
            variation_wave_demandes -= demande.montant
            variation_cash_demandes += demande.montant
        elif demande.type_echange == 'cash_to_uv':
            variation_cash_demandes -= demande.montant
            variation_uv_demandes += demande.montant
        elif demande.type_echange == 'cash_to_wave':
            variation_cash_demandes -= demande.montant
            variation_wave_demandes += demande.montant
    
    # ========== VARIATION TOTALE (transactions + demandes) ==========
    variation_cash_today = variation_cash_transactions + variation_cash_demandes
    variation_uv_today = variation_uv_transactions + variation_uv_demandes
    variation_wave_today = variation_wave_transactions + variation_wave_demandes
    
    # ========== SOLDES D'HIER ==========
    if caisse.last_balance_update == today:
        solde_cash_hier = caisse.solde_cash_hier
        solde_uv_hier = caisse.solde_uv_hier
        solde_wave_hier = caisse.solde_wave_hier
    else:
        solde_cash_hier = caisse.solde_cash - variation_cash_today
        solde_uv_hier = caisse.solde_uv - variation_uv_today
        solde_wave_hier = caisse.solde_wave - variation_wave_today
        
        caisse.solde_cash_hier = solde_cash_hier
        caisse.solde_uv_hier = solde_uv_hier
        caisse.solde_wave_hier = solde_wave_hier
        caisse.last_balance_update = today
        caisse.save()
    
    evolution_cash = caisse.solde_cash - solde_cash_hier
    evolution_uv = caisse.solde_uv - solde_uv_hier
    evolution_wave = caisse.solde_wave - solde_wave_hier
    
    # ========== DEMANDES REÇUES PAR L'ASSISTANT ==========
    demandes = DemandeApprovisionnement.objects.filter(
        destinataire_type='assistant',
        assistant_destinataire=assistant
    ).order_by('-date_demande')
    
    # ========== STATS DEMANDES ==========
    demande_stats = {
        'total': demandes.count(),
        'en_attente': demandes.filter(statut='en_attente').count(),
        'validees': demandes.filter(statut='valide').count(),
        'refusees': demandes.filter(statut='refuse').count(),
        'montant_total': demandes.aggregate(Sum('montant'))['montant__sum'] or 0,
    }
    
    # ========== FILTRES DEMANDES ==========
    demande_date_debut = request.GET.get('demande_date_debut')
    demande_date_fin = request.GET.get('demande_date_fin')
    demande_statut = request.GET.get('demande_statut')
    demande_type = request.GET.get('demande_type')
    
    demandes_filtrees = demandes
    if demande_date_debut:
        try:
            date_debut_obj = datetime.strptime(demande_date_debut, '%Y-%m-%d').date()
            demandes_filtrees = demandes_filtrees.filter(date_demande__date__gte=date_debut_obj)
        except:
            pass
    if demande_date_fin:
        try:
            date_fin_obj = datetime.strptime(demande_date_fin, '%Y-%m-%d').date()
            demandes_filtrees = demandes_filtrees.filter(date_demande__date__lte=date_fin_obj)
        except:
            pass
    if demande_statut:
        demandes_filtrees = demandes_filtrees.filter(statut=demande_statut)
    if demande_type:
        demandes_filtrees = demandes_filtrees.filter(type_echange=demande_type)
    
    # ========== STATS TRANSACTIONS ==========
    total_entree = transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    total_sortie = transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    total_commission = transactions.aggregate(Sum('commission'))['commission__sum'] or 0
    nombre_transactions = transactions.count()
    
    # ========== FILTRES TRANSACTIONS ==========
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    type_filtre = request.GET.get('type')
    operateur_filtre = request.GET.get('operateur')
    show_all = request.GET.get('show_all')
    
    transactions_filtrees = transactions
    if not show_all and not date_debut and not date_fin and not type_filtre and not operateur_filtre:
        transactions_filtrees = transactions_filtrees.filter(date__date=today)
    else:
        if date_debut:
            try:
                date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
                transactions_filtrees = transactions_filtrees.filter(date__date__gte=date_debut_obj)
            except:
                pass
        if date_fin:
            try:
                date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
                transactions_filtrees = transactions_filtrees.filter(date__date__lte=date_fin_obj)
            except:
                pass
        if type_filtre:
            transactions_filtrees = transactions_filtrees.filter(type_transaction=type_filtre)
        if operateur_filtre:
            transactions_filtrees = transactions_filtrees.filter(operateur=operateur_filtre)
    
    # ========== PAGINATION ==========
    page = request.GET.get('page', 1)
    paginator = Paginator(transactions_filtrees, 15)
    transactions_page = paginator.get_page(page)
    
    context = {
        'title': f'Assistant - {assistant.nom}',
        'assistant': assistant,
        'caisse': caisse,
        'transactions': transactions_page,
        'total_entree': total_entree,
        'total_sortie': total_sortie,
        'total_commission': total_commission,
        'nombre_transactions': nombre_transactions,
        'demandes': demandes_filtrees[:20],
        'demande_stats': demande_stats,
        'demande_date_debut': demande_date_debut,
        'demande_date_fin': demande_date_fin,
        'demande_statut': demande_statut,
        'demande_type': demande_type,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'type_filtre': type_filtre,
        'operateur_filtre': operateur_filtre,
        'show_all': show_all,
        'solde_cash_hier': solde_cash_hier,
        'solde_uv_hier': solde_uv_hier,
        'solde_wave_hier': solde_wave_hier,
        'evolution_cash': evolution_cash,
        'evolution_uv': evolution_uv,
        'evolution_wave': evolution_wave,
    }
    return render(request, 'transactions/detail_assistant.html', context)

@login_required
def modifier_assistant(request, assistant_id):
    """Modifier un assistant"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    assistant = get_object_or_404(Assistant, id=assistant_id, admin=admin)
    
    if request.method == 'POST':
        assistant.nom = request.POST.get('nom', assistant.nom)
        assistant.telephone = request.POST.get('telephone', assistant.telephone)
        assistant.email = request.POST.get('email', assistant.email)
        assistant.est_actif = request.POST.get('est_actif') == 'true'
        assistant.save()
        
        messages.success(request, f'✅ Assistant "{assistant.nom}" modifié avec succès.')
        return redirect('gestion_assistants')
    
    return redirect('gestion_assistants')


@login_required
def activer_assistant(request, assistant_id):
    """Activer/Désactiver un assistant"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    assistant = get_object_or_404(Assistant, id=assistant_id, admin=admin)
    assistant.est_actif = not assistant.est_actif
    assistant.save()
    
    status = "activé" if assistant.est_actif else "désactivé"
    messages.success(request, f'✅ Assistant "{assistant.nom}" {status}.')
    return redirect('gestion_assistants')


@login_required
def supprimer_assistant(request, assistant_id):
    """Supprimer un assistant"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    assistant = get_object_or_404(Assistant, id=assistant_id, admin=admin)
    nom = assistant.nom
    user = assistant.user
    
    assistant.delete()
    user.delete()
    
    messages.success(request, f'✅ Assistant "{nom}" supprimé définitivement.')
    return redirect('gestion_assistants')


@login_required
def detail_assistant(request, assistant_id):
    """Page dédiée à un assistant avec ses demandes reçues"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    assistant = get_object_or_404(Assistant, id=assistant_id, admin=admin)
    caisse = assistant.admin.user.caisse  # L'assistant partage la caisse de l'admin
    
    today = datetime.now().date()
    
    # ========== TRANSACTIONS DE L'ASSISTANT ==========
    transactions = Transaction.objects.filter(user=assistant.user).order_by('-date')
    
    # ========== TRANSACTIONS D'AUJOURD'HUI ==========
    transactions_today = transactions.filter(date__date=today)
    
    # Calcul des variations des transactions
    cash_depot_today = transactions_today.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    cash_retrait_today = transactions_today.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    variation_cash_transactions = cash_depot_today - cash_retrait_today
    
    uv_depot_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_retrait_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_credit_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='credit'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_uv_transactions = uv_retrait_today - uv_depot_today - uv_credit_today
    
    wave_depot_today = transactions_today.filter(
        operateur='wave',
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    wave_retrait_today = transactions_today.filter(
        operateur='wave',
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_wave_transactions = wave_retrait_today - wave_depot_today
    
    # ========== DEMANDES VALIDÉES PAR L'ASSISTANT AUJOURD'HUI ==========
    demandes_validees_today = DemandeApprovisionnement.objects.filter(
        destinataire_type='assistant',
        assistant_destinataire=assistant,
        statut='valide',
        date_traitement__date=today
    )
    
    # Calcul des variations des demandes
    variation_cash_demandes = 0
    variation_uv_demandes = 0
    variation_wave_demandes = 0
    
    for demande in demandes_validees_today:
        if demande.type_echange == 'uv_to_cash':
            variation_uv_demandes -= demande.montant
            variation_cash_demandes += demande.montant
        elif demande.type_echange == 'wave_to_cash':
            variation_wave_demandes -= demande.montant
            variation_cash_demandes += demande.montant
        elif demande.type_echange == 'cash_to_uv':
            variation_cash_demandes -= demande.montant
            variation_uv_demandes += demande.montant
        elif demande.type_echange == 'cash_to_wave':
            variation_cash_demandes -= demande.montant
            variation_wave_demandes += demande.montant
    
    # ========== VARIATION TOTALE (transactions + demandes) ==========
    variation_cash_today = variation_cash_transactions + variation_cash_demandes
    variation_uv_today = variation_uv_transactions + variation_uv_demandes
    variation_wave_today = variation_wave_transactions + variation_wave_demandes
    
    # ========== SOLDES D'HIER ==========
    if caisse.last_balance_update == today:
        solde_cash_hier = caisse.solde_cash_hier
        solde_uv_hier = caisse.solde_uv_hier
        solde_wave_hier = caisse.solde_wave_hier
    else:
        solde_cash_hier = caisse.solde_cash - variation_cash_today
        solde_uv_hier = caisse.solde_uv - variation_uv_today
        solde_wave_hier = caisse.solde_wave - variation_wave_today
        
        caisse.solde_cash_hier = solde_cash_hier
        caisse.solde_uv_hier = solde_uv_hier
        caisse.solde_wave_hier = solde_wave_hier
        caisse.last_balance_update = today
        caisse.save()
    
    evolution_cash = caisse.solde_cash - solde_cash_hier
    evolution_uv = caisse.solde_uv - solde_uv_hier
    evolution_wave = caisse.solde_wave - solde_wave_hier
    
    # ========== FILTRES TRANSACTIONS ==========
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    type_filtre = request.GET.get('type')
    operateur_filtre = request.GET.get('operateur')
    show_all = request.GET.get('show_all')
    
    # ========== TRANSACTIONS ==========
    if not show_all and not date_debut and not date_fin and not type_filtre and not operateur_filtre:
        transactions = transactions.filter(date__date=today)
    else:
        if date_debut:
            try:
                date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
                transactions = transactions.filter(date__date__gte=date_debut_obj)
            except:
                pass
        if date_fin:
            try:
                date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
                transactions = transactions.filter(date__date__lte=date_fin_obj)
            except:
                pass
        if type_filtre:
            transactions = transactions.filter(type_transaction=type_filtre)
        if operateur_filtre:
            transactions = transactions.filter(operateur=operateur_filtre)
    
    # ========== STATS TRANSACTIONS ==========
    total_entree = transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    total_sortie = transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    total_commission = transactions.aggregate(Sum('commission'))['commission__sum'] or 0
    nombre_transactions = transactions.count()
    
    # ========== DEMANDES REÇUES PAR L'ASSISTANT ==========
    show_all_demandes = request.GET.get('show_all_demandes')
    demande_statut = request.GET.get('demande_statut')
    demande_type = request.GET.get('demande_type')
    demande_date_debut = request.GET.get('demande_date_debut')
    demande_date_fin = request.GET.get('demande_date_fin')
    
    # Demandes REÇUES par l'assistant
    demandes = DemandeApprovisionnement.objects.filter(
        destinataire_type='assistant',
        assistant_destinataire=assistant
    ).order_by('-date_demande')
    
    # Filtres des demandes
    if demande_date_debut:
        try:
            date_debut_obj = datetime.strptime(demande_date_debut, '%Y-%m-%d').date()
            demandes = demandes.filter(date_demande__date__gte=date_debut_obj)
        except:
            pass
    if demande_date_fin:
        try:
            date_fin_obj = datetime.strptime(demande_date_fin, '%Y-%m-%d').date()
            demandes = demandes.filter(date_demande__date__lte=date_fin_obj)
        except:
            pass
    if demande_statut:
        demandes = demandes.filter(statut=demande_statut)
    if demande_type:
        demandes = demandes.filter(type_echange=demande_type)
    
    if not show_all_demandes:
        demandes = demandes.filter(date_demande__date=today)
    
    total_demandes = demandes.count()
    
    # Stats des demandes (toutes, sans filtre pour les stats globales)
    demandes_all = DemandeApprovisionnement.objects.filter(
        destinataire_type='assistant',
        assistant_destinataire=assistant
    )
    demande_stats = {
        'attente': demandes_all.filter(statut='en_attente').count(),
        'valide': demandes_all.filter(statut='valide').count(),
        'refuse': demandes_all.filter(statut='refuse').count(),
        'total': demandes_all.count(),
        'montant_total': demandes_all.aggregate(Sum('montant'))['montant__sum'] or 0,
    }
    
    # ========== EXPORT ==========
    export_format = request.GET.get('export')
    if export_format in ['csv', 'excel']:
        return export_transactions(transactions, assistant, caisse, total_entree, total_sortie, total_commission, demandes, export_format)
    
    # ========== PAGINATION ==========
    page = request.GET.get('page', 1)
    paginator = Paginator(transactions, 15)
    transactions_page = paginator.get_page(page)
    
    context = {
        'title': f'Assistant - {assistant.nom}',
        'assistant': assistant,
        'caisse': caisse,
        'transactions': transactions_page,
        'total_entree': total_entree,
        'total_sortie': total_sortie,
        'total_commission': total_commission,
        'nombre_transactions': nombre_transactions,
        'demandes': demandes[:20],  # Limite à 20 pour l'affichage
        'demande_stats': demande_stats,
        'total_demandes': total_demandes,
        'solde_cash_hier': solde_cash_hier,
        'solde_uv_hier': solde_uv_hier,
        'solde_wave_hier': solde_wave_hier,
        'evolution_cash': evolution_cash,
        'evolution_uv': evolution_uv,
        'evolution_wave': evolution_wave,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'type_filtre': type_filtre,
        'operateur_filtre': operateur_filtre,
        'show_all': show_all,
        'show_all_demandes': show_all_demandes,
        'demande_statut': demande_statut,
        'demande_type': demande_type,
        'demande_date_debut': demande_date_debut,
        'demande_date_fin': demande_date_fin,
    }
    return render(request, 'transactions/detail_assistant.html', context)


@login_required
def modifier_mot_de_passe_assistant(request, assistant_id):
    """Modifier le mot de passe d'un assistant"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    assistant = get_object_or_404(Assistant, id=assistant_id, admin=admin)
    
    if request.method == 'POST':
        nouveau_password = request.POST.get('nouveau_password')
        
        if not nouveau_password or len(nouveau_password) < 4:
            messages.error(request, 'Le mot de passe doit contenir au moins 4 caractères.')
            return redirect('gestion_assistants')
        
        assistant.user.set_password(nouveau_password)
        assistant.user.save()
        
        messages.success(request, f'✅ Mot de passe modifié pour "{assistant.nom}".')
        return redirect('gestion_assistants')
    
    return redirect('gestion_assistants')
# ==================== RAPPORTS ADMIN ====================

import json
import csv
from datetime import datetime, timedelta
from django.db.models import Sum
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import (
    Transaction, Agent, Admin, Caisse, DemandeApprovisionnement,
    Facture, PaiementFacture, Dette, RemboursementDette,
    CompteEpargne, OperationCompte
)

   
@login_required
def rapports_admin(request):
    """Page principale des rapports et gestion"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, "Vous n'etes pas autorise.")
        return redirect('login')
    
    # ========== RECUPERATION DES UTILISATEURS ==========
    admins = Admin.objects.all()
    agents = Agent.objects.all()
    assistants = Assistant.objects.filter(est_actif=True)
    
    total_users = admins.count() + agents.filter(est_actif=True).count() + assistants.count()
    agents_actifs_count = agents.filter(est_actif=True).count()
    
    today = timezone.now().date()
    default_date_debut = today - timedelta(days=30)
    
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    if date_debut_str and date_fin_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            date_debut = default_date_debut
            date_fin = today
    else:
        date_debut = default_date_debut
        date_fin = today
    
    transactions = Transaction.objects.filter(
        date__date__gte=date_debut,
        date__date__lte=date_fin
    )
    
    total_transactions = transactions.count()
    total_volume = int(transactions.aggregate(total=Sum('montant'))['total'] or 0)
    total_commission = int(transactions.aggregate(total=Sum('commission'))['total'] or 0)
    
    # ========== FILTRAGE POUR L'HISTORIQUE ==========
    filtre_type = request.GET.get('filtre_type', 'toutes')
    filtre_compte = request.GET.get('filtre_compte', 'tous')
    filtre_date_debut = request.GET.get('filtre_date_debut', '')
    filtre_date_fin = request.GET.get('filtre_date_fin', '')
    
    # ========== GESTION DE CAISSE ADMIN UNIFIEE ==========
    # Encaissement (Depot) - avec selection du compte
    if request.method == 'POST' and 'encaissement' in request.POST:
        montant = request.POST.get('montant_encaissement')
        description = request.POST.get('description_encaissement', 'Encaissement')
        compte_concerne = request.POST.get('compte_concerne', 'cash')
        
        try:
            montant = int(montant)
            if montant > 0:
                caisse_admin, created = Caisse.objects.get_or_create(user=request.user)
                
                if compte_concerne == 'cash':
                    caisse_admin.solde_cash += montant
                    caisse_admin.save()
                    OperationCaisse.objects.create(
                        caisse=caisse_admin,
                        type_operation='encaissement',
                        montant=montant,
                        description=description,
                        user=request.user
                    )
                    messages.success(request, f'Encaissement de {montant:,.0f} FCFA sur Espèces effectue')
                    
                elif compte_concerne == 'uv_touchpoint':
                    caisse_admin.solde_uv += montant
                    caisse_admin.save()
                    OperationUv.objects.create(
                        caisse=caisse_admin,
                        type_operation='ajout',
                        type_uv='touchpoint',
                        montant=montant,
                        description=description,
                        user=request.user
                    )
                    messages.success(request, f'Ajout de {montant:,.0f} FCFA sur UV Touchpoint effectue')
                    
                elif compte_concerne == 'uv_wave':
                    caisse_admin.solde_wave += montant
                    caisse_admin.save()
                    OperationUv.objects.create(
                        caisse=caisse_admin,
                        type_operation='ajout',
                        type_uv='wave',
                        montant=montant,
                        description=description,
                        user=request.user
                    )
                    messages.success(request, f'Ajout de {montant:,.0f} FCFA sur UV Wave effectue')
                    
                elif compte_concerne == 'epargne':
                    compte_epargne, created = CompteEpargneAdmin.objects.get_or_create(
                        user=request.user,
                        defaults={'solde': 0, 'titulaire': request.user.username}
                    )
                    compte_epargne.solde += montant
                    compte_epargne.save()
                    OperationEpargne.objects.create(
                        compte=compte_epargne,
                        type_operation='depot',
                        montant=montant,
                        description=description
                    )
                    messages.success(request, f'Ajout de {montant:,.0f} FCFA sur Epargne effectue')
            else:
                messages.error(request, 'Montant invalide')
        except ValueError:
            messages.error(request, 'Montant invalide')
        return redirect('rapports_admin')
    
    # Decaissement (Retrait) - avec selection du compte
    if request.method == 'POST' and 'decaissement' in request.POST:
        montant = request.POST.get('montant_decaissement')
        description = request.POST.get('description_decaissement', 'Decaissement')
        compte_concerne = request.POST.get('compte_concerne', 'cash')
        
        try:
            montant = int(montant)
            if montant > 0:
                caisse_admin, created = Caisse.objects.get_or_create(user=request.user)
                erreur = False
                
                if compte_concerne == 'cash':
                    if montant <= caisse_admin.solde_cash:
                        caisse_admin.solde_cash -= montant
                        caisse_admin.save()
                        OperationCaisse.objects.create(
                            caisse=caisse_admin,
                            type_operation='decaissement',
                            montant=montant,
                            description=description,
                            user=request.user
                        )
                        messages.success(request, f'Decaissement de {montant:,.0f} FCFA sur Espèces effectue')
                    else:
                        messages.error(request, 'Solde Espèces insuffisant')
                        erreur = True
                        
                elif compte_concerne == 'uv_touchpoint':
                    if montant <= caisse_admin.solde_uv:
                        caisse_admin.solde_uv -= montant
                        caisse_admin.save()
                        OperationUv.objects.create(
                            caisse=caisse_admin,
                            type_operation='retrait',
                            type_uv='touchpoint',
                            montant=montant,
                            description=description,
                            user=request.user
                        )
                        messages.success(request, f'Retrait de {montant:,.0f} FCFA sur UV Touchpoint effectue')
                    else:
                        messages.error(request, 'Solde UV Touchpoint insuffisant')
                        erreur = True
                        
                elif compte_concerne == 'uv_wave':
                    if montant <= caisse_admin.solde_wave:
                        caisse_admin.solde_wave -= montant
                        caisse_admin.save()
                        OperationUv.objects.create(
                            caisse=caisse_admin,
                            type_operation='retrait',
                            type_uv='wave',
                            montant=montant,
                            description=description,
                            user=request.user
                        )
                        messages.success(request, f'Retrait de {montant:,.0f} FCFA sur UV Wave effectue')
                    else:
                        messages.error(request, 'Solde UV Wave insuffisant')
                        erreur = True
                        
                elif compte_concerne == 'epargne':
                    compte_epargne = CompteEpargneAdmin.objects.filter(user=request.user).first()
                    if compte_epargne and montant <= compte_epargne.solde:
                        compte_epargne.solde -= montant
                        compte_epargne.save()
                        OperationEpargne.objects.create(
                            compte=compte_epargne,
                            type_operation='retrait',
                            montant=montant,
                            description=description
                        )
                        messages.success(request, f'Retrait de {montant:,.0f} FCFA sur Epargne effectue')
                    else:
                        messages.error(request, 'Solde Epargne insuffisant')
                        erreur = True
                
                if not erreur:
                    pass
            else:
                messages.error(request, 'Montant invalide')
        except ValueError:
            messages.error(request, 'Montant invalide')
        return redirect('rapports_admin')
    
    # ========== PERFORMANCE PAR UTILISATEUR ==========
    user_performance = []
    
    for agent in agents:
        agent_transactions = transactions.filter(user=agent.user)
        volume = int(agent_transactions.aggregate(total=Sum('montant'))['total'] or 0)
        if volume > 0:
            user_performance.append({'nom': agent.nom, 'montant': volume, 'id': agent.id, 'type': 'agent'})
    
    for assistant in assistants:
        assistant_transactions = transactions.filter(user=assistant.user)
        volume = int(assistant_transactions.aggregate(total=Sum('montant'))['total'] or 0)
        if volume > 0:
            user_performance.append({'nom': assistant.nom, 'montant': volume, 'id': assistant.id, 'type': 'assistant'})
    
    for admin_user in admins:
        admin_transactions = transactions.filter(user=admin_user.user)
        volume = int(admin_transactions.aggregate(total=Sum('montant'))['total'] or 0)
        if volume > 0:
            user_performance.append({'nom': admin_user.nom, 'montant': volume, 'id': admin_user.id, 'type': 'admin'})
    
    user_performance = sorted(user_performance, key=lambda x: x['montant'], reverse=True)[:5]
    max_volume = user_performance[0]['montant'] if user_performance else 1
    
    for up in user_performance:
        up['percentage'] = int((up['montant'] / max_volume * 100)) if max_volume > 0 else 0
    
    demandes_attente = DemandeApprovisionnement.objects.filter(statut='en_attente')
    
    stats_today = {
        'nombre': Transaction.objects.filter(date__date=today).count(),
        'depots': int(Transaction.objects.filter(date__date=today, type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
        'retraits': int(Transaction.objects.filter(date__date=today, type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
        'credits': int(Transaction.objects.filter(date__date=today, type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0),
    }
    
    stats_yesterday = {'nombre': Transaction.objects.filter(date__date=today - timedelta(days=1)).count()}
    evolution = {'nombre': ((stats_today['nombre'] - stats_yesterday['nombre']) / stats_yesterday['nombre'] * 100) if stats_yesterday['nombre'] > 0 else 0}
    
    top_users = []
    
    for agent in agents.filter(est_actif=True):
        trans_aujourdhui = Transaction.objects.filter(user=agent.user, date__date=today)
        count = trans_aujourdhui.count()
        total = int(trans_aujourdhui.aggregate(Sum('montant'))['montant__sum'] or 0)
        if count > 0:
            top_users.append({'user': agent, 'type': 'agent', 'transactions': count, 'montant': total})
    
    for assistant in assistants:
        trans_aujourdhui = Transaction.objects.filter(user=assistant.user, date__date=today)
        count = trans_aujourdhui.count()
        total = int(trans_aujourdhui.aggregate(Sum('montant'))['montant__sum'] or 0)
        if count > 0:
            top_users.append({'user': assistant, 'type': 'assistant', 'transactions': count, 'montant': total})
    
    top_users = sorted(top_users, key=lambda x: x['transactions'], reverse=True)[:5]
    dernieres_transactions = Transaction.objects.all().order_by('-date')[:10]
    
    # Caisse de l'admin
    try:
        caisse = Caisse.objects.get(user=request.user)
        caisse.solde_cash = int(caisse.solde_cash or 0)
        caisse.solde_uv = int(caisse.solde_uv or 0)
        caisse.solde_wave = int(caisse.solde_wave or 0)
    except Caisse.DoesNotExist:
        caisse = Caisse.objects.create(
            user=request.user,
            solde_cash=0,
            solde_uv=0,
            solde_wave=0
        )
    
    # Compte epargne
    compte_epargne, created = CompteEpargneAdmin.objects.get_or_create(
        user=request.user,
        defaults={'solde': 0, 'titulaire': request.user.username}
    )
    compte_epargne.solde = int(compte_epargne.solde or 0)
    
    # ========== CALCUL DES TOTAUX D'ENCAISSEMENT ET DECAISSEMENT ==========
    # Totaux sur les operations de caisse
    total_encaissements_caisse = OperationCaisse.objects.filter(
        user=request.user, 
        type_operation='encaissement'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    total_decaissements_caisse = OperationCaisse.objects.filter(
        user=request.user, 
        type_operation='decaissement'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Totaux sur les operations UV (ajouts = encaissements, retraits = decaissements)
    total_ajouts_uv = OperationUv.objects.filter(
        user=request.user, 
        type_operation='ajout'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    total_retraits_uv = OperationUv.objects.filter(
        user=request.user, 
        type_operation='retrait'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Totaux sur les operations Epargne (depot = encaissements, retrait = decaissements)
    total_depots_epargne = OperationEpargne.objects.filter(
        compte=compte_epargne, 
        type_operation='depot'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    total_retraits_epargne = OperationEpargne.objects.filter(
        compte=compte_epargne, 
        type_operation='retrait'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Totaux generaux
    total_encaissements = int(total_encaissements_caisse + total_ajouts_uv + total_depots_epargne)
    total_decaissements = int(total_decaissements_caisse + total_retraits_uv + total_retraits_epargne)
    
    # ========== RECUPERATION DE L'HISTORIQUE ==========
    operations = []
    
    # Operations caisse
    queryset_caisse = OperationCaisse.objects.filter(user=request.user)
    for op in queryset_caisse.order_by('-date_operation')[:50]:
        operations.append({
            'date': op.date_operation.strftime('%d/%m/%Y %H:%M:%S'),
            'compte_icon': '💰',
            'compte_nom': 'Especes',
            'type': 'Encaissement' if op.type_operation == 'encaissement' else 'Decaissement',
            'signe': '+' if op.type_operation == 'encaissement' else '-',
            'montant': f'{op.montant:,.0f}',
            'couleur': 'color:#10b981' if op.type_operation == 'encaissement' else 'color:#ef4444',
            'description': op.description or '-'
        })
    
    # Operations UV
    for op in OperationUv.objects.filter(user=request.user).order_by('-date_operation')[:50]:
        operations.append({
            'date': op.date_operation.strftime('%d/%m/%Y %H:%M:%S'),
            'compte_icon': '📱' if op.type_uv == 'touchpoint' else '💳',
            'compte_nom': 'UV Touchpoint' if op.type_uv == 'touchpoint' else 'UV Wave',
            'type': 'Ajout' if op.type_operation == 'ajout' else 'Retrait',
            'signe': '+' if op.type_operation == 'ajout' else '-',
            'montant': f'{op.montant:,.0f}',
            'couleur': 'color:#10b981' if op.type_operation == 'ajout' else 'color:#ef4444',
            'description': op.description or '-'
        })
    
    # Operations Epargne
    if compte_epargne:
        for op in OperationEpargne.objects.filter(compte=compte_epargne).order_by('-date_operation')[:50]:
            operations.append({
                'date': op.date_operation.strftime('%d/%m/%Y %H:%M:%S'),
                'compte_icon': '🏦',
                'compte_nom': 'Epargne',
                'type': 'Ajout' if op.type_operation == 'depot' else 'Retrait',
                'signe': '+' if op.type_operation == 'depot' else '-',
                'montant': f'{op.montant:,.0f}',
                'couleur': 'color:#10b981' if op.type_operation == 'depot' else 'color:#ef4444',
                'description': op.description or '-'
            })
    
    # Trier par date decroissante
    operations.sort(key=lambda x: x['date'], reverse=True)
    
    # Totaux generaux des comptes de l'admin connecte
    total_cash = int(caisse.solde_cash or 0)
    total_uv = int(caisse.solde_uv or 0)
    total_wave = int(caisse.solde_wave or 0)
    total_general = int(total_cash + total_uv + total_wave + (compte_epargne.solde or 0))
    
    context = {
        'title': 'Rapports et Gestion',
        'admin': admin,
        'caisse': caisse,
        'compte_epargne': compte_epargne,
        'operations': operations[:50],
        'total_encaissements': total_encaissements,
        'total_decaissements': total_decaissements,
        'admins': admins,
        'agents': agents,
        'assistants': assistants,
        'agents_actifs_count': agents_actifs_count,
        'total_users': total_users,
        'stats_today': stats_today,
        'evolution': evolution,
        'demandes_attente': demandes_attente,
        'top_agents': top_users,
        'top_users': top_users,
        'dernieres_transactions': dernieres_transactions,
        'total_transactions': total_transactions,
        'total_volume': total_volume,
        'total_commission': total_commission,
        'evolution_transactions': evolution['nombre'],
        'top_agents_performance': user_performance,
        'date_debut': date_debut.strftime('%Y-%m-%d'),
        'date_fin': date_fin.strftime('%Y-%m-%d'),
        'total_general': total_general,
        'filtre_type': filtre_type,
        'filtre_compte': filtre_compte,
        'filtre_date_debut': filtre_date_debut,
        'filtre_date_fin': filtre_date_fin,
    }
    return render(request, 'transactions/rapports_admin.html', context)


@login_required
def api_historique_operations(request):
    """API pour recuperer l'historique des operations"""
    from django.core.paginator import Paginator
    from django.http import JsonResponse
    from datetime import datetime
    
    page = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 10))
    
    operations = []
    
    # 1. Opérations Caisse
    for op in OperationCaisse.objects.filter(user=request.user).order_by('-date_operation'):
        operations.append({
            'date': op.date_operation.strftime('%d/%m/%Y %H:%M:%S'),
            'compte_icon': '💰',
            'compte_nom': 'Espèces',
            'type': 'Encaissement' if op.type_operation == 'encaissement' else 'Décaissement',
            'signe': '+' if op.type_operation == 'encaissement' else '-',
            'montant': f"{op.montant:,.0f}",
            'couleur': '#10b981' if op.type_operation == 'encaissement' else '#ef4444',
            'description': op.description or '-'
        })
    
    # 2. Opérations UV
    for op in OperationUv.objects.filter(user=request.user).order_by('-date_operation'):
        operations.append({
            'date': op.date_operation.strftime('%d/%m/%Y %H:%M:%S'),
            'compte_icon': '📱' if op.type_uv == 'touchpoint' else '💳',
            'compte_nom': 'UV Touchpoint' if op.type_uv == 'touchpoint' else 'UV Wave',
            'type': 'Ajout' if op.type_operation == 'ajout' else 'Retrait',
            'signe': '+' if op.type_operation == 'ajout' else '-',
            'montant': f"{op.montant:,.0f}",
            'couleur': '#10b981' if op.type_operation == 'ajout' else '#ef4444',
            'description': op.description or '-'
        })
    
    # 3. Opérations Epargne
    compte_epargne = CompteEpargneAdmin.objects.filter(user=request.user).first()
    if compte_epargne:
        for op in OperationEpargne.objects.filter(compte=compte_epargne).order_by('-date_operation'):
            operations.append({
                'date': op.date_operation.strftime('%d/%m/%Y %H:%M:%S'),
                'compte_icon': '🏦',
                'compte_nom': 'Epargne',
                'type': 'Dépôt' if op.type_operation == 'depot' else 'Retrait',
                'signe': '+' if op.type_operation == 'depot' else '-',
                'montant': f"{op.montant:,.0f}",
                'couleur': '#10b981' if op.type_operation == 'depot' else '#ef4444',
                'description': op.description or '-'
            })
    
    # Trier par date
    operations.sort(key=lambda x: x['date'], reverse=True)
    
    # Calculer les totaux
    total_encaissements = sum(int(op['montant'].replace(',', '').replace(' ', '')) for op in operations if op['signe'] == '+')
    total_decaissements = sum(int(op['montant'].replace(',', '').replace(' ', '')) for op in operations if op['signe'] == '-')
    
    # Pagination
    paginator = Paginator(operations, per_page)
    try:
        page_obj = paginator.page(page)
    except:
        page_obj = paginator.page(1)
    
    return JsonResponse({
        'success': True,
        'operations': list(page_obj),
        'total': paginator.count,
        'page': page_obj.number,
        'total_pages': paginator.num_pages,
        'has_next': page_obj.has_next(),
        'has_previous': page_obj.has_previous(),
        'per_page': per_page,
        'total_encaissements': f"{total_encaissements:,.0f}",
        'total_decaissements': f"{total_decaissements:,.0f}"
    })


@login_required
def api_totaux_operations(request):
    """API pour recuperer les totaux d'encaissements et decaissements"""
    from django.db.models import Sum
    from django.http import JsonResponse
    
    # Totaux caisse
    total_encaissements_caisse = OperationCaisse.objects.filter(
        user=request.user, 
        type_operation='encaissement'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    total_decaissements_caisse = OperationCaisse.objects.filter(
        user=request.user, 
        type_operation='decaissement'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Totaux UV
    total_ajouts_uv = OperationUv.objects.filter(
        user=request.user, 
        type_operation='ajout'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    total_retraits_uv = OperationUv.objects.filter(
        user=request.user, 
        type_operation='retrait'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Totaux Epargne
    compte_epargne = CompteEpargneAdmin.objects.filter(user=request.user).first()
    total_depots_epargne = 0
    total_retraits_epargne = 0
    
    if compte_epargne:
        total_depots_epargne = OperationEpargne.objects.filter(
            compte=compte_epargne, 
            type_operation='depot'
        ).aggregate(total=Sum('montant'))['total'] or 0
        
        total_retraits_epargne = OperationEpargne.objects.filter(
            compte=compte_epargne, 
            type_operation='retrait'
        ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Totaux generaux
    total_encaissements = int(total_encaissements_caisse + total_ajouts_uv + total_depots_epargne)
    total_decaissements = int(total_decaissements_caisse + total_retraits_uv + total_retraits_epargne)
    
    return JsonResponse({
        'total_encaissements': total_encaissements,
        'total_decaissements': total_decaissements
    })

@login_required
def generer_rapport_admin(request):
    """
    RAPPORT ADMINISTRATEUR - Version Finale
    =======================================
    Contient :
    1. Onglet "RECAP ADMIN" - Totaux généraux + Stats Admin + Soldes Admin + Liste des Agents + Liste des Assistants
    2. Onglet "TRANSACTIONS" - Toutes les transactions
    3. Onglet "DEMANDES" - Toutes les demandes
    4. Onglets individuels pour chaque Agent
    5. Onglets individuels pour chaque Assistant
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.utils.timezone import now
    from datetime import datetime, timedelta
    from django.db.models import Sum, Q
    from transactions.models import Transaction, Agent, Assistant, Admin, Caisse, DemandeApprovisionnement
    from django.contrib.auth.models import User
    
    # ==================== PARAMETRES ====================
    format_type = request.GET.get('format', 'excel')
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    today = now().date()
    
    if date_debut_str and date_fin_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            date_debut = today - timedelta(days=30)
            date_fin = today
    else:
        date_debut = today - timedelta(days=30)
        date_fin = today
    
    # ==================== STYLES EXCEL ====================
    COLOR_PRIMARY = "1a73e8"
    COLOR_DARK = "202124"
    COLOR_SUCCESS = "28a745"
    COLOR_DANGER = "dc3545"
    
    font_title = Font(bold=True, size=16, color="FFFFFF")
    font_subtitle = Font(bold=True, size=12, color=COLOR_PRIMARY)
    font_header = Font(bold=True, size=10, color="FFFFFF")
    font_bold = Font(bold=True)
    font_money = Font(bold=True, size=11, color=COLOR_SUCCESS)
    font_entree = Font(bold=True, size=14, color=COLOR_SUCCESS)
    font_sortie = Font(bold=True, size=14, color=COLOR_DANGER)
    
    fill_title = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
    fill_header = PatternFill(start_color=COLOR_DARK, end_color=COLOR_DARK, fill_type="solid")
    fill_success = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    fill_warning = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    fill_info = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
    fill_entree = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    fill_sortie = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    fill_admin = PatternFill(start_color="e2f0d9", end_color="e2f0d9", fill_type="solid")
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    # ==================== COLLECTE DES DONNEES AVEC FILTRES DATES ====================
    # IMPORTANT: Toutes les transactions sont filtrées par la période
    all_transactions = Transaction.objects.filter(
        date__date__gte=date_debut,
        date__date__lte=date_fin
    ).select_related('user')
    
    # Totaux généraux (déjà filtrés par date)
    total_entree = int(all_transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
    total_sortie = int(all_transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
    total_commission = int(all_transactions.aggregate(Sum('commission'))['commission__sum'] or 0)
    total_transactions = all_transactions.count()
    
    # ==================== STATISTIQUES ADMIN ====================
    admin_stats = {
        'nom': "ADMINISTRATEUR",
        'solde_cash': 0,
        'solde_uv': 0,
        'solde_wave': 0,
        'solde_cash_hier': 0,
        'solde_uv_hier': 0,
        'solde_wave_hier': 0,
        'total_entree': 0,
        'total_sortie': 0,
        'commission': 0,
        'total_transactions': 0,
    }
    
    try:
        admin_user = User.objects.filter(is_superuser=True).first()
        if admin_user:
            caisse_admin = Caisse.objects.get(user=admin_user)
            admin_stats['solde_cash'] = int(caisse_admin.solde_cash or 0)
            admin_stats['solde_uv'] = int(caisse_admin.solde_uv or 0)
            admin_stats['solde_wave'] = int(caisse_admin.solde_wave or 0)
            
            # Transactions de l'admin sur la période
            admin_transactions = all_transactions.filter(user=admin_user)
            admin_stats['total_entree'] = int(admin_transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
            admin_stats['total_sortie'] = int(admin_transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
            admin_stats['commission'] = int(admin_transactions.aggregate(Sum('commission'))['commission__sum'] or 0)
            admin_stats['total_transactions'] = admin_transactions.count()
            
            # Calcul des soldes d'hier (basé sur les transactions d'aujourd'hui)
            transactions_auj = Transaction.objects.filter(date__date=today)
            cash_depot = int(transactions_auj.filter(type_transaction='depot', user=admin_user).aggregate(Sum('montant'))['montant__sum'] or 0)
            cash_retrait = int(transactions_auj.filter(type_transaction='retrait', user=admin_user).aggregate(Sum('montant'))['montant__sum'] or 0)
            admin_stats['solde_cash_hier'] = admin_stats['solde_cash'] - (cash_depot - cash_retrait)
            
            uv_depot = int(transactions_auj.filter(operateur__in=['orange','malitel','telecel'], type_transaction='depot', user=admin_user).aggregate(Sum('montant'))['montant__sum'] or 0)
            uv_retrait = int(transactions_auj.filter(operateur__in=['orange','malitel','telecel'], type_transaction='retrait', user=admin_user).aggregate(Sum('montant'))['montant__sum'] or 0)
            admin_stats['solde_uv_hier'] = admin_stats['solde_uv'] - (uv_retrait - uv_depot)
            
            wave_depot = int(transactions_auj.filter(operateur='wave', type_transaction='depot', user=admin_user).aggregate(Sum('montant'))['montant__sum'] or 0)
            wave_retrait = int(transactions_auj.filter(operateur='wave', type_transaction='retrait', user=admin_user).aggregate(Sum('montant'))['montant__sum'] or 0)
            admin_stats['solde_wave_hier'] = admin_stats['solde_wave'] - (wave_retrait - wave_depot)
    except Exception as e:
        print(f"Erreur admin stats: {e}")
    
    # Statistiques par opérateur (filtrées par date)
    stats = {
        'orange': {
            'depot': int(all_transactions.filter(operateur='orange', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
            'retrait': int(all_transactions.filter(operateur='orange', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
            'credit': int(all_transactions.filter(operateur='orange', type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0),
        },
        'malitel': {
            'depot': int(all_transactions.filter(operateur='malitel', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
            'retrait': int(all_transactions.filter(operateur='malitel', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
            'credit': int(all_transactions.filter(operateur='malitel', type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0),
        },
        'telecel': {
            'depot': int(all_transactions.filter(operateur='telecel', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
            'retrait': int(all_transactions.filter(operateur='telecel', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
            'credit': int(all_transactions.filter(operateur='telecel', type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0),
        },
        'wave': {
            'depot': int(all_transactions.filter(operateur='wave', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
            'retrait': int(all_transactions.filter(operateur='wave', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
        },
    }
    
    # Demandes filtrées par date
    all_demandes = DemandeApprovisionnement.objects.filter(
        date_demande__date__gte=date_debut,
        date_demande__date__lte=date_fin
    ).order_by('-date_demande')
    
    # ==================== COLLECTE DES AGENTS (avec filtres date) ====================
    agents_data = []
    
    for agent in Agent.objects.all():
        # Transactions de l'agent sur la période UNIQUEMENT
        agent_transactions = all_transactions.filter(user=agent.user)
        
        try:
            caisse = Caisse.objects.get(user=agent.user)
            solde_cash = int(caisse.solde_cash or 0)
            solde_uv = int(caisse.solde_uv or 0)
            solde_wave = int(caisse.solde_wave or 0)
        except Caisse.DoesNotExist:
            solde_cash = solde_uv = solde_wave = 0
        
        # Soldes d'hier (basé sur les transactions d'aujourd'hui seulement)
        transactions_auj = Transaction.objects.filter(date__date=today, user=agent.user)
        cash_depot = int(transactions_auj.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        cash_retrait = int(transactions_auj.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        solde_cash_hier = solde_cash - (cash_depot - cash_retrait)
        
        uv_depot = int(transactions_auj.filter(operateur__in=['orange','malitel','telecel'], type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        uv_retrait = int(transactions_auj.filter(operateur__in=['orange','malitel','telecel'], type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        uv_credit = int(transactions_auj.filter(operateur__in=['orange','malitel','telecel'], type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0)
        solde_uv_hier = solde_uv - (uv_retrait - uv_depot - uv_credit)
        
        wave_depot = int(transactions_auj.filter(operateur='wave', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        wave_retrait = int(transactions_auj.filter(operateur='wave', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        solde_wave_hier = solde_wave - (wave_retrait - wave_depot)
        
        # Stats sur la période
        total_entree_agent = int(agent_transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        total_sortie_agent = int(agent_transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        commission_agent = int(agent_transactions.aggregate(Sum('commission'))['commission__sum'] or 0)
        
        agents_data.append({
            'nom': agent.nom,
            'telephone': agent.telephone or '-',
            'email': agent.user.email if agent.user else '-',
            'solde_cash': solde_cash,
            'solde_cash_hier': solde_cash_hier,
            'solde_uv': solde_uv,
            'solde_uv_hier': solde_uv_hier,
            'solde_wave': solde_wave,
            'solde_wave_hier': solde_wave_hier,
            'total_entree': total_entree_agent,
            'total_sortie': total_sortie_agent,
            'commission': commission_agent,
            'total_transactions': agent_transactions.count(),
            'transactions': agent_transactions,
            'demandes': all_demandes.filter(agent=agent),
        })
    
    # ==================== COLLECTE DES ASSISTANTS (avec filtres date) ====================
    assistants_data = []
    
    for assistant in Assistant.objects.all():
        # Transactions de l'assistant sur la période UNIQUEMENT
        assistant_transactions = all_transactions.filter(user=assistant.user)
        
        total_entree_assistant = int(assistant_transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        total_sortie_assistant = int(assistant_transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        commission_assistant = int(assistant_transactions.aggregate(Sum('commission'))['commission__sum'] or 0)
        
        assistants_data.append({
            'nom': assistant.nom,
            'telephone': assistant.telephone or '-',
            'email': assistant.user.email if assistant.user else '-',
            'total_entree': total_entree_assistant,
            'total_sortie': total_sortie_assistant,
            'commission': commission_assistant,
            'total_transactions': assistant_transactions.count(),
            'transactions': assistant_transactions,
            'demandes': all_demandes.filter(assistant_destinataire=assistant),
        })
    
    # ==================== EXPORT EXCEL ====================
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="rapport_admin_{date_debut}_{date_fin}.xlsx"'
    
    wb = Workbook()
    
    # ==================== ONGLET 1: RECAP ADMIN ====================
    ws = wb.active
    ws.title = "1. RECAP ADMIN"
    
    # Titre
    ws.merge_cells('A1:H1')
    ws['A1'] = "📊 RAPPORT ADMINISTRATEUR"
    ws['A1'].font = font_title
    ws['A1'].fill = fill_title
    ws['A1'].alignment = align_center
    
    ws['A2'] = f"📅 Periode: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    ws['A2'].font = font_bold
    ws['A3'] = f"⏰ Date d'export: {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}"
    
    current_row = 5
    
    # ========== SECTION 1: TOTAUX GENERAUX ==========
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "💰 TOTAUX GENERAUX"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_info
    current_row += 1
    
    ws[f'A{current_row}'] = "📈 TOTAL ENTREES"
    ws[f'A{current_row}'].font = font_entree
    ws[f'B{current_row}'] = f"{total_entree:,.0f} FCFA"
    ws[f'B{current_row}'].font = font_entree
    ws[f'B{current_row}'].fill = fill_entree
    ws[f'D{current_row}'] = "📉 TOTAL SORTIES"
    ws[f'D{current_row}'].font = font_sortie
    ws[f'E{current_row}'] = f"{total_sortie:,.0f} FCFA"
    ws[f'E{current_row}'].font = font_sortie
    ws[f'E{current_row}'].fill = fill_sortie
    current_row += 1
    
    ws[f'A{current_row}'] = "🎯 TOTAL COMMISSION"
    ws[f'A{current_row}'].font = font_bold
    ws[f'B{current_row}'] = f"{total_commission:,.0f} FCFA"
    ws[f'B{current_row}'].font = font_money
    ws[f'D{current_row}'] = "📋 NOMBRE DE TRANSACTIONS"
    ws[f'D{current_row}'].font = font_bold
    ws[f'E{current_row}'] = total_transactions
    current_row += 2
    
    # ========== SECTION 2: SOLDES ET ACTIVITE DE L'ADMIN ==========
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "👨‍💼 ADMINISTRATEUR"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_admin
    current_row += 1
    
    # Sous-section: Soldes
    ws[f'A{current_row}'] = "💰 SOLDES"
    ws[f'A{current_row}'].font = font_bold
    current_row += 1
    
    ws[f'A{current_row}'] = "Compte"
    ws[f'B{current_row}'] = "Solde Aujourd'hui"
    ws[f'C{current_row}'] = "Solde Hier"
    ws[f'D{current_row}'] = "Variation"
    for col in range(1, 5):
        ws.cell(row=current_row, column=col).font = font_header
        ws.cell(row=current_row, column=col).fill = fill_header
    current_row += 1
    
    ws[f'A{current_row}'] = "💵 Cash (Espèces)"
    ws[f'B{current_row}'] = f"{admin_stats['solde_cash']:,.0f} FCFA"
    ws[f'C{current_row}'] = f"{admin_stats['solde_cash_hier']:,.0f} FCFA"
    ws[f'D{current_row}'] = f"{admin_stats['solde_cash'] - admin_stats['solde_cash_hier']:+,.0f} FCFA"
    current_row += 1
    
    ws[f'A{current_row}'] = "📱 UV Touchpiont"
    ws[f'B{current_row}'] = f"{admin_stats['solde_uv']:,.0f} FCFA"
    ws[f'C{current_row}'] = f"{admin_stats['solde_uv_hier']:,.0f} FCFA"
    ws[f'D{current_row}'] = f"{admin_stats['solde_uv'] - admin_stats['solde_uv_hier']:+,.0f} FCFA"
    current_row += 1
    
    ws[f'A{current_row}'] = "🌊 UV Wave"
    ws[f'B{current_row}'] = f"{admin_stats['solde_wave']:,.0f} FCFA"
    ws[f'C{current_row}'] = f"{admin_stats['solde_wave_hier']:,.0f} FCFA"
    ws[f'D{current_row}'] = f"{admin_stats['solde_wave'] - admin_stats['solde_wave_hier']:+,.0f} FCFA"
    current_row += 2
    
    # Sous-section: Activité sur la période
    ws[f'A{current_row}'] = "📊 ACTIVITE SUR LA PERIODE"
    ws[f'A{current_row}'].font = font_bold
    current_row += 1
    
    ws[f'A{current_row}'] = f"💰 Entrées: {admin_stats['total_entree']:,.0f} FCFA"
    ws[f'C{current_row}'] = f"💸 Sorties: {admin_stats['total_sortie']:,.0f} FCFA"
    ws[f'E{current_row}'] = f"🎯 Commission: {admin_stats['commission']:,.0f} FCFA"
    ws[f'G{current_row}'] = f"📋 Nb Ops: {admin_stats['total_transactions']}"
    current_row += 2
    
    # ========== SECTION 3: STATISTIQUES PAR OPERATEUR ==========
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "📱 STATISTIQUES PAR OPERATEUR"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_info
    current_row += 1
    
    headers = ['Opérateur', 'Dépôts', 'Retraits', 'Crédits', 'Total', 'Commission']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    current_row += 1
    
    ops_data = [('Orange', 'orange'), ('Malitel', 'malitel'), ('Telecel', 'telecel'), ('Wave', 'wave')]
    for op_name, op_key in ops_data:
        ws.cell(row=current_row, column=1, value=op_name)
        ws.cell(row=current_row, column=2, value=f"{stats[op_key]['depot']:,.0f} FCFA")
        ws.cell(row=current_row, column=3, value=f"{stats[op_key]['retrait']:,.0f} FCFA")
        ws.cell(row=current_row, column=4, value=f"{stats[op_key]['credit']:,.0f} FCFA" if op_key != 'wave' else "-")
        total = stats[op_key]['depot'] + stats[op_key]['retrait'] + (stats[op_key].get('credit', 0) if op_key != 'wave' else 0)
        ws.cell(row=current_row, column=5, value=f"{total:,.0f} FCFA")
        ws.cell(row=current_row, column=6, value=f"{int(total * 0.01):,.0f} FCFA")
        current_row += 1
    current_row += 2
    
    # ========== SECTION 4: LISTE DES AGENTS ==========
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "👥 LISTE DES AGENTS"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_info
    current_row += 1
    
    for agent in agents_data:
        # Nom de l'agent
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = f"👤 {agent['nom']}"
        ws[f'A{current_row}'].font = font_bold
        ws[f'A{current_row}'].fill = fill_info
        current_row += 1
        
        # Téléphone et email
        ws[f'A{current_row}'] = f"📞 {agent['telephone']}  |  ✉️ {agent['email']}"
        current_row += 1
        
        # En-tête du tableau des soldes
        ws[f'A{current_row}'] = "Compte"
        ws[f'B{current_row}'] = "Solde Aujourd'hui"
        ws[f'C{current_row}'] = "Solde Hier"
        ws[f'D{current_row}'] = "Variation"
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).font = font_header
            ws.cell(row=current_row, column=col).fill = fill_header
        current_row += 1
        
        # Cash
        ws[f'A{current_row}'] = "💵 Cash (Espèces)"
        ws[f'B{current_row}'] = f"{agent['solde_cash']:,.0f} FCFA"
        ws[f'C{current_row}'] = f"{agent['solde_cash_hier']:,.0f} FCFA"
        var_cash = agent['solde_cash'] - agent['solde_cash_hier']
        ws[f'D{current_row}'] = f"{var_cash:+,.0f} FCFA"
        ws[f'D{current_row}'].fill = fill_success if var_cash >= 0 else fill_warning
        current_row += 1
        
        # UV
        ws[f'A{current_row}'] = "📱 UV Touchpiont"
        ws[f'B{current_row}'] = f"{agent['solde_uv']:,.0f} FCFA"
        ws[f'C{current_row}'] = f"{agent['solde_uv_hier']:,.0f} FCFA"
        ws[f'D{current_row}'] = f"{agent['solde_uv'] - agent['solde_uv_hier']:+,.0f} FCFA"
        current_row += 1
        
        # Wave
        ws[f'A{current_row}'] = "🌊 UV Wave"
        ws[f'B{current_row}'] = f"{agent['solde_wave']:,.0f} FCFA"
        ws[f'C{current_row}'] = f"{agent['solde_wave_hier']:,.0f} FCFA"
        ws[f'D{current_row}'] = f"{agent['solde_wave'] - agent['solde_wave_hier']:+,.0f} FCFA"
        current_row += 1
        
        # Résumé période
        ws[f'A{current_row}'] = "📊 RÉSUMÉ PÉRIODE"
        ws[f'B{current_row}'] = f"Entrées: {agent['total_entree']:,.0f} FCFA | Sorties: {agent['total_sortie']:,.0f} FCFA | Commission: {agent['commission']:,.0f} FCFA | Nb Ops: {agent['total_transactions']}"
        current_row += 2
        
        # Séparateur
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "-" * 60
        current_row += 1
    
    # ========== SECTION 5: LISTE DES ASSISTANTS (AJOUTÉE) ==========
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "👥 LISTE DES ASSISTANTS"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_info
    current_row += 1
    
    headers_assistants = ['Assistant', 'Téléphone', 'Email', 'Entrées', 'Sorties', 'Commission', 'Nb Ops']
    for col, header in enumerate(headers_assistants, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    current_row += 1
    
    for assistant in assistants_data:
        ws.cell(row=current_row, column=1, value=assistant['nom'])
        ws.cell(row=current_row, column=2, value=assistant['telephone'])
        ws.cell(row=current_row, column=3, value=assistant['email'])
        ws.cell(row=current_row, column=4, value=f"{assistant['total_entree']:,.0f} FCFA")
        ws.cell(row=current_row, column=5, value=f"{assistant['total_sortie']:,.0f} FCFA")
        ws.cell(row=current_row, column=6, value=f"{assistant['commission']:,.0f} FCFA")
        ws.cell(row=current_row, column=7, value=assistant['total_transactions'])
        current_row += 1
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12
    
    # ==================== ONGLET 2: TRANSACTIONS ====================
    ws2 = wb.create_sheet("2. TRANSACTIONS")
    ws2.merge_cells('A1:H1')
    ws2['A1'] = "DETAIL DES TRANSACTIONS"
    ws2['A1'].font = font_title
    ws2['A1'].fill = fill_title
    ws2['A1'].alignment = align_center
    ws2['A2'] = f"📅 Periode: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    
    row = 4
    headers_trans = ['Date', 'Référence', 'Type', 'Opérateur', 'Client', 'Montant', 'Commission', 'Utilisateur']
    for col, header in enumerate(headers_trans, 1):
        cell = ws2.cell(row=row, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    row += 1
    
    for t in all_transactions.order_by('-date'):
        ws2.cell(row=row, column=1, value=t.date.strftime('%d/%m/%Y %H:%M'))
        ws2.cell(row=row, column=2, value=t.reference)
        ws2.cell(row=row, column=3, value=t.get_type_transaction_display())
        ws2.cell(row=row, column=4, value=t.get_operateur_display())
        ws2.cell(row=row, column=5, value=t.numero_client)
        ws2.cell(row=row, column=6, value=f"{int(t.montant):,.0f}")
        ws2.cell(row=row, column=6).font = font_money
        ws2.cell(row=row, column=7, value=f"{int(t.commission):,.0f}")
        ws2.cell(row=row, column=8, value=t.user.username if t.user else "-")
        row += 1
    
    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 16
    
    # ==================== ONGLET 3: DEMANDES ====================
    if all_demandes.exists():
        ws3 = wb.create_sheet("3. DEMANDES")
        ws3.merge_cells('A1:G1')
        ws3['A1'] = "DEMANDES D'APPROVISIONNEMENT"
        ws3['A1'].font = font_title
        ws3['A1'].fill = fill_title
        ws3['A1'].alignment = align_center
        ws3['A2'] = f"📅 Periode: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        
        row = 4
        headers_dem = ['Date', 'Type', 'Montant', 'Statut', 'Motif', 'Agent', 'Assistant']
        for col, header in enumerate(headers_dem, 1):
            cell = ws3.cell(row=row, column=col, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        row += 1
        
        for d in all_demandes:
            ws3.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y %H:%M'))
            ws3.cell(row=row, column=2, value=d.get_type_echange_display())
            ws3.cell(row=row, column=3, value=f"{int(d.montant):,.0f} FCFA")
            ws3.cell(row=row, column=4, value=d.get_statut_display())
            ws3.cell(row=row, column=5, value=d.motif or "-")
            ws3.cell(row=row, column=6, value=d.agent.nom if d.agent else "-")
            ws3.cell(row=row, column=7, value=d.assistant_destinataire.nom if d.assistant_destinataire else "-")
            row += 1
        
        for col in range(1, 8):
            ws3.column_dimensions[get_column_letter(col)].width = 18
    
    # ==================== ONGLETS INDIVIDUELS POUR CHAQUE AGENT ====================
    for agent in agents_data:
        nom_feuille = f"Agent_{agent['nom'][:20]}".replace(' ', '_').replace('-', '_')
        ws_agent = wb.create_sheet(nom_feuille[:25])
        
        ws_agent.merge_cells('A1:D1')
        ws_agent['A1'] = f"AGENT : {agent['nom']}"
        ws_agent['A1'].font = font_title
        ws_agent['A1'].fill = fill_title
        ws_agent['A1'].alignment = align_center
        
        ws_agent['A2'] = f"📞 {agent['telephone']}  |  ✉️ {agent['email']}"
        ws_agent['A3'] = f"📅 Periode: {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        
        row = 5
        
        # SOLDES
        ws_agent.merge_cells(f'A{row}:D{row}')
        ws_agent[f'A{row}'] = "💰 SOLDES"
        ws_agent[f'A{row}'].font = font_subtitle
        ws_agent[f'A{row}'].fill = fill_info
        row += 1
        
        ws_agent[f'A{row}'] = "Compte"
        ws_agent[f'B{row}'] = "Aujourd'hui"
        ws_agent[f'C{row}'] = "Hier"
        ws_agent[f'D{row}'] = "Variation"
        for col in range(1, 5):
            ws_agent.cell(row=row, column=col).font = font_header
            ws_agent.cell(row=row, column=col).fill = fill_header
        row += 1
        
        ws_agent[f'A{row}'] = "💵 Cash"
        ws_agent[f'B{row}'] = f"{agent['solde_cash']:,.0f} FCFA"
        ws_agent[f'C{row}'] = f"{agent['solde_cash_hier']:,.0f} FCFA"
        ws_agent[f'D{row}'] = f"{agent['solde_cash'] - agent['solde_cash_hier']:+,.0f} FCFA"
        row += 1
        
        ws_agent[f'A{row}'] = "📱 UV Touchpiont"
        ws_agent[f'B{row}'] = f"{agent['solde_uv']:,.0f} FCFA"
        ws_agent[f'C{row}'] = f"{agent['solde_uv_hier']:,.0f} FCFA"
        ws_agent[f'D{row}'] = f"{agent['solde_uv'] - agent['solde_uv_hier']:+,.0f} FCFA"
        row += 1
        
        ws_agent[f'A{row}'] = "🌊 UV Wave"
        ws_agent[f'B{row}'] = f"{agent['solde_wave']:,.0f} FCFA"
        ws_agent[f'C{row}'] = f"{agent['solde_wave_hier']:,.0f} FCFA"
        ws_agent[f'D{row}'] = f"{agent['solde_wave'] - agent['solde_wave_hier']:+,.0f} FCFA"
        row += 2
        
        ws_agent[f'A{row}'] = "📊 RÉSUMÉ PÉRIODE"
        ws_agent[f'A{row}'].font = font_bold
        ws_agent[f'B{row}'] = f"Entrées: {agent['total_entree']:,.0f} FCFA"
        ws_agent[f'C{row}'] = f"Sorties: {agent['total_sortie']:,.0f} FCFA"
        ws_agent[f'D{row}'] = f"Commission: {agent['commission']:,.0f} FCFA"
        row += 2
        
        # DEMANDES
        if agent['demandes'].exists():
            ws_agent.merge_cells(f'A{row}:D{row}')
            ws_agent[f'A{row}'] = "📨 DEMANDES"
            ws_agent[f'A{row}'].font = font_subtitle
            ws_agent[f'A{row}'].fill = fill_info
            row += 1
            
            headers_dem = ['Date', 'Type', 'Montant', 'Statut']
            for col, header in enumerate(headers_dem, 1):
                cell = ws_agent.cell(row=row, column=col, value=header)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
            row += 1
            
            for d in agent['demandes']:
                ws_agent.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y'))
                ws_agent.cell(row=row, column=2, value=d.get_type_echange_display())
                ws_agent.cell(row=row, column=3, value=f"{int(d.montant):,.0f} FCFA")
                ws_agent.cell(row=row, column=4, value=d.get_statut_display())
                row += 1
        
        # TRANSACTIONS (colonne F à K)
        col_trans = 6
        row_trans = 5
        
        ws_agent.merge_cells(start_row=row_trans, start_column=col_trans, end_row=row_trans, end_column=col_trans+5)
        ws_agent.cell(row=row_trans, column=col_trans, value="📊 TRANSACTIONS SUR LA PÉRIODE")
        ws_agent.cell(row=row_trans, column=col_trans).font = font_subtitle
        ws_agent.cell(row=row_trans, column=col_trans).fill = fill_info
        ws_agent.cell(row=row_trans, column=col_trans).alignment = align_center
        row_trans += 1
        
        headers_trans_pers = ['Date', 'Type', 'Opérateur', 'Client', 'Montant', 'Commission']
        for col, header in enumerate(headers_trans_pers, col_trans):
            cell = ws_agent.cell(row=row_trans, column=col, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        row_trans += 1
        
        for t in agent['transactions'].order_by('-date'):
            ws_agent.cell(row=row_trans, column=col_trans, value=t.date.strftime('%d/%m/%Y'))
            ws_agent.cell(row=row_trans, column=col_trans+1, value=t.get_type_transaction_display())
            ws_agent.cell(row=row_trans, column=col_trans+2, value=t.get_operateur_display())
            ws_agent.cell(row=row_trans, column=col_trans+3, value=t.numero_client)
            ws_agent.cell(row=row_trans, column=col_trans+4, value=f"{int(t.montant):,.0f}")
            ws_agent.cell(row=row_trans, column=col_trans+5, value=f"{int(t.commission):,.0f}")
            row_trans += 1
        
        if agent['transactions'].count() == 0:
            ws_agent.cell(row=row_trans, column=col_trans, value="Aucune transaction sur cette période")
        
        # Ajuster largeurs
        for col in range(1, 5):
            ws_agent.column_dimensions[get_column_letter(col)].width = 20
        for col in range(6, 12):
            ws_agent.column_dimensions[get_column_letter(col)].width = 16
    
    # ==================== ONGLETS INDIVIDUELS POUR CHAQUE ASSISTANT ====================
    for assistant in assistants_data:
        nom_feuille = f"Assistant_{assistant['nom'][:20]}".replace(' ', '_').replace('-', '_')
        ws_assistant = wb.create_sheet(nom_feuille[:25])
        
        ws_assistant.merge_cells('A1:D1')
        ws_assistant['A1'] = f"ASSISTANT : {assistant['nom']}"
        ws_assistant['A1'].font = font_title
        ws_assistant['A1'].fill = fill_title
        ws_assistant['A1'].alignment = align_center
        
        ws_assistant['A2'] = f"📞 {assistant['telephone']}  |  ✉️ {assistant['email']}"
        ws_assistant['A3'] = f"📅 Periode: {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        ws_assistant['A4'] = "ℹ️ Les assistants partagent le solde de l'administrateur"
        ws_assistant['A4'].fill = fill_info
        
        row = 6
        
        # RÉSUMÉ
        ws_assistant.merge_cells(f'A{row}:D{row}')
        ws_assistant[f'A{row}'] = "📊 RÉSUMÉ DES OPÉRATIONS"
        ws_assistant[f'A{row}'].font = font_subtitle
        ws_assistant[f'A{row}'].fill = fill_info
        row += 1
        
        ws_assistant[f'A{row}'] = "💰 Total Entrées"
        ws_assistant[f'B{row}'] = f"{assistant['total_entree']:,.0f} FCFA"
        ws_assistant[f'C{row}'] = "💸 Total Sorties"
        ws_assistant[f'D{row}'] = f"{assistant['total_sortie']:,.0f} FCFA"
        row += 1
        
        ws_assistant[f'A{row}'] = "🎯 Commission totale"
        ws_assistant[f'B{row}'] = f"{assistant['commission']:,.0f} FCFA"
        ws_assistant[f'C{row}'] = "📋 Nombre de transactions"
        ws_assistant[f'D{row}'] = assistant['total_transactions']
        row += 2
        
        # DEMANDES
        if assistant['demandes'].exists():
            ws_assistant.merge_cells(f'A{row}:D{row}')
            ws_assistant[f'A{row}'] = "📨 DEMANDES TRAITÉES"
            ws_assistant[f'A{row}'].font = font_subtitle
            ws_assistant[f'A{row}'].fill = fill_info
            row += 1
            
            headers_dem = ['Date', 'Type', 'Montant', 'Statut']
            for col, header in enumerate(headers_dem, 1):
                cell = ws_assistant.cell(row=row, column=col, value=header)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
            row += 1
            
            for d in assistant['demandes']:
                ws_assistant.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y'))
                ws_assistant.cell(row=row, column=2, value=d.get_type_echange_display())
                ws_assistant.cell(row=row, column=3, value=f"{int(d.montant):,.0f} FCFA")
                ws_assistant.cell(row=row, column=4, value=d.get_statut_display())
                row += 1
        
        # TRANSACTIONS
        col_trans = 6
        row_trans = 6
        
        ws_assistant.merge_cells(start_row=row_trans, start_column=col_trans, end_row=row_trans, end_column=col_trans+5)
        ws_assistant.cell(row=row_trans, column=col_trans, value="📊 TRANSACTIONS SUR LA PÉRIODE")
        ws_assistant.cell(row=row_trans, column=col_trans).font = font_subtitle
        ws_assistant.cell(row=row_trans, column=col_trans).fill = fill_info
        ws_assistant.cell(row=row_trans, column=col_trans).alignment = align_center
        row_trans += 1
        
        headers_trans_pers = ['Date', 'Type', 'Opérateur', 'Client', 'Montant', 'Commission']
        for col, header in enumerate(headers_trans_pers, col_trans):
            cell = ws_assistant.cell(row=row_trans, column=col, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        row_trans += 1
        
        for t in assistant['transactions'].order_by('-date'):
            ws_assistant.cell(row=row_trans, column=col_trans, value=t.date.strftime('%d/%m/%Y'))
            ws_assistant.cell(row=row_trans, column=col_trans+1, value=t.get_type_transaction_display())
            ws_assistant.cell(row=row_trans, column=col_trans+2, value=t.get_operateur_display())
            ws_assistant.cell(row=row_trans, column=col_trans+3, value=t.numero_client)
            ws_assistant.cell(row=row_trans, column=col_trans+4, value=f"{int(t.montant):,.0f}")
            ws_assistant.cell(row=row_trans, column=col_trans+5, value=f"{int(t.commission):,.0f}")
            row_trans += 1
        
        if assistant['transactions'].count() == 0:
            ws_assistant.cell(row=row_trans, column=col_trans, value="Aucune transaction sur cette période")
        
        # Ajuster largeurs
        for col in range(1, 5):
            ws_assistant.column_dimensions[get_column_letter(col)].width = 20
        for col in range(6, 12):
            ws_assistant.column_dimensions[get_column_letter(col)].width = 16
    
    wb.save(response)
    return response
    
def export_complete_report(transactions, nom, user_type, caisse, total_entree, total_sortie, total_commission, demandes, format_type, date_debut, date_fin, stats_orange, stats_malitel, stats_telecel, stats_wave):
    """Export complet avec recap, demandes et transactions"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import csv
    from django.http import HttpResponse
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    today = datetime.now().date()
    
    # Styles pour Excel
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="0f766e", end_color="0f766e", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Calcul des soldes hier si caisse disponible
    solde_cash_hier = 0
    solde_uv_hier = 0
    solde_wave_hier = 0
    
    if caisse and nom != "Tous":
        transactions_today = transactions.filter(date__date=today)
        
        cash_depot_today = int(transactions_today.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        cash_retrait_today = int(transactions_today.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        variation_cash_today = cash_depot_today - cash_retrait_today
        
        uv_depot_today = int(transactions_today.filter(
            operateur__in=['orange', 'malitel', 'telecel'],
            type_transaction='depot'
        ).aggregate(Sum('montant'))['montant__sum'] or 0)
        uv_retrait_today = int(transactions_today.filter(
            operateur__in=['orange', 'malitel', 'telecel'],
            type_transaction='retrait'
        ).aggregate(Sum('montant'))['montant__sum'] or 0)
        uv_credit_today = int(transactions_today.filter(
            operateur__in=['orange', 'malitel', 'telecel'],
            type_transaction='credit'
        ).aggregate(Sum('montant'))['montant__sum'] or 0)
        variation_uv_today = uv_retrait_today - uv_depot_today - uv_credit_today
        
        wave_depot_today = int(transactions_today.filter(
            operateur='wave',
            type_transaction='depot'
        ).aggregate(Sum('montant'))['montant__sum'] or 0)
        wave_retrait_today = int(transactions_today.filter(
            operateur='wave',
            type_transaction='retrait'
        ).aggregate(Sum('montant'))['montant__sum'] or 0)
        variation_wave_today = wave_retrait_today - wave_depot_today
        
        solde_cash_hier = int(caisse.solde_cash - variation_cash_today)
        solde_uv_hier = int(caisse.solde_uv - variation_uv_today)
        solde_wave_hier = int(caisse.solde_wave - variation_wave_today)
    
    type_label = ""
    if user_type == "agent":
        type_label = "Agent"
    elif user_type == "assistant":
        type_label = "Assistant"
    elif user_type == "admin":
        type_label = "Admin"
    
    # ==================== EXPORT CSV ====================
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        if nom != "Tous":
            response['Content-Disposition'] = f'attachment; filename="rapport_{nom}_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
        else:
            response['Content-Disposition'] = f'attachment; filename="rapport_global_{date_debut}_{date_fin}.csv"'
        response.write('\ufeff')
        writer = csv.writer(response)
        
        # En-tete
        if nom != "Tous":
            writer.writerow([f"RAPPORT DETAILLE - {nom} ({type_label})"])
        else:
            writer.writerow(["RAPPORT FINANCIER GLOBAL"])
        writer.writerow([f"Periode: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"])
        writer.writerow([f"Date export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        writer.writerow([])
        
        # SOLDES
        if caisse and nom != "Tous":
            writer.writerow(["=== SOLDES ==="])
            writer.writerow(["Compte", "Solde actuel", "Solde hier", "Variation"])
            writer.writerow(["Argent Cash", f"{caisse.solde_cash:,.0f} FCFA", f"{solde_cash_hier:,.0f} FCFA", f"{caisse.solde_cash - solde_cash_hier:+,.0f} FCFA"])
            writer.writerow(["UV Touspiont", f"{caisse.solde_uv:,.0f} FCFA", f"{solde_uv_hier:,.0f} FCFA", f"{caisse.solde_uv - solde_uv_hier:+,.0f} FCFA"])
            writer.writerow(["UV Wave", f"{caisse.solde_wave:,.0f} FCFA", f"{solde_wave_hier:,.0f} FCFA", f"{caisse.solde_wave - solde_wave_hier:+,.0f} FCFA"])
            writer.writerow([])
        
        # TOTAUX TRANSACTIONS
        writer.writerow(["=== TOTAUX DES TRANSACTIONS ==="])
        writer.writerow(["Total Entrees", f"{total_entree:,.0f} FCFA"])
        writer.writerow(["Total Sorties", f"{total_sortie:,.0f} FCFA"])
        writer.writerow(["Total Commission", f"{total_commission:,.0f} FCFA"])
        writer.writerow(["Nombre de transactions", transactions.count()])
        writer.writerow([])
        
        # STATISTIQUES PAR OPERATEUR
        writer.writerow(["=== STATISTIQUES PAR OPERATEUR ==="])
        writer.writerow(["Operateur", "Depots", "Retraits", "Credits", "Total"])
        writer.writerow(["Orange", f"{stats_orange['depot']:,.0f}", f"{stats_orange['retrait']:,.0f}", f"{stats_orange['credit']:,.0f}", f"{stats_orange['depot'] + stats_orange['retrait'] + stats_orange['credit']:,.0f}"])
        writer.writerow(["Malitel", f"{stats_malitel['depot']:,.0f}", f"{stats_malitel['retrait']:,.0f}", f"{stats_malitel['credit']:,.0f}", f"{stats_malitel['depot'] + stats_malitel['retrait'] + stats_malitel['credit']:,.0f}"])
        writer.writerow(["Telecel", f"{stats_telecel['depot']:,.0f}", f"{stats_telecel['retrait']:,.0f}", f"{stats_telecel['credit']:,.0f}", f"{stats_telecel['depot'] + stats_telecel['retrait'] + stats_telecel['credit']:,.0f}"])
        writer.writerow(["Wave", f"{stats_wave['depot']:,.0f}", f"{stats_wave['retrait']:,.0f}", "0", f"{stats_wave['depot'] + stats_wave['retrait']:,.0f}"])
        writer.writerow([])
        
        # DEMANDES
        if demandes:
            writer.writerow(["=== DEMANDES D'APPROVISIONNEMENT ==="])
            writer.writerow(["Date", "Type", "Montant", "Statut", "Motif", "Agent", "Assistant"])
            for d in demandes:
                writer.writerow([
                    d.date_demande.strftime('%d/%m/%Y %H:%M'),
                    d.get_type_echange_display(),
                    f"{int(d.montant):,.0f} FCFA",
                    d.get_statut_display(),
                    d.motif or "",
                    d.agent.nom if d.agent else "-",
                    d.assistant_destinataire.nom if d.assistant_destinataire else "-"
                ])
            writer.writerow([])
        
        # DETAIL DES TRANSACTIONS
        writer.writerow(["=== DETAIL DES TRANSACTIONS ==="])
        writer.writerow(['Reference', 'Type', 'Operateur', 'Client', 'Montant (FCFA)', 'Commission (FCFA)', 'Date'])
        for t in transactions:
            writer.writerow([
                t.reference,
                t.get_type_transaction_display(),
                t.get_operateur_display(),
                t.numero_client,
                f"{int(t.montant):,}",
                f"{int(t.commission):,}",
                t.date.strftime('%d/%m/%Y %H:%M:%S')
            ])
        
        return response
    
    # ==================== EXPORT EXCEL ====================
    else:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        if nom != "Tous":
            response['Content-Disposition'] = f'attachment; filename="rapport_{nom}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        else:
            response['Content-Disposition'] = f'attachment; filename="rapport_global_{date_debut}_{date_fin}.xlsx"'
        
        wb = Workbook()
        
        # ========== FEUILLE 1: RECAPITULATIF ==========
        ws_summary = wb.active
        ws_summary.title = "Recapitulatif"
        
        # Titre principal
        ws_summary.merge_cells('A1:E1')
        if nom != "Tous":
            ws_summary['A1'] = f"RAPPORT - {nom} ({type_label})"
        else:
            ws_summary['A1'] = "RAPPORT FINANCIER GLOBAL"
        ws_summary['A1'].font = title_font
        ws_summary['A1'].fill = title_fill
        ws_summary['A1'].alignment = center_align
        
        ws_summary['A2'] = f"Periode: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        ws_summary['A3'] = f"Date export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        row = 5
        
        # SECTION SOLDES
        if caisse and nom != "Tous":
            ws_summary[f'A{row}'] = "SOLDES"
            ws_summary[f'A{row}'].font = subheader_font
            row += 1
            
            for col, header in enumerate(['Compte', 'Solde actuel', 'Solde hier', 'Variation'], 1):
                cell = ws_summary.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            
            soldes_data = [
                ["Argent Cash", f"{caisse.solde_cash:,.0f} FCFA", f"{solde_cash_hier:,.0f} FCFA", f"{caisse.solde_cash - solde_cash_hier:+,.0f} FCFA"],
                ["UV Touspiont", f"{caisse.solde_uv:,.0f} FCFA", f"{solde_uv_hier:,.0f} FCFA", f"{caisse.solde_uv - solde_uv_hier:+,.0f} FCFA"],
                ["UV Wave", f"{caisse.solde_wave:,.0f} FCFA", f"{solde_wave_hier:,.0f} FCFA", f"{caisse.solde_wave - solde_wave_hier:+,.0f} FCFA"],
            ]
            for data in soldes_data:
                for col, val in enumerate(data, 1):
                    ws_summary.cell(row=row, column=col, value=val)
                row += 1
            row += 1
        
        # SECTION TOTAUX TRANSACTIONS
        ws_summary[f'A{row}'] = "TOTAUX DES TRANSACTIONS"
        ws_summary[f'A{row}'].font = subheader_font
        row += 1
        
        totals_data = [
            ["Total Entrees", f"{total_entree:,.0f} FCFA"],
            ["Total Sorties", f"{total_sortie:,.0f} FCFA"],
            ["Total Commission", f"{total_commission:,.0f} FCFA"],
            ["Nombre de transactions", transactions.count()],
        ]
        for data in totals_data:
            ws_summary[f'A{row}'] = data[0]
            ws_summary[f'B{row}'] = data[1]
            row += 1
        row += 1
        
        # SECTION STATISTIQUES PAR OPERATEUR
        ws_summary[f'A{row}'] = "STATISTIQUES PAR OPERATEUR"
        ws_summary[f'A{row}'].font = subheader_font
        row += 1
        
        for col, header in enumerate(['Operateur', 'Depots', 'Retraits', 'Credits', 'Total'], 1):
            cell = ws_summary.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1
        
        stats_data = [
            ["Orange", stats_orange['depot'], stats_orange['retrait'], stats_orange['credit'], stats_orange['depot'] + stats_orange['retrait'] + stats_orange['credit']],
            ["Malitel", stats_malitel['depot'], stats_malitel['retrait'], stats_malitel['credit'], stats_malitel['depot'] + stats_malitel['retrait'] + stats_malitel['credit']],
            ["Telecel", stats_telecel['depot'], stats_telecel['retrait'], stats_telecel['credit'], stats_telecel['depot'] + stats_telecel['retrait'] + stats_telecel['credit']],
            ["Wave", stats_wave['depot'], stats_wave['retrait'], 0, stats_wave['depot'] + stats_wave['retrait']],
        ]
        for data in stats_data:
            ws_summary[f'A{row}'] = data[0]
            ws_summary[f'B{row}'] = f"{data[1]:,.0f} FCFA"
            ws_summary[f'C{row}'] = f"{data[2]:,.0f} FCFA"
            ws_summary[f'D{row}'] = f"{data[3]:,.0f} FCFA" if data[3] > 0 else "-"
            ws_summary[f'E{row}'] = f"{data[4]:,.0f} FCFA"
            row += 1
        
        # Ajustement largeurs colonnes
        for col in range(1, 6):
            ws_summary.column_dimensions[get_column_letter(col)].width = 22
        
        # ========== FEUILLE 2: DEMANDES ==========
        if demandes:
            ws_demandes = wb.create_sheet("Demandes")
            
            headers_demandes = ['Date', 'Type', 'Montant', 'Statut', 'Motif', 'Agent', 'Assistant']
            for col, header in enumerate(headers_demandes, 1):
                cell = ws_demandes.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            
            for row, d in enumerate(demandes, 2):
                ws_demandes.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y %H:%M'))
                ws_demandes.cell(row=row, column=2, value=d.get_type_echange_display())
                ws_demandes.cell(row=row, column=3, value=f"{int(d.montant):,.0f} FCFA")
                ws_demandes.cell(row=row, column=4, value=d.get_statut_display())
                ws_demandes.cell(row=row, column=5, value=d.motif or "")
                ws_demandes.cell(row=row, column=6, value=d.agent.nom if d.agent else "-")
                ws_demandes.cell(row=row, column=7, value=d.assistant_destinataire.nom if d.assistant_destinataire else "-")
            
            for col in range(1, 8):
                ws_demandes.column_dimensions[get_column_letter(col)].width = 20
        
        # ========== FEUILLE 3: TRANSACTIONS ==========
        ws_trans = wb.create_sheet("Transactions")
        
        headers_trans = ['Reference', 'Type', 'Operateur', 'Client', 'Montant (FCFA)', 'Commission (FCFA)', 'Date']
        for col, header in enumerate(headers_trans, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        for row, t in enumerate(transactions, 2):
            ws_trans.cell(row=row, column=1, value=t.reference)
            ws_trans.cell(row=row, column=2, value=t.get_type_transaction_display())
            ws_trans.cell(row=row, column=3, value=t.get_operateur_display())
            ws_trans.cell(row=row, column=4, value=t.numero_client)
            ws_trans.cell(row=row, column=5, value=int(t.montant))
            ws_trans.cell(row=row, column=6, value=int(t.commission))
            ws_trans.cell(row=row, column=7, value=t.date.strftime('%d/%m/%Y %H:%M:%S'))
        
        for col in range(1, 8):
            ws_trans.column_dimensions[get_column_letter(col)].width = 18
        
        wb.save(response)
        return response


# ==================== VUES FACTURES ===================
@login_required
def creer_facture(request):
    """Creer une nouvelle facture"""
    if request.method == 'POST':
        type_facture = request.POST.get('type_facture')
        numero = request.POST.get('numero')
        personne_nom = request.POST.get('personne_nom')
        montant_total = request.POST.get('montant_total')
        date_echeance = request.POST.get('date_echeance')
        
        try:
            montant_total = int(montant_total)
            if montant_total <= 0:
                messages.error(request, "Le montant doit etre superieur a 0")
                return redirect('rapports_admin')
            
            # Generer un numero de facture unique
            import random
            import string
            numero = f"FAC-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            
            facture = Facture.objects.create(
                numero=numero,
                type_facture=type_facture,
                
                personne_nom=personne_nom,
                montant_total=montant_total,
                montant_paye=0,
                date_echeance=date_echeance,
                statut='en_attente',
                cree_par=request.user
            )
            
            messages.success(request, f"Facture {numero} cree avec succes")
            
        except ValueError:
            messages.error(request, "Montant invalide")
        except Exception as e:
            messages.error(request, f"Erreur lors de la creation: {str(e)}")
        
        return redirect('rapports_admin')
    
    return redirect('rapports_admin')

@login_required
def api_factures(request):
    """API pour recuperer les factures"""
    factures = Facture.objects.filter(cree_par=request.user).order_by('-date_emission')
    
    data = {
        'factures': [
            {
                'id': f.id,
                'numero': f.numero,
                'type_facture': f.type_facture,
                'personne_nom': f.personne_nom,
                'montant_total': f.montant_total,
                'montant_paye': f.montant_paye,
                'reste': f.montant_total - f.montant_paye,
                'date_emission': f.date_emission.strftime('%d/%m/%Y'),
                'date_echeance': f.date_echeance.strftime('%d/%m/%Y'),
                'statut': f.statut
            } for f in factures
        ]
    }
    return JsonResponse(data)


@login_required
def rechercher_client_api(request):
    """API pour rechercher un client par téléphone"""
    numero = request.GET.get('numero', '')
    if numero:
        transactions = Transaction.objects.filter(
            numero_client=numero
        ).select_related('user__agent').order_by('-date')[:5]
        
        factures = Facture.objects.filter(
            personne_telephone=numero
        ).order_by('-date_emission')[:5]
        
        clients_data = []
        
        for t in transactions:
            agent_nom = t.user.agent.nom if hasattr(t.user, 'agent') else 'Agent'
            clients_data.append({
                'numero': t.numero_client,
                'nom': f"Client {t.numero_client}",
                'source': f'Transaction du {t.date.strftime("%d/%m/%Y")}',
                'montant_moyen': float(t.montant),
                'agent': agent_nom
            })
        
        for f in factures:
            clients_data.append({
                'numero': f.personne_telephone,
                'nom': f.personne_nom,
                'source': f'Facture {f.numero}',
                'type': f.get_type_facture_display(),
                'montant_total': float(f.montant_total)
            })
        
        uniques = {}
        for client in clients_data:
            if client['numero'] not in uniques:
                uniques[client['numero']] = client
        
        return JsonResponse({
            'success': True,
            'clients': list(uniques.values()),
            'found': len(uniques) > 0
        })
    
    return JsonResponse({'success': False, 'clients': []})


@login_required
def detail_facture(request, facture_id):
    facture = get_object_or_404(Facture, id=facture_id)
    return render(request, 'transactions/detail_facture.html', {'facture': facture})


@login_required
def modifier_facture(request, facture_id):
    """Modifier une facture"""
    facture = get_object_or_404(Facture, id=facture_id)
    
    if request.method == 'POST':
        facture.client_nom = request.POST.get('client_nom', facture.client_nom)
        facture.client_email = request.POST.get('client_email', facture.client_email)
        facture.client_telephone = request.POST.get('client_telephone', facture.client_telephone)
        facture.description = request.POST.get('description', facture.description)
        facture.save()
        messages.success(request, 'Facture modifiée avec succès')
    
    return redirect('rapports_admin')


@login_required
def supprimer_facture(request, facture_id):
    """Supprimer/Annuler une facture"""
    facture = get_object_or_404(Facture, id=facture_id)
    
    if request.method == 'POST':
        facture.statut = 'annulee'
        facture.save()
        messages.success(request, 'Facture annulée')
    
    return redirect('rapports_admin')


@login_required
def enregistrer_paiement_facture(request, facture_id):
    """Enregistrer un paiement sur une facture"""
    facture = get_object_or_404(Facture, id=facture_id)
    
    if request.method == 'POST':
        try:
            montant = request.POST.get('montant', 0)
            try:
                montant = int(montant)
            except ValueError:
                montant = 0
            
            mode = request.POST.get('mode_paiement', 'cash')
            
            if montant <= 0:
                messages.error(request, 'Montant invalide')
            elif montant > facture.reste_a_payer:
                messages.error(request, f'Montant dépasse le reste à payer ({facture.reste_a_payer:,.0f} FCFA)')
            else:
                facture.montant_paye += montant
                facture.save()
                
                PaiementFacture.objects.create(
                    facture=facture,
                    montant=montant,
                    mode_paiement=mode,
                    cree_par=request.user
                )
                messages.success(request, f'Paiement de {montant:,.0f} FCFA enregistré')
        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')
    
    return redirect('rapports_admin')


# ==================== VUES DETTES ====================
# views.py - VERSION CORRECTE AVEC VOS MODÈLES
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Sum
from django.http import JsonResponse, FileResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from io import BytesIO

from .models import Dette, RemboursementDette, Debiteur


# ==================== VUES PRINCIPALES ====================

@login_required
def ajouter_dette(request):
    """Ajouter une nouvelle dette - SAISIE LIBRE DU NOM"""
    if request.method == 'POST':
        try:
            montant = request.POST.get('montant', 0)
            try:
                montant = int(montant)
            except ValueError:
                montant = 0
            
            if montant <= 0:
                messages.error(request, 'Le montant doit être supérieur à 0')
                return redirect('rapports_admin')
            
            nom_debiteur = request.POST.get('nom_debiteur', '').strip()
            if not nom_debiteur:
                messages.error(request, 'Veuillez saisir le nom du débiteur')
                return redirect('rapports_admin')
            
            # Récupérer la valeur de la case "ajouter à la caisse"
            ajouter_caisse = request.POST.get('ajouter_caisse') == 'on'
            
            debiteur, created = Debiteur.objects.get_or_create(
                nom__iexact=nom_debiteur,
                defaults={
                    'nom': nom_debiteur,
                    'telephone': request.POST.get('telephone', ''),
                    'email': request.POST.get('email', ''),
                    'adresse': request.POST.get('adresse', ''),
                    'cree_par': request.user
                }
            )
            
            if not created:
                if request.POST.get('telephone') and not debiteur.telephone:
                    debiteur.telephone = request.POST.get('telephone')
                if request.POST.get('email') and not debiteur.email:
                    debiteur.email = request.POST.get('email')
                if request.POST.get('adresse') and not debiteur.adresse:
                    debiteur.adresse = request.POST.get('adresse')
                debiteur.save()
            
            dette = Dette.objects.create(
                debiteur=debiteur,
                montant=montant,
                date_echeance=request.POST.get('date_echeance'),
                motif=request.POST.get('motif', ''),
                cree_par=request.user
            )
            
            message = f'Dette de {montant:,.0f} FCFA ajoutée pour {debiteur.nom}'
            
            # ========== CRÉATION DETTE : ON SOUSTRAIT DE LA CAISSE ==========
            if ajouter_caisse:
                try:
                    caisse, cree = Caisse.objects.get_or_create(
                        user=request.user,
                        defaults={'solde_cash': 0, 'solde_uv': 0, 'solde_wave': 0}
                    )
                    # SOUSTRAIRE de la caisse (l'argent sort)
                    caisse.solde_cash -= montant
                    caisse.save()
                    message += f" et retiré {montant:,.0f} FCFA de la caisse espèces"
                except Exception as e:
                    message += f" (⚠️ erreur caisse: {str(e)})"
            
            if created:
                messages.success(request, f'Nouveau débiteur "{nom_debiteur}" créé. {message}')
            else:
                messages.success(request, message)
                
        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')
    
    return redirect('rapports_admin')


@login_required
def modifier_dette(request, dette_id):
    """Modifier une dette"""
    dette = get_object_or_404(Dette, id=dette_id)
    
    if request.method == 'POST':
        try:
            montant = request.POST.get('montant', dette.montant)
            try:
                montant = int(montant)
            except ValueError:
                montant = dette.montant
            
            nouveau_nom = request.POST.get('nom_debiteur', '').strip()
            if nouveau_nom and nouveau_nom != dette.debiteur.nom:
                debiteur_existant = Debiteur.objects.filter(nom__iexact=nouveau_nom).first()
                if debiteur_existant:
                    dette.debiteur = debiteur_existant
                else:
                    dette.debiteur.nom = nouveau_nom
                    if request.POST.get('telephone'):
                        dette.debiteur.telephone = request.POST.get('telephone')
                    if request.POST.get('email'):
                        dette.debiteur.email = request.POST.get('email')
                    dette.debiteur.save()
            
            dette.montant = montant
            dette.date_echeance = request.POST.get('date_echeance', dette.date_echeance)
            dette.motif = request.POST.get('motif', dette.motif)
            dette.save()
            messages.success(request, 'Dette modifiée avec succès')
        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')
    
    return redirect('rapports_admin')


@login_required
def supprimer_dette(request, dette_id):
    """Supprimer une dette"""
    dette = get_object_or_404(Dette, id=dette_id)
    
    if request.method == 'POST':
        dette.delete()
        messages.success(request, 'Dette supprimée')
    
    return redirect('rapports_admin')

 
@login_required
def enregistrer_remboursement_dette(request, dette_id):
    """Enregistrer un remboursement et AJOUTER à la caisse"""
    dette = get_object_or_404(Dette, id=dette_id)
    
    if request.method == 'POST':
        from django.urls import reverse
        try:
            montant = request.POST.get('montant', 0)
            try:
                montant = int(montant)
            except ValueError:
                montant = 0
            
            mode = request.POST.get('mode_paiement', 'cash')
            generer_recu = request.POST.get('generer_recu') == 'on'
            ajouter_caisse = request.POST.get('ajouter_caisse') == 'on'
            
            if montant <= 0:
                messages.error(request, 'Montant invalide')
            elif montant > dette.reste_a_payer:
                messages.error(request, f'Montant dépasse le reste à payer ({dette.reste_a_payer:,.0f} FCFA)')
            else:
                # Enregistrer le remboursement
                remboursement = RemboursementDette.objects.create(
                    dette=dette,
                    montant=montant,
                    mode_paiement=mode,
                    cree_par=request.user
                )
                
                # Mettre à jour la dette
                dette.montant_rembourse += montant
                dette.save()
                
                message = f'✅ Remboursement de {montant:,.0f} FCFA enregistré'
                
                # ========== REMBOURSEMENT : ON AJOUTE À LA CAISSE ==========
                if ajouter_caisse:
                    try:
                        caisse, created = Caisse.objects.get_or_create(
                            user=request.user,
                            defaults={'solde_cash': 0, 'solde_uv': 0, 'solde_wave': 0}
                        )
                        
                        mode_labels = {'cash': 'espèces', 'uv': 'UV Touchpiont', 'wave': 'UV Wave'}
                        
                        # AJOUTER à la caisse selon le mode de paiement
                        if mode == 'cash':
                            caisse.solde_cash += montant
                        elif mode == 'uv':
                            caisse.solde_uv += montant
                        elif mode == 'wave':
                            caisse.solde_wave += montant
                        
                        caisse.save()
                        message += f" et ajouté {montant:,.0f} FCFA au solde {mode_labels.get(mode, mode)}"
                    except Exception as e:
                        message += f" (⚠️ erreur caisse: {str(e)})"
                
                # Générer le reçu si demandé (toujours rediriger, téléchargement auto via JS)
                if generer_recu:
                    return redirect(f"{reverse('rapports_admin')}?recu={remboursement.id}")
                
                messages.success(request, message)
                    
        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')
    
    return redirect('rapports_admin')


def generer_recu_pdf(remboursement):
    """Génère un PDF de reçu pour un paiement (format ticket 80mm)"""
    buffer = BytesIO()
    avail = 184
    doc = SimpleDocTemplate(buffer, pagesize=(200, 380), topMargin=8, bottomMargin=8, leftMargin=8, rightMargin=8)
    
    c_gray = colors.HexColor('#6b7280')
    c_dark = colors.HexColor('#111827')
    c_green = colors.HexColor('#059669')
    c_red = colors.HexColor('#dc2626')
    
    s_title = ParagraphStyle('T', fontSize=12, fontName='Helvetica-Bold', textColor=c_dark, alignment=TA_CENTER, spaceAfter=2)
    s_subtitle = ParagraphStyle('ST', fontSize=7, textColor=c_gray, alignment=TA_CENTER, spaceAfter=6)
    s_label = ParagraphStyle('L', fontSize=7, fontName='Helvetica-Bold', textColor=c_gray, spaceAfter=0)
    s_value = ParagraphStyle('V', fontSize=8, textColor=c_dark, spaceAfter=4)
    s_amount = ParagraphStyle('A', fontSize=18, fontName='Helvetica-Bold', textColor=c_green, alignment=TA_CENTER, spaceAfter=2)
    s_rest = ParagraphStyle('R', fontSize=12, fontName='Helvetica-Bold', textColor=c_red, alignment=TA_CENTER, spaceAfter=4)
    s_footer = ParagraphStyle('F', fontSize=6, textColor=c_gray, alignment=TA_CENTER, spaceAfter=0)
    
    story = []
    def hr(c='#d1d5db'): story.append(Table([['']], colWidths=[avail], rowHeights=[1], style=TableStyle([('LINEBELOW',(0,0),(-1,-1),0.5,c)])))
    
    story.append(Paragraph("<b>KONE SERVICES</b>", s_title))
    story.append(Paragraph("Tél : 76 89 77 31", ParagraphStyle('Tel', fontSize=8, fontName='Helvetica-Bold', textColor=c_dark, alignment=TA_CENTER, spaceAfter=2)))
    story.append(Paragraph("Reçu de paiement", s_subtitle))
    hr()
    story.append(Spacer(1, 3))
    
    info = [
        [Paragraph("N° reçu", s_label), Paragraph(f"REC-{remboursement.id:06d}", s_value)],
        [Paragraph("Date", s_label), Paragraph(f"{remboursement.date_remboursement.strftime('%d/%m/%Y %H:%M')}", s_value)],
        [Paragraph("Agent", s_label), Paragraph(f"{remboursement.cree_par.username[:18]}", s_value)],
        [Paragraph("Débiteur", s_label), Paragraph(f"{remboursement.dette.debiteur.nom[:22]}", s_value)],
    ]
    if remboursement.dette.debiteur.telephone:
        info.append([Paragraph("Tél", s_label), Paragraph(f"{remboursement.dette.debiteur.telephone}", s_value)])
    motif = remboursement.dette.motif or "Prêt"
    info.append([Paragraph("Motif", s_label), Paragraph(f"{motif[:22]}", s_value)])
    
    t = Table(info, colWidths=[50, avail-50])
    t.setStyle(TableStyle([
        ('ALIGN', (0,0), (0,-1), 'LEFT'), ('ALIGN', (1,0), (1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 0), ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ('LEFTPADDING', (0,0), (-1,-1), 0), ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(t)
    story.append(Spacer(1, 4))
    hr()
    story.append(Spacer(1, 4))
    
    def sf(v): return f"{v:,.0f}".replace(",", " ")
    story.append(Paragraph(f"{sf(remboursement.montant)} FCFA", s_amount))
    story.append(Spacer(1, 10))
    paye_label = remboursement.get_mode_paiement_display()
    story.append(Paragraph(f"Payé par {paye_label}", s_subtitle))
    story.append(Spacer(1, 8))
    hr()
    story.append(Spacer(1, 4))
    
    recap = [
        [Paragraph("Total dette", s_label), Paragraph(f"{sf(remboursement.dette.montant)} FCFA", s_value)],
        [Paragraph("Déjà remboursé", s_label), Paragraph(f"{sf(remboursement.dette.montant_rembourse - remboursement.montant)} FCFA", s_value)],
        [Paragraph("Ce paiement", s_label), Paragraph(f"{sf(remboursement.montant)} FCFA", s_value)],
    ]
    tr = Table(recap, colWidths=[75, avail-75])
    tr.setStyle(TableStyle([
        ('ALIGN', (0,0), (0,-1), 'LEFT'), ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 0), ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ('LEFTPADDING', (0,0), (-1,-1), 0), ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(tr)
    story.append(Spacer(1, 4))
    
    story.append(Paragraph("RESTE À PAYER", s_label))
    story.append(Paragraph(f"{sf(remboursement.dette.reste_a_payer)} FCFA", s_rest))
    story.append(Spacer(1, 3))
    hr()
    story.append(Spacer(1, 4))
    
    story.append(Paragraph("Merci de votre confiance", ParagraphStyle('Thanks', fontSize=7, textColor=c_gray, alignment=TA_CENTER, spaceAfter=0)))
    story.append(Paragraph(f"Édité le {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", s_footer))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


@login_required
def download_recu(request, remboursement_id):
    """Télécharger le reçu PDF"""
    remboursement = get_object_or_404(RemboursementDette, id=remboursement_id)
    
    if remboursement.cree_par != request.user and not request.user.is_staff:
        messages.error(request, "Permission non accordée")
        return redirect('rapports_admin')
    
    buffer = generer_recu_pdf(remboursement)
    
    filename = f"recu_{remboursement.dette.debiteur.nom}_{remboursement.date_remboursement.strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return FileResponse(
        buffer,
        as_attachment=True,
        filename=filename,
        content_type='application/pdf'
    )


# ==================== API REST ====================

@login_required
def api_dettes(request):
    """API REST pour les dettes"""
    if request.method == 'GET':
        dettes = Dette.objects.all().order_by('-date_creation')
        
        statut = request.GET.get('statut')
        debiteur_nom = request.GET.get('debiteur_nom')
        
        if statut and statut != 'all':
            dettes = dettes.filter(statut=statut)
        if debiteur_nom:
            dettes = dettes.filter(debiteur__nom__icontains=debiteur_nom)
        
        total_montant = dettes.aggregate(Sum('montant'))['montant__sum'] or 0
        total_rembourse = dettes.aggregate(Sum('montant_rembourse'))['montant_rembourse__sum'] or 0
        
        data = {
            'success': True,
            'dettes': [
                {
                    'id': d.id,
                    'debiteur_nom': d.debiteur.nom,
                    'debiteur_id': d.debiteur.id,
                    'montant': float(d.montant),
                    'montant_rembourse': float(d.montant_rembourse),
                    'reste': float(d.reste_a_payer),
                    'statut': d.statut,
                    'date_creation': d.date_creation.strftime('%d/%m/%Y'),
                    'date_echeance': d.date_echeance.strftime('%d/%m/%Y'),
                    'motif': d.motif,
                } for d in dettes
            ],
            'total': float(total_montant),
            'payees': float(total_rembourse),
            'attente': float(total_montant - total_rembourse),
            'nombre': dettes.count(),
        }
        return JsonResponse(data)
    
    return JsonResponse({'success': False, 'error': 'Method not allowed'})


@login_required
def api_dette_detail(request, dette_id):
    """API REST pour le détail d'une dette"""
    if request.method == 'GET':
        dette = get_object_or_404(Dette, id=dette_id)
        
        data = {
            'success': True,
            'id': dette.id,
            'debiteur_nom': dette.debiteur.nom,
            'debiteur_id': dette.debiteur.id,
            'debiteur_telephone': dette.debiteur.telephone or '',
            'debiteur_email': dette.debiteur.email or '',
            'montant': float(dette.montant),
            'montant_rembourse': float(dette.montant_rembourse),
            'reste': float(dette.reste_a_payer),
            'statut': dette.statut,
            'date_creation': dette.date_creation.strftime('%d/%m/%Y'),
            'date_echeance': dette.date_echeance.strftime('%d/%m/%Y'),
            'motif': dette.motif,
            'remboursements': [
                {
                    'id': r.id,
                    'montant': float(r.montant),
                    'mode_paiement': r.mode_paiement,
                    'mode_display': r.get_mode_paiement_display(),
                    'date': r.date_remboursement.strftime('%d/%m/%Y %H:%M:%S'),
                    'cree_par': r.cree_par.username,
                } for r in dette.remboursements.all()
            ]
        }
        return JsonResponse(data)
    
    return JsonResponse({'success': False, 'error': 'Method not allowed'})


@login_required
def api_chercher_debiteurs(request):
    """API pour rechercher des débiteurs par nom (autocomplétion)"""
    if request.method == 'GET':
        terme = request.GET.get('q', '')
        if len(terme) >= 2:
            debiteurs = Debiteur.objects.filter(nom__icontains=terme)[:10]
            data = {
                'success': True,
                'debiteurs': [
                    {
                        'id': d.id,
                        'nom': d.nom,
                        'telephone': d.telephone or '',
                        'email': d.email or '',
                    } for d in debiteurs
                ]
            }
        else:
            data = {'success': True, 'debiteurs': []}
        return JsonResponse(data)
    
    return JsonResponse({'success': False, 'error': 'Method not allowed'})


@login_required
def api_analyse_stats(request):
    """API pour les statistiques d'analyse"""
    today = timezone.now().date()
    
    transactions_aujourdhui = Transaction.objects.filter(date__date=today)
    volume_aujourdhui = transactions_aujourdhui.aggregate(Sum('montant'))['montant__sum'] or 0
    
    month_ago = today - timedelta(days=30)
    best_agent = None
    best_agent_volume = 0
    for agent in Agent.objects.filter(est_actif=True):
        volume = Transaction.objects.filter(
            user=agent.user, 
            date__date__gte=month_ago
        ).aggregate(Sum('montant'))['montant__sum'] or 0
        if volume > best_agent_volume:
            best_agent_volume = volume
            best_agent = agent.nom
    
    operator_stats = {}
    for op_code, op_name in [('orange', 'Orange Money'), ('wave', 'Wave'), ('malitel', 'Moov Africa'), ('telecel', 'Telecel')]:
        volume = Transaction.objects.filter(
            operateur=op_code,
            date__date__gte=month_ago
        ).aggregate(Sum('montant'))['montant__sum'] or 0
        operator_stats[op_name] = volume
    preferred_operator = max(operator_stats, key=operator_stats.get) if operator_stats else "Aucun"
    
    three_months_ago = today - timedelta(days=90)
    stats_3mois = []
    for i in range(3):
        month_start = today - timedelta(days=30 * (i + 1))
        month_end = today - timedelta(days=30 * i)
        volume = Transaction.objects.filter(
            date__date__gte=month_start,
            date__date__lt=month_end
        ).aggregate(Sum('montant'))['montant__sum'] or 0
        stats_3mois.append(volume)
    
    avg_volume = sum(stats_3mois) / len(stats_3mois) if stats_3mois else 0
    prevision = avg_volume * 1.1
    
    evolution_12mois = []
    for i in range(12):
        month_start = today - timedelta(days=30 * (i + 1))
        month_end = today - timedelta(days=30 * i)
        volume = Transaction.objects.filter(
            date__date__gte=month_start,
            date__date__lt=month_end
        ).aggregate(Sum('montant'))['montant__sum'] or 0
        evolution_12mois.insert(0, float(volume) / 1000000)
    
    return JsonResponse({
        'success': True,
        'performance_journaliere': {
            'volume': float(volume_aujourdhui),
            'transactions': transactions_aujourdhui.count(),
        },
        'meilleur_agent': best_agent or "Aucun",
        'operateur_prefere': preferred_operator,
        'prevision_mensuelle': float(prevision),
        'evolution_12_mois': evolution_12mois,
    })


# ==================== API COMPTES ÉPARGNE ====================

@login_required
def api_comptes_epargne(request):
    """API REST pour les comptes epargne"""
    from django.core.serializers import json
    from django.http import JsonResponse
    
    if request.method == 'GET':
        comptes = CompteEpargneAdmin.objects.all().order_by('-date_creation')
        total_solde = comptes.aggregate(Sum('solde'))['solde__sum'] or 0
        
        data = {
            'success': True,
            'comptes': [
                {
                    'id': c.id,
                    'titulaire': c.titulaire,
                    'solde': int(c.solde),
                    'date_ouverture': c.date_creation.strftime('%d/%m/%Y'),
                } for c in comptes
            ],
            'total': int(total_solde),
            'nombre': comptes.count(),
        }
        return JsonResponse(data)
    
    return JsonResponse({'success': False, 'error': 'Method not allowed'})


@login_required
def operation_compte(request, compte_id):
    """Effectuer une operation sur un compte epargne"""
    from django.shortcuts import get_object_or_404, redirect
    from django.contrib import messages
    
    compte = get_object_or_404(CompteEpargneAdmin, id=compte_id)
    
    if request.method == 'POST':
        type_operation = request.POST.get('type_operation')
        montant = request.POST.get('montant', 0)
        
        try:
            montant = int(montant)
        except ValueError:
            montant = 0
        
        if montant <= 0:
            messages.error(request, 'Montant invalide')
        elif type_operation == 'retrait' and montant > compte.solde:
            messages.error(request, 'Solde insuffisant')
        else:
            if type_operation == 'depot':
                compte.solde += montant
                description = f"Depot de {montant:,.0f} FCFA"
            else:
                compte.solde -= montant
                description = f"Retrait de {montant:,.0f} FCFA"
            
            compte.save()
            OperationEpargne.objects.create(
                compte=compte,
                type_operation=type_operation,
                montant=montant,
                description=description
            )
            messages.success(request, f'{description} effectue')
    
    return redirect('rapports_admin')


@login_required
def get_caisse_operations(request):
    """API pour recuperer les operations de caisse"""
    from django.http import JsonResponse
    
    operations = OperationCaisse.objects.filter(user=request.user).order_by('-date_operation')[:50]
    data = {
        'success': True,
        'operations': [
            {
                'id': op.id,
                'type': op.type_operation,
                'montant': int(op.montant),
                'description': op.description,
                'date': op.date_operation.strftime('%d/%m/%Y %H:%M:%S')
            } for op in operations
        ]
    }
    return JsonResponse(data)
# ==================== PDF GENERATION ====================

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from django.http import HttpResponse
from datetime import datetime
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from django.http import HttpResponse
from datetime import datetime


@login_required
def generer_facture_pdf(request, facture_id):
    """Générer une facture PDF professionnelle"""
    
    facture = get_object_or_404(Facture, id=facture_id)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="FACTURE_{facture.numero}.pdf"'
    
    doc = SimpleDocTemplate(
        response, 
        pagesize=A4,
        topMargin=2*cm,
        bottomMargin=2*cm,
        leftMargin=2*cm,
        rightMargin=2*cm
    )
    
    styles = getSampleStyleSheet()
    normal = styles['Normal']
    
    # ========== COULEURS AMÉLIORÉES ==========
    primary = colors.HexColor('#0f766e')      # Teal élégant
    gold = colors.HexColor('#f59e0b')         # Orange doré
    danger = colors.HexColor('#dc2626')       # Rouge plus vif
    success = colors.HexColor('#059669')      # Vert plus profond
    light = colors.HexColor('#f0fdf4')        # Vert très clair
    border = colors.HexColor('#cbd5e1')       # Bordure plus douce
    medium = colors.HexColor('#475569')       # Gris plus soutenu
    dark = colors.HexColor('#1e293b')         # Gris très foncé
    
    style_h1 = ParagraphStyle('H1', parent=styles['Heading1'], fontSize=22, textColor=primary, alignment=TA_LEFT, fontName='Helvetica-Bold')
    style_h1_right = ParagraphStyle('H1Right', parent=styles['Heading1'], fontSize=22, textColor=primary, alignment=TA_RIGHT, fontName='Helvetica-Bold')
    style_title = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=20, textColor=primary, alignment=TA_CENTER, spaceAfter=5, fontName='Helvetica-Bold')
    style_subtitle = ParagraphStyle('Subtitle', parent=normal, alignment=TA_CENTER, fontSize=9, textColor=medium, spaceAfter=15)
    style_section = ParagraphStyle('Section', parent=normal, fontSize=12, textColor=primary, spaceAfter=8, fontName='Helvetica-Bold')
    style_label = ParagraphStyle('Label', parent=normal, fontSize=8, textColor=medium, fontName='Helvetica-Bold')
    style_value = ParagraphStyle('Value', parent=normal, fontSize=9, textColor=dark)
    style_small = ParagraphStyle('Small', parent=normal, fontSize=7, textColor=medium)
    style_small_right = ParagraphStyle('SmallRight', parent=normal, fontSize=7, textColor=medium, alignment=TA_RIGHT)
    
    elements = []
    
    # En-tête
    header_row1 = [[Paragraph("KONE SERVICES", style_h1), Paragraph("FACTURE", style_h1_right)]]
    header_table1 = Table(header_row1, colWidths=[8.5*cm, 8.5*cm])
    header_table1.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    elements.append(header_table1)
    elements.append(Spacer(1, 5))
    
    header_row2 = [[Paragraph("✉ contact@koneservices.com", style_small), Paragraph(f"TYPE : {facture.type_facture}", style_small_right)]]
    header_table2 = Table(header_row2, colWidths=[8.5*cm, 8.5*cm])
    header_table2.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    elements.append(header_table2)
    elements.append(Spacer(1, 3))
    
    header_row3 = [[Paragraph("Bamako - Mali", style_small), Paragraph(f"N° {facture.numero}", style_small_right)]]
    header_table3 = Table(header_row3, colWidths=[8.5*cm, 8.5*cm])
    header_table3.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    elements.append(header_table3)
    elements.append(Spacer(1, 3))
    
    header_row4 = [[Paragraph("📞 +223 76 12 34 56", style_small), Paragraph(f"EMISSION : {facture.date_emission.strftime('%d/%m/%Y')}", style_small_right)]]
    header_table4 = Table(header_row4, colWidths=[8.5*cm, 8.5*cm])
    header_table4.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    elements.append(header_table4)
    elements.append(Spacer(1, 3))
    
    header_row5 = [[Paragraph("", style_small), Paragraph(f"ECHEANCE : {facture.date_echeance.strftime('%d/%m/%Y')}", style_small_right)]]
    header_table5 = Table(header_row5, colWidths=[8.5*cm, 8.5*cm])
    header_table5.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    elements.append(header_table5)
    elements.append(Spacer(1, 10))
    
    line = Table([['']], colWidths=[17*cm], rowHeights=[2])
    line.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), primary)]))
    elements.append(line)
    elements.append(Spacer(1, 20))
    
    # Titre
    elements.append(Paragraph(f"FACTURE {facture.type_facture.upper()}", style_title))
    elements.append(Spacer(1, 20))
    # Informations client

    
    client_data = [
        [Paragraph("NOM", style_label), Paragraph(facture.personne_nom or "Non renseigné", style_value)],
        [Paragraph("TÉLÉPHONE", style_label), Paragraph(facture.personne_telephone or "Non renseigné", style_value)],
    ]
    
    client_table = Table(client_data, colWidths=[4.5*cm, 12*cm])
    client_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), light),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    elements.append(client_table)
    elements.append(Spacer(1, 20))
    
    elements.append(Spacer(1, 20))
    
    # Détails
    elements.append(Paragraph("DETAILS", style_section))
    elements.append(Spacer(1, 8))
    
    details_header = [
        [Paragraph("DESIGNATION", style_label), Paragraph("QTE", style_label), Paragraph("MONTANT", style_label)]
    ]
    details_row = [
        
    ]
    
    if facture.description:
        details_row.insert(0, [Paragraph(facture.description[:50], style_value), Paragraph("1", style_value), Paragraph(f"{facture.montant_total:,.0f} FCFA", style_value)])
    
    details_table = Table(details_header + details_row, colWidths=[9*cm, 3*cm, 5*cm])
    details_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('ALIGN', (1,0), (2,0), 'CENTER'),
        ('ALIGN', (2,1), (2,1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('GRID', (0,1), (-1,-1), 0.5, border),
    ]))
    elements.append(details_table)
    elements.append(Spacer(1, 20))
    
    # Totaux
    totals_data = [
        ["MONTANT TOTAL", f"{facture.montant_total:,.0f} FCFA"],
    ]
    
    if facture.montant_paye > 0:
        totals_data.insert(0, ["MONTANT PAYÉ", f"{facture.montant_paye:,.0f} FCFA"])
    
    totals_table = Table(totals_data, colWidths=[5*cm, 5*cm])
    totals_table.setStyle(TableStyle([
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), gold),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,-1), (-1,-1), 11),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,-1), (-1,-1), 15),
        ('RIGHTPADDING', (0,-1), (-1,-1), 15),
    ]))
    
    totals_container = Table([[totals_table]], colWidths=[17*cm])
    totals_container.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'RIGHT')]))
    elements.append(totals_container)
    elements.append(Spacer(1, 20))
    
    # Pied de page
    footer_line = Table([['']], colWidths=[17*cm], rowHeights=[1])
    footer_line.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), border)]))
    elements.append(footer_line)
    elements.append(Spacer(1, 8))
    
    footer = "MERCI POUR VOTRE CONFIANCE"
    elements.append(Paragraph(footer, style_small))
    elements.append(Paragraph(f"Genere le {datetime.now().strftime('%d/%m/%Y a %H:%M')}", style_small))
    
    doc.build(elements)
    return response


@login_required
def generer_facture_80mm(request, facture_id):
    """Générer un ticket 80mm professionnel avec QR code, détails complets et design pro"""
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from django.http import HttpResponse
    from datetime import datetime
    import qrcode
    from io import BytesIO
    from reportlab.lib.utils import ImageReader
    
    facture = get_object_or_404(Facture, id=facture_id)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="TICKET_{facture.numero}.pdf"'
    
    # Dimensions exactes pour papier 80mm (182mm de haut pour 80mm de large)
    page_width = 8.0 * cm
    page_height = 28.0 * cm
    
    doc = SimpleDocTemplate(
        response, 
        pagesize=(page_width, page_height),
        topMargin=0.3*cm,
        bottomMargin=0.3*cm,
        leftMargin=0.3*cm,
        rightMargin=0.3*cm
    )
    
    styles = getSampleStyleSheet()
    
    # Styles personnalisés professionnels
    style_logo = ParagraphStyle(
        'Logo', parent=styles['Normal'], 
        alignment=TA_CENTER, fontSize=14, 
        fontName='Helvetica-Bold', 
        textColor=colors.HexColor('#0f766e'),
        spaceAfter=4
    )
    
    style_soustitre = ParagraphStyle(
        'Soustitre', parent=styles['Normal'], 
        alignment=TA_CENTER, fontSize=9, 
        textColor=colors.HexColor('#374151'),
        spaceAfter=2
    )
    
    style_sep = ParagraphStyle(
        'Sep', parent=styles['Normal'], 
        alignment=TA_CENTER, fontSize=7, 
        textColor=colors.HexColor('#9ca3af'),
        spaceAfter=4,
        spaceBefore=4
    )
    
    style_title = ParagraphStyle(
        'Title', parent=styles['Normal'], 
        alignment=TA_CENTER, fontSize=12, 
        fontName='Helvetica-Bold', 
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=6,
        spaceBefore=4
    )
    
    style_info = ParagraphStyle(
        'Info', parent=styles['Normal'], 
        alignment=TA_LEFT, fontSize=8, 
        fontName='Helvetica',
        spaceAfter=2
    )
    
    style_info_center = ParagraphStyle(
        'InfoCenter', parent=styles['Normal'], 
        alignment=TA_CENTER, fontSize=8,
        spaceAfter=2
    )
    
    style_label = ParagraphStyle(
        'Label', parent=styles['Normal'], 
        alignment=TA_LEFT, fontSize=8, 
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#4b5563')
    )
    
    style_valeur = ParagraphStyle(
        'Valeur', parent=styles['Normal'], 
        alignment=TA_RIGHT, fontSize=8, 
        fontName='Helvetica'
    )
    
    style_montant = ParagraphStyle(
        'Montant', parent=styles['Normal'], 
        alignment=TA_CENTER, fontSize=16, 
        fontName='Helvetica-Bold', 
        textColor=colors.HexColor('#0f766e'),
        spaceAfter=6,
        spaceBefore=4
    )
    
    style_total = ParagraphStyle(
        'Total', parent=styles['Normal'], 
        alignment=TA_CENTER, fontSize=18, 
        fontName='Helvetica-Bold', 
        textColor=colors.HexColor('#ef4444'),
        spaceAfter=6,
        spaceBefore=4
    )
    
    style_badge = ParagraphStyle(
        'Badge', parent=styles['Normal'], 
        alignment=TA_CENTER, fontSize=9, 
        fontName='Helvetica-Bold',
        spaceAfter=6,
        spaceBefore=4
    )
    
    style_footer = ParagraphStyle(
        'Footer', parent=styles['Normal'], 
        alignment=TA_CENTER, fontSize=7, 
        textColor=colors.HexColor('#6b7280'),
        spaceAfter=2
    )
    
    elements = []
    
    # ==================== ENTÊTE ====================
    elements.append(Paragraph("🏦 KONE SERVICES ", style_logo))
    elements.append(Paragraph("• Solutions Financières •", style_soustitre))
    elements.append(Paragraph("Tel: +223 76 12 34 56", style_soustitre))
    elements.append(Paragraph("Email: contact@entreprise.com", style_soustitre))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#0f766e'), spaceAfter=4, spaceBefore=4))
    
    # ==================== TYPE DE DOCUMENT ====================
    if facture.type_facture == 'cliente':
        elements.append(Paragraph("📄 FACTURE CLIENT", style_title))
    elif facture.type_facture == 'fournisseur':
        elements.append(Paragraph("📄 FACTURE FOURNISSEUR", style_title))
    else:
        elements.append(Paragraph(f"📄 {facture.type_facture.upper()}", style_title))
    
    elements.append(HRFlowable(width="100%", thickness=0.3, color=colors.HexColor('#d1d5db'), spaceAfter=4, spaceBefore=2))
    
    # ==================== INFORMATIONS FACTURE ====================
    # Tableau des infos facture
    info_data = [
        [Paragraph("<b>N° Facture</b>", style_label), Paragraph(facture.numero, style_valeur)],
        [Paragraph("<b>Date émission</b>", style_label), Paragraph(facture.date_emission.strftime('%d/%m/%Y à %H:%M'), style_valeur)],
    ]
    
    if facture.date_echeance:
        info_data.append([Paragraph("<b>Date échéance</b>", style_label), Paragraph(facture.date_echeance.strftime('%d/%m/%Y'), style_valeur)])
    
    info_table = Table(info_data, colWidths=[3.2*cm, 3.8*cm])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(info_table)
    
    elements.append(HRFlowable(width="100%", thickness=0.3, color=colors.HexColor('#d1d5db'), spaceAfter=4, spaceBefore=4))
    
    # ==================== INFORMATIONS CLIENT ====================
    elements.append(Paragraph("<b>👤 INFORMATIONS CLIENT</b>", style_label))
    
    client_data = [
        [Paragraph("<b>Nom</b>", style_label), Paragraph(facture.personne_nom or "-", style_valeur)],
    ]
    
    if facture.personne_telephone:
        client_data.append([Paragraph("<b>Téléphone</b>", style_label), Paragraph(facture.personne_telephone, style_valeur)])
    
    client_table = Table(client_data, colWidths=[3.2*cm, 3.8*cm])
    client_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(client_table)
    
    elements.append(HRFlowable(width="100%", thickness=0.3, color=colors.HexColor('#d1d5db'), spaceAfter=4, spaceBefore=4))
    
    # ==================== DÉTAIL DE LA FACTURE ====================
    elements.append(Paragraph("<b>📋 DÉTAIL DE LA FACTURE</b>", style_label))
    
    # Tableau des lignes de détail
    detail_data = [
        [Paragraph("<b>Désignation</b>", style_label), Paragraph("<b>Montant</b>", style_valeur)],
        [Paragraph("Prestation de service", style_info), Paragraph(f"{facture.montant_total:,.0f} FCFA", style_valeur)],
    ]
    
    if facture.description:
        detail_data.insert(2, [Paragraph(facture.description[:50], style_info), Paragraph("", style_valeur)])
    
    detail_table = Table(detail_data, colWidths=[4.5*cm, 2.5*cm])
    detail_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, 0), 0.5, colors.HexColor('#d1d5db')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(detail_table)
    
    elements.append(HRFlowable(width="100%", thickness=0.3, color=colors.HexColor('#d1d5db'), spaceAfter=4, spaceBefore=4))
    
    # ==================== RÉCAPITULATIF FINANCIER (sans TVA ni RESTE À PAYER) ====================
    recap_data = [
        [Paragraph("<b>Montant total</b>", style_label), Paragraph(f"{facture.montant_total:,.0f} FCFA", style_valeur)],
    ]
    
    if facture.montant_paye > 0:
        recap_data.append([Paragraph("<b>Montant payé</b>", style_label), Paragraph(f"{facture.montant_paye:,.0f} FCFA", style_valeur)])
    
    recap_table = Table(recap_data, colWidths=[4.5*cm, 2.5*cm])
    recap_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(recap_table)
    
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#0f766e'), spaceAfter=4, spaceBefore=4))
    
    # ==================== STATUT ====================
   
    
    # ==================== QR CODE (POUR PAIEMENT RAPIDE) ====================
    try:
        # Génération du QR code avec les infos de la facture
        qr_data = f"FACTURE:{facture.numero}|MONTANT:{facture.montant_total - facture.montant_paye}|CLIENT:{facture.personne_nom}"
        qr = qrcode.QRCode(version=1, box_size=2, border=1)
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="#0f766e", back_color="white")
        
        # Convertir en bytes pour ReportLab
        qr_bytes = BytesIO()
        qr_img.save(qr_bytes, format='PNG')
        qr_bytes.seek(0)
        qr_reader = ImageReader(qr_bytes)
        
        # Ajouter le QR code
        from reportlab.platypus import Image
        qr_image = Image(qr_reader, width=1.5*cm, height=1.5*cm)
        qr_image.hAlign = 'CENTER'
        elements.append(Spacer(1, 4))
        elements.append(qr_image)
        elements.append(Paragraph("Scannez pour régler", style_footer))
    except Exception:
        pass  # Si erreur QR code, on continue sans
    
    elements.append(HRFlowable(width="100%", thickness=0.3, color=colors.HexColor('#d1d5db'), spaceAfter=4, spaceBefore=4))
    
    # ==================== PIED DE PAGE ====================
    elements.append(Paragraph("Merci de votre confiance !", style_footer))
    
    elements.append(Paragraph("Conservez ce ticket comme justificatif", style_footer))
    
    # Construction du PDF
    doc.build(elements)
    return response



# views.py - Modifier l'API
@login_required
def api_user_password(request, user_id):
    """API pour récupérer le mot de passe d'un utilisateur"""
    try:
        user = User.objects.get(id=user_id)
        
        # Essayer de récupérer depuis Agent
        try:
            agent = Agent.objects.get(user=user)
            if agent.mot_de_passe_clair:
                return JsonResponse({
                    'success': True,
                    'password': agent.mot_de_passe_clair
                })
        except Agent.DoesNotExist:
            pass
        
        # Essayer de récupérer depuis Assistant
        try:
            assistant = Assistant.objects.get(user=user)
            if assistant.mot_de_passe_clair:
                return JsonResponse({
                    'success': True,
                    'password': assistant.mot_de_passe_clair
                })
        except Assistant.DoesNotExist:
            pass
        
        return JsonResponse({
            'success': True,
            'password': 'Mot de passe non stocké en clair'
        })
        
    except User.DoesNotExist:
        return JsonResponse({
            'success': False,
            'password': 'Utilisateur introuvable'
        })

from django.db.models import Sum, Count, Avg
from django.utils import timezone
from datetime import timedelta, datetime
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from decimal import Decimal

@login_required
def api_analyse_stats(request):
    """
    API pour les statistiques d'analyse
    """
    try:
        today = timezone.now().date()
        
        # ========== PERFORMANCE JOURNALIERE ==========
        transactions_today = Transaction.objects.filter(date__date=today)
        volume_today = transactions_today.aggregate(Sum('montant'))['montant__sum'] or 0
        transactions_count_today = transactions_today.count()
        
        # Convertir Decimal en int/float
        if isinstance(volume_today, Decimal):
            volume_today = int(volume_today)
        
        # Taux de croissance (comparaison avec hier)
        yesterday = today - timedelta(days=1)
        transactions_yesterday = Transaction.objects.filter(date__date=yesterday)
        volume_yesterday = transactions_yesterday.aggregate(Sum('montant'))['montant__sum'] or 0
        
        if isinstance(volume_yesterday, Decimal):
            volume_yesterday = int(volume_yesterday)
        
        if volume_yesterday > 0:
            croissance = ((volume_today - volume_yesterday) / volume_yesterday) * 100
        else:
            croissance = 0 if volume_today == 0 else 100
        
        # ========== MEILLEUR AGENT / UTILISATEUR ==========
        from django.contrib.auth.models import User
        
        user_stats = []
        for user_obj in User.objects.filter(is_active=True):
            total_volume = Transaction.objects.filter(
                user=user_obj,
                date__date__gte=today - timedelta(days=30)
            ).aggregate(Sum('montant'))['montant__sum'] or 0
            
            if isinstance(total_volume, Decimal):
                total_volume = int(total_volume)
            
            if total_volume > 0:
                nom_complet = f"{user_obj.first_name} {user_obj.last_name}".strip()
                if not nom_complet:
                    nom_complet = user_obj.username
                user_stats.append({
                    'nom': nom_complet,
                    'volume': total_volume
                })
        
        user_stats.sort(key=lambda x: x['volume'], reverse=True)
        
        if user_stats:
            meilleur_agent = user_stats[0]['nom']
        else:
            meilleur_agent = "Aucune transaction"
        
        # ========== OPERATEUR PREFERE ==========
        operateur_stats = Transaction.objects.values('operateur')\
            .annotate(total_volume=Sum('montant'), total_count=Count('id'))\
            .order_by('-total_volume')\
            .first()
        
        operateur_labels = {
            'orange': 'Orange Money',
            'malitel': 'Malitel',
            'telecel': 'Telecel',
            'wave': 'Wave'
        }
        
        if operateur_stats:
            operateur_prefere = operateur_labels.get(operateur_stats['operateur'], operateur_stats['operateur'])
        else:
            operateur_prefere = "Aucune opération"
        
        # ========== PREVISION MENSUELLE ==========
        # Calculer la moyenne des 3 derniers mois complets
        three_months_ago = today - timedelta(days=90)
        start_date = datetime(three_months_ago.year, three_months_ago.month, 1).date()
        end_date = datetime(today.year, today.month, 1).date() - timedelta(days=1)
        
        if start_date <= end_date:
            monthly_transactions = Transaction.objects.filter(
                date__date__gte=start_date,
                date__date__lte=end_date
            )
            total_volume_3mois = monthly_transactions.aggregate(Sum('montant'))['montant__sum'] or 0
            
            if isinstance(total_volume_3mois, Decimal):
                total_volume_3mois = int(total_volume_3mois)
            
            prevision_mensuelle = total_volume_3mois / 3 if total_volume_3mois > 0 else 0
        else:
            prevision_mensuelle = 0
        
        # ========== EVOLUTION SUR 12 MOIS ==========
        evolution_12_mois = []
        for i in range(11, -1, -1):
            date_cible = today.replace(day=1) - timedelta(days=30 * i)
            mois = date_cible.month
            annee = date_cible.year
            
            # Premier jour du mois
            debut_mois = date_cible.replace(day=1)
            # Dernier jour du mois
            if mois == 12:
                fin_mois = debut_mois.replace(year=annee+1, month=1) - timedelta(days=1)
            else:
                fin_mois = debut_mois.replace(month=mois+1) - timedelta(days=1)
            
            volume_mois = Transaction.objects.filter(
                date__date__gte=debut_mois,
                date__date__lte=fin_mois
            ).aggregate(Sum('montant'))['montant__sum'] or 0
            
            if isinstance(volume_mois, Decimal):
                volume_mois = int(volume_mois)
            
            # Convertir en millions (avec 1 décimale)
            evolution_12_mois.append(round(float(volume_mois / 1000000), 1))
        
        # ========== TOP 5 USERS ==========
        top_5_users = []
        for us in user_stats[:5]:
            user_obj = User.objects.filter(
                username__icontains=us['nom'].split()[-1] if ' ' in us['nom'] else us['nom']
            ).first()
            if user_obj:
                if hasattr(user_obj, 'admin_profile') and user_obj.admin_profile:
                    role = 'admin'
                elif hasattr(user_obj, 'agent_profile') and user_obj.agent_profile:
                    role = 'agent'
                elif hasattr(user_obj, 'assistant_profile') and user_obj.assistant_profile:
                    role = 'assistant'
                else:
                    role = 'agent'
            else:
                role = 'agent'
            top_5_users.append({
                'nom': us['nom'],
                'volume': us['volume'],
                'type': role
            })
        
        return JsonResponse({
            'success': True,
            'performance_journaliere': {
                'volume': int(volume_today),
                'transactions': transactions_count_today,
                'croissance': round(float(croissance), 1)
            },
            'meilleur_agent': meilleur_agent,
            'operateur_prefere': operateur_prefere,
            'prevision_mensuelle': int(prevision_mensuelle),
            'evolution_12_mois': evolution_12_mois,
            'top_5_users': top_5_users
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)



@login_required
def supprimer_facture(request, facture_id):
    print(f"=== Tentative de suppression de la facture {facture_id} par {request.user} ===")
    
    try:
        from transactions.models import Facture
        print("Modèle Facture importé")
        
        facture = Facture.objects.get(id=facture_id)
        print(f"Facture trouvée: {facture.numero}")
        
        facture.delete()
        print("Facture supprimée avec succès")
        
        return JsonResponse({
            'success': True,
            'message': 'Facture supprimée avec succès'
        })
        
    except Facture.DoesNotExist:
        print(f"Facture {facture_id} non trouvée")
        return JsonResponse({
            'success': False,
            'error': f'Facture avec ID {facture_id} non trouvée'
        }, status=404)
    except Exception as e:
        print(f"Erreur: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)



def get_demandes_filter_for_user(user):
    q = Q(statut='en_attente')
    if hasattr(user, 'assistant_profile'):
        q &= Q(destinataire_type='assistant', assistant_destinataire=user.assistant_profile)
    else:
        q &= Q(destinataire_type='admin')
    return q

def api_demandes_attente_count(request):
    q = get_demandes_filter_for_user(request.user)
    count = DemandeApprovisionnement.objects.filter(q).count()
    return JsonResponse({'count': count})

def api_demandes_attente_list(request):
    q = get_demandes_filter_for_user(request.user)
    demandes = DemandeApprovisionnement.objects.filter(q).select_related('agent')
    data = []
    for d in demandes:
        data.append({
            'id': d.id,
            'agent_nom': d.agent.nom,
            'type_echange': d.get_type_echange_display(),
            'montant': f"{d.montant:,.0f}".replace(',', ' '),
            'motif': d.motif[:40] if d.motif else '',
        })
    return JsonResponse({'demandes': data})

@login_required
def api_annuler_transaction(request):
    import json
    try:
        data = json.loads(request.body)
        reference = data.get('reference', '').strip().upper()
        motif = data.get('motif', '').strip()
        chercher = data.get('chercher', False)
    except (json.JSONDecodeError, AttributeError):
        reference = request.POST.get('reference', '').strip().upper()
        motif = request.POST.get('motif', '').strip()
        chercher = request.POST.get('chercher') == 'true'

    if not reference:
        return JsonResponse({'success': False, 'error': 'Référence obligatoire.'})

    try:
        transaction = Transaction.objects.select_related('user__agent_profile', 'user__admin_profile', 'user__assistant_profile').get(reference=reference)
    except Transaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Aucune transaction trouvée avec cette référence.'})

    if chercher:
        agent_nom = ''
        if hasattr(transaction.user, 'agent_profile'):
            agent_nom = transaction.user.agent_profile.nom
        elif hasattr(transaction.user, 'assistant_profile'):
            agent_nom = f"{transaction.user.assistant_profile.nom} (Assistant)"
        elif hasattr(transaction.user, 'admin_profile'):
            agent_nom = f"{transaction.user.admin_profile.nom} (Admin)"
        return JsonResponse({
            'success': True,
            'transaction': {
                'reference': transaction.reference,
                'date': transaction.date.strftime('%d/%m/%Y %H:%M'),
                'type': transaction.get_type_transaction_display(),
                'operateur': transaction.get_operateur_display(),
                'montant': f"{transaction.montant:,.0f}".replace(',', ' '),
                'client': f"{transaction.numero_client} {transaction.nom_client or ''}",
                'agent': agent_nom,
                'est_annule': transaction.est_annule,
            }
        })

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'})

    if transaction.est_annule:
        return JsonResponse({'success': False, 'error': 'Cette transaction est déjà annulée.'})

    from django.utils import timezone
    transaction.annuler(user=request.user, motif=motif)
    return JsonResponse({'success': True, 'message': f'Transaction {reference} annulée avec succès.'})


def sign_message(request):
    """
    Signe un message pour QZ Tray (signature HTTPS sans popup)
    Endpoint public : la securite repose sur le certificat QZ,
    pas sur la session navigateur (une session expiree ne doit
    pas casser le signature, sinon QZ redemande a chaque fois).
    Utilise cryptography si disponible, sinon openssl CLI
    """
    msg = request.GET.get('request', '')
    if not msg:
        return HttpResponse('', content_type='text/plain')

    key_path = Path(__file__).resolve().parent.parent / 'certs' / 'private-key.pem'
    if not key_path.exists():
        return HttpResponse('', content_type='text/plain')

    try:
        from cryptography.hazmat.primitives import hashes, serialization
        from cryptography.hazmat.primitives.asymmetric import padding
        from cryptography.hazmat.backends import default_backend
        with open(key_path, 'rb') as f:
            key = serialization.load_pem_private_key(f.read(), password=None, backend=default_backend())
        sig = key.sign(msg.encode('utf-8'), padding.PKCS1v15(), hashes.SHA512())
        return HttpResponse(base64.b64encode(sig), content_type='text/plain')
    except ImportError:
        import subprocess
        result = subprocess.run(
            ['openssl', 'dgst', '-sha512', '-sign', str(key_path)],
            input=msg.encode('utf-8'),
            capture_output=True
        )
        if result.returncode == 0:
            return HttpResponse(base64.b64encode(result.stdout), content_type='text/plain')
        return HttpResponse('', content_type='text/plain')