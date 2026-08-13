# Auto-split depuis transactions/views.py (refactor)
# imports partages (niveau module d'origine)

import base64

from pathlib import Path

import django

from django.shortcuts import render, redirect, get_object_or_404

from django.contrib.auth.decorators import login_required

from django.contrib import messages

from django.http import JsonResponse

from django.views.decorators.http import require_POST

from django.db.models import Sum, Q

from datetime import datetime, timedelta

from decimal import Decimal

import json

from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator

from django.contrib.auth.models import User

from django.contrib.auth.hashers import make_password

from django.contrib.auth import logout

from django.views.decorators.http import require_http_methods

from django.utils import timezone

import csv

from ..models import Admin, Agent, Assistant, Caisse, CompteEpargneAdmin, OperationCaisse, OperationEpargne, OperationUv, Transaction, DemandeApprovisionnement, ApprovisionnementDirect

from openpyxl import Workbook

from openpyxl.styles import Font, Alignment, PatternFill

from reportlab.lib.pagesizes import A4

from reportlab.lib.units import cm

from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from django.http import HttpResponse

from django.template.loader import get_template

from xhtml2pdf import pisa

import io

from ..models import (
    Admin, Agent, Assistant, Caisse, Transaction, DemandeApprovisionnement, 
    ApprovisionnementDirect, Facture, PaiementFacture, Dette, RemboursementDette,
    CompteEpargne, OperationCompte
)

from ..forms import (OrangeTransactionForm, WaveTransactionForm, 
                   MalitelTransactionForm, TelecelTransactionForm)

from django.shortcuts import get_object_or_404

from django.http import JsonResponse

from django.contrib.auth.decorators import login_required

from django.template.loader import render_to_string

from django.utils import timezone

from decimal import Decimal

from django.contrib.auth.decorators import login_required

from django.http import JsonResponse

from django.shortcuts import redirect, render, get_object_or_404

from django.contrib import messages

from django.db.models import Sum

from transactions.models import Agent, Assistant, Admin, Caisse, HistoriqueAgent

from django.core.paginator import Paginator

from datetime import datetime

from django.db.models import Sum, Q

from transactions.models import HistoriqueAgent, Agent

import json

import csv

from datetime import datetime, timedelta

from django.db.models import Sum

from django.core.paginator import Paginator

from django.http import JsonResponse, HttpResponse

from django.shortcuts import render, redirect, get_object_or_404

from django.contrib.auth.decorators import login_required

from django.contrib import messages

from django.utils import timezone

from openpyxl import Workbook

from openpyxl.styles import Font, Alignment, PatternFill

from ..models import (
    Transaction, Agent, Admin, Caisse, DemandeApprovisionnement,
    Facture, PaiementFacture, Dette, RemboursementDette,
    CompteEpargne, OperationCompte
)

from django.contrib.auth.decorators import login_required

from django.shortcuts import render, redirect, get_object_or_404

from django.contrib import messages

from django.db.models import Sum

from django.http import JsonResponse, FileResponse

from django.utils import timezone

from django.views.decorators.csrf import csrf_exempt

from django.views.decorators.http import require_http_methods

import json

from datetime import datetime

from reportlab.lib import colors

from reportlab.lib.pagesizes import A4

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from reportlab.lib.units import cm

from io import BytesIO

from ..models import Dette, RemboursementDette, Debiteur

from reportlab.lib.pagesizes import A4

from reportlab.lib.units import cm

from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from django.http import HttpResponse

from datetime import datetime

from reportlab.lib.units import cm

from reportlab.lib import colors

from reportlab.lib.pagesizes import A4

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from django.http import HttpResponse

from datetime import datetime

from django.db.models import Sum, Count, Avg

from django.utils import timezone

from datetime import timedelta, datetime

from django.http import JsonResponse

from django.contrib.auth.decorators import login_required

from decimal import Decimal

@login_required
def rapports_admin(request):
    """Page principale des rapports et gestion"""
    try:
        admin = Admin.objects.get(user=request.user)
    except Admin.DoesNotExist:
        messages.error(request, "Vous n'etes pas autorise.")
        return redirect('login')
    
    # ========== RECUPERATION DES UTILISATEURS ==========
    admins = Admin.objects.all()
    agents = Agent.objects.all()
    assistants = Assistant.objects.filter(est_actif=True)
    
    total_users = admins.count() + agents.filter(est_actif=True).count() + assistants.count()
    agents_actifs_count = agents.filter(est_actif=True).count()
    
    today = timezone.now().date()
    default_date_debut = today - timedelta(days=30)
    
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    if date_debut_str and date_fin_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            date_debut = default_date_debut
            date_fin = today
    else:
        date_debut = default_date_debut
        date_fin = today
    
    transactions = Transaction.objects.filter(
        date__date__gte=date_debut,
        date__date__lte=date_fin
    )
    
    total_transactions = transactions.count()
    total_volume = int(transactions.aggregate(total=Sum('montant'))['total'] or 0)
    total_commission = int(transactions.aggregate(total=Sum('commission'))['total'] or 0)
    
    # ========== FILTRAGE POUR L'HISTORIQUE ==========
    filtre_type = request.GET.get('filtre_type', 'toutes')
    filtre_compte = request.GET.get('filtre_compte', 'tous')
    filtre_date_debut = request.GET.get('filtre_date_debut', '')
    filtre_date_fin = request.GET.get('filtre_date_fin', '')
    
    # ========== GESTION DE CAISSE ADMIN UNIFIEE ==========
    # Encaissement (Depot) - avec selection du compte
    if request.method == 'POST' and 'encaissement' in request.POST:
        montant = request.POST.get('montant_encaissement')
        description = request.POST.get('description_encaissement', 'Encaissement')
        compte_concerne = request.POST.get('compte_concerne', 'cash')
        
        try:
            montant = int(montant)
            if montant > 0:
                caisse_admin, created = Caisse.objects.get_or_create(user=request.user)
                
                if compte_concerne == 'cash':
                    caisse_admin.solde_cash += montant
                    caisse_admin.save()
                    OperationCaisse.objects.create(
                        caisse=caisse_admin,
                        type_operation='encaissement',
                        montant=montant,
                        description=description,
                        user=request.user
                    )
                    messages.success(request, f'Encaissement de {montant:,.0f} FCFA sur Espèces effectue')
                    
                elif compte_concerne == 'uv_touchpoint':
                    caisse_admin.solde_uv += montant
                    caisse_admin.save()
                    OperationUv.objects.create(
                        caisse=caisse_admin,
                        type_operation='ajout',
                        type_uv='touchpoint',
                        montant=montant,
                        description=description,
                        user=request.user
                    )
                    messages.success(request, f'Ajout de {montant:,.0f} FCFA sur UV Touchpoint effectue')
                    
                elif compte_concerne == 'uv_wave':
                    caisse_admin.solde_wave += montant
                    caisse_admin.save()
                    OperationUv.objects.create(
                        caisse=caisse_admin,
                        type_operation='ajout',
                        type_uv='wave',
                        montant=montant,
                        description=description,
                        user=request.user
                    )
                    messages.success(request, f'Ajout de {montant:,.0f} FCFA sur UV Wave effectue')
                    
                elif compte_concerne == 'epargne':
                    compte_epargne, created = CompteEpargneAdmin.objects.get_or_create(
                        user=request.user,
                        defaults={'solde': 0, 'titulaire': request.user.username}
                    )
                    compte_epargne.solde += montant
                    compte_epargne.save()
                    OperationEpargne.objects.create(
                        compte=compte_epargne,
                        type_operation='depot',
                        montant=montant,
                        description=description
                    )
                    messages.success(request, f'Ajout de {montant:,.0f} FCFA sur Epargne effectue')
            else:
                messages.error(request, 'Montant invalide')
        except ValueError:
            messages.error(request, 'Montant invalide')
        return redirect('rapports_admin')
    
    # Decaissement (Retrait) - avec selection du compte
    if request.method == 'POST' and 'decaissement' in request.POST:
        montant = request.POST.get('montant_decaissement')
        description = request.POST.get('description_decaissement', 'Decaissement')
        compte_concerne = request.POST.get('compte_concerne', 'cash')
        
        try:
            montant = int(montant)
            if montant > 0:
                caisse_admin, created = Caisse.objects.get_or_create(user=request.user)
                erreur = False
                
                if compte_concerne == 'cash':
                    if montant <= caisse_admin.solde_cash:
                        caisse_admin.solde_cash -= montant
                        caisse_admin.save()
                        OperationCaisse.objects.create(
                            caisse=caisse_admin,
                            type_operation='decaissement',
                            montant=montant,
                            description=description,
                            user=request.user
                        )
                        messages.success(request, f'Decaissement de {montant:,.0f} FCFA sur Espèces effectue')
                    else:
                        messages.error(request, 'Solde Espèces insuffisant')
                        erreur = True
                        
                elif compte_concerne == 'uv_touchpoint':
                    if montant <= caisse_admin.solde_uv:
                        caisse_admin.solde_uv -= montant
                        caisse_admin.save()
                        OperationUv.objects.create(
                            caisse=caisse_admin,
                            type_operation='retrait',
                            type_uv='touchpoint',
                            montant=montant,
                            description=description,
                            user=request.user
                        )
                        messages.success(request, f'Retrait de {montant:,.0f} FCFA sur UV Touchpoint effectue')
                    else:
                        messages.error(request, 'Solde UV Touchpoint insuffisant')
                        erreur = True
                        
                elif compte_concerne == 'uv_wave':
                    if montant <= caisse_admin.solde_wave:
                        caisse_admin.solde_wave -= montant
                        caisse_admin.save()
                        OperationUv.objects.create(
                            caisse=caisse_admin,
                            type_operation='retrait',
                            type_uv='wave',
                            montant=montant,
                            description=description,
                            user=request.user
                        )
                        messages.success(request, f'Retrait de {montant:,.0f} FCFA sur UV Wave effectue')
                    else:
                        messages.error(request, 'Solde UV Wave insuffisant')
                        erreur = True
                        
                elif compte_concerne == 'epargne':
                    compte_epargne = CompteEpargneAdmin.objects.filter(user=request.user).first()
                    if compte_epargne and montant <= compte_epargne.solde:
                        compte_epargne.solde -= montant
                        compte_epargne.save()
                        OperationEpargne.objects.create(
                            compte=compte_epargne,
                            type_operation='retrait',
                            montant=montant,
                            description=description
                        )
                        messages.success(request, f'Retrait de {montant:,.0f} FCFA sur Epargne effectue')
                    else:
                        messages.error(request, 'Solde Epargne insuffisant')
                        erreur = True
                
                if not erreur:
                    pass
            else:
                messages.error(request, 'Montant invalide')
        except ValueError:
            messages.error(request, 'Montant invalide')
        return redirect('rapports_admin')
    
    # ========== PERFORMANCE PAR UTILISATEUR ==========
    user_performance = []
    
    for agent in agents:
        agent_transactions = transactions.filter(user=agent.user)
        volume = int(agent_transactions.aggregate(total=Sum('montant'))['total'] or 0)
        if volume > 0:
            user_performance.append({'nom': agent.nom, 'montant': volume, 'id': agent.id, 'type': 'agent'})
    
    for assistant in assistants:
        assistant_transactions = transactions.filter(user=assistant.user)
        volume = int(assistant_transactions.aggregate(total=Sum('montant'))['total'] or 0)
        if volume > 0:
            user_performance.append({'nom': assistant.nom, 'montant': volume, 'id': assistant.id, 'type': 'assistant'})
    
    for admin_user in admins:
        admin_transactions = transactions.filter(user=admin_user.user)
        volume = int(admin_transactions.aggregate(total=Sum('montant'))['total'] or 0)
        if volume > 0:
            user_performance.append({'nom': admin_user.nom, 'montant': volume, 'id': admin_user.id, 'type': 'admin'})
    
    user_performance = sorted(user_performance, key=lambda x: x['montant'], reverse=True)[:5]
    max_volume = user_performance[0]['montant'] if user_performance else 1
    
    for up in user_performance:
        up['percentage'] = int((up['montant'] / max_volume * 100)) if max_volume > 0 else 0
    
    demandes_attente = DemandeApprovisionnement.objects.filter(statut='en_attente')
    
    stats_today = {
        'nombre': Transaction.objects.filter(date__date=today).count(),
        'depots': int(Transaction.objects.filter(date__date=today, type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
        'retraits': int(Transaction.objects.filter(date__date=today, type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
        'credits': int(Transaction.objects.filter(date__date=today, type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0),
    }
    
    stats_yesterday = {'nombre': Transaction.objects.filter(date__date=today - timedelta(days=1)).count()}
    evolution = {'nombre': ((stats_today['nombre'] - stats_yesterday['nombre']) / stats_yesterday['nombre'] * 100) if stats_yesterday['nombre'] > 0 else 0}
    
    top_users = []
    
    for agent in agents.filter(est_actif=True):
        trans_aujourdhui = Transaction.objects.filter(user=agent.user, date__date=today)
        count = trans_aujourdhui.count()
        total = int(trans_aujourdhui.aggregate(Sum('montant'))['montant__sum'] or 0)
        if count > 0:
            top_users.append({'user': agent, 'type': 'agent', 'transactions': count, 'montant': total})
    
    for assistant in assistants:
        trans_aujourdhui = Transaction.objects.filter(user=assistant.user, date__date=today)
        count = trans_aujourdhui.count()
        total = int(trans_aujourdhui.aggregate(Sum('montant'))['montant__sum'] or 0)
        if count > 0:
            top_users.append({'user': assistant, 'type': 'assistant', 'transactions': count, 'montant': total})
    
    top_users = sorted(top_users, key=lambda x: x['transactions'], reverse=True)[:5]
    dernieres_transactions = Transaction.objects.all().order_by('-date')[:10]
    
    # Caisse de l'admin
    try:
        caisse = Caisse.objects.get(user=request.user)
        caisse.solde_cash = int(caisse.solde_cash or 0)
        caisse.solde_uv = int(caisse.solde_uv or 0)
        caisse.solde_wave = int(caisse.solde_wave or 0)
    except Caisse.DoesNotExist:
        caisse = Caisse.objects.create(
            user=request.user,
            solde_cash=0,
            solde_uv=0,
            solde_wave=0
        )
    
    # Compte epargne
    compte_epargne, created = CompteEpargneAdmin.objects.get_or_create(
        user=request.user,
        defaults={'solde': 0, 'titulaire': request.user.username}
    )
    compte_epargne.solde = int(compte_epargne.solde or 0)
    
    # ========== CALCUL DES TOTAUX D'ENCAISSEMENT ET DECAISSEMENT ==========
    # Totaux sur les operations de caisse
    total_encaissements_caisse = OperationCaisse.objects.filter(
        user=request.user, 
        type_operation='encaissement'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    total_decaissements_caisse = OperationCaisse.objects.filter(
        user=request.user, 
        type_operation='decaissement'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Totaux sur les operations UV (ajouts = encaissements, retraits = decaissements)
    total_ajouts_uv = OperationUv.objects.filter(
        user=request.user, 
        type_operation='ajout'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    total_retraits_uv = OperationUv.objects.filter(
        user=request.user, 
        type_operation='retrait'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Totaux sur les operations Epargne (depot = encaissements, retrait = decaissements)
    total_depots_epargne = OperationEpargne.objects.filter(
        compte=compte_epargne, 
        type_operation='depot'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    total_retraits_epargne = OperationEpargne.objects.filter(
        compte=compte_epargne, 
        type_operation='retrait'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Totaux generaux
    total_encaissements = int(total_encaissements_caisse + total_ajouts_uv + total_depots_epargne)
    total_decaissements = int(total_decaissements_caisse + total_retraits_uv + total_retraits_epargne)
    
    # ========== RECUPERATION DE L'HISTORIQUE ==========
    operations = []
    
    # Operations caisse
    queryset_caisse = OperationCaisse.objects.filter(user=request.user)
    for op in queryset_caisse.order_by('-date_operation')[:50]:
        operations.append({
            'date': op.date_operation.strftime('%d/%m/%Y %H:%M:%S'),
            'compte_icon': '💰',
            'compte_nom': 'Especes',
            'type': 'Encaissement' if op.type_operation == 'encaissement' else 'Decaissement',
            'signe': '+' if op.type_operation == 'encaissement' else '-',
            'montant': f'{op.montant:,.0f}',
            'couleur': 'color:#10b981' if op.type_operation == 'encaissement' else 'color:#ef4444',
            'description': op.description or '-'
        })
    
    # Operations UV
    for op in OperationUv.objects.filter(user=request.user).order_by('-date_operation')[:50]:
        operations.append({
            'date': op.date_operation.strftime('%d/%m/%Y %H:%M:%S'),
            'compte_icon': '📱' if op.type_uv == 'touchpoint' else '💳',
            'compte_nom': 'UV Touchpoint' if op.type_uv == 'touchpoint' else 'UV Wave',
            'type': 'Ajout' if op.type_operation == 'ajout' else 'Retrait',
            'signe': '+' if op.type_operation == 'ajout' else '-',
            'montant': f'{op.montant:,.0f}',
            'couleur': 'color:#10b981' if op.type_operation == 'ajout' else 'color:#ef4444',
            'description': op.description or '-'
        })
    
    # Operations Epargne
    if compte_epargne:
        for op in OperationEpargne.objects.filter(compte=compte_epargne).order_by('-date_operation')[:50]:
            operations.append({
                'date': op.date_operation.strftime('%d/%m/%Y %H:%M:%S'),
                'compte_icon': '🏦',
                'compte_nom': 'Epargne',
                'type': 'Ajout' if op.type_operation == 'depot' else 'Retrait',
                'signe': '+' if op.type_operation == 'depot' else '-',
                'montant': f'{op.montant:,.0f}',
                'couleur': 'color:#10b981' if op.type_operation == 'depot' else 'color:#ef4444',
                'description': op.description or '-'
            })
    
    # Trier par date decroissante
    operations.sort(key=lambda x: x['date'], reverse=True)
    
    # Totaux generaux des comptes de l'admin connecte
    total_cash = int(caisse.solde_cash or 0)
    total_uv = int(caisse.solde_uv or 0)
    total_wave = int(caisse.solde_wave or 0)
    total_general = int(total_cash + total_uv + total_wave + (compte_epargne.solde or 0))
    
    context = {
        'title': 'Rapports et Gestion',
        'admin': admin,
        'caisse': caisse,
        'compte_epargne': compte_epargne,
        'operations': operations[:50],
        'total_encaissements': total_encaissements,
        'total_decaissements': total_decaissements,
        'admins': admins,
        'agents': agents,
        'assistants': assistants,
        'agents_actifs_count': agents_actifs_count,
        'total_users': total_users,
        'stats_today': stats_today,
        'evolution': evolution,
        'demandes_attente': demandes_attente,
        'top_agents': top_users,
        'top_users': top_users,
        'dernieres_transactions': dernieres_transactions,
        'total_transactions': total_transactions,
        'total_volume': total_volume,
        'total_commission': total_commission,
        'evolution_transactions': evolution['nombre'],
        'top_agents_performance': user_performance,
        'date_debut': date_debut.strftime('%Y-%m-%d'),
        'date_fin': date_fin.strftime('%Y-%m-%d'),
        'total_general': total_general,
        'filtre_type': filtre_type,
        'filtre_compte': filtre_compte,
        'filtre_date_debut': filtre_date_debut,
        'filtre_date_fin': filtre_date_fin,
    }
    return render(request, 'transactions/rapports_admin.html', context)

@login_required
def api_historique_operations(request):
    """API pour recuperer l'historique des operations"""
    from django.core.paginator import Paginator
    from django.http import JsonResponse
    from datetime import datetime
    
    page = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 10))
    
    operations = []
    
    # 1. Opérations Caisse
    for op in OperationCaisse.objects.filter(user=request.user).order_by('-date_operation'):
        operations.append({
            'date': op.date_operation.strftime('%d/%m/%Y %H:%M:%S'),
            'compte_icon': '💰',
            'compte_nom': 'Espèces',
            'type': 'Encaissement' if op.type_operation == 'encaissement' else 'Décaissement',
            'signe': '+' if op.type_operation == 'encaissement' else '-',
            'montant': f"{op.montant:,.0f}",
            'couleur': '#10b981' if op.type_operation == 'encaissement' else '#ef4444',
            'description': op.description or '-'
        })
    
    # 2. Opérations UV
    for op in OperationUv.objects.filter(user=request.user).order_by('-date_operation'):
        operations.append({
            'date': op.date_operation.strftime('%d/%m/%Y %H:%M:%S'),
            'compte_icon': '📱' if op.type_uv == 'touchpoint' else '💳',
            'compte_nom': 'UV Touchpoint' if op.type_uv == 'touchpoint' else 'UV Wave',
            'type': 'Ajout' if op.type_operation == 'ajout' else 'Retrait',
            'signe': '+' if op.type_operation == 'ajout' else '-',
            'montant': f"{op.montant:,.0f}",
            'couleur': '#10b981' if op.type_operation == 'ajout' else '#ef4444',
            'description': op.description or '-'
        })
    
    # 3. Opérations Epargne
    compte_epargne = CompteEpargneAdmin.objects.filter(user=request.user).first()
    if compte_epargne:
        for op in OperationEpargne.objects.filter(compte=compte_epargne).order_by('-date_operation'):
            operations.append({
                'date': op.date_operation.strftime('%d/%m/%Y %H:%M:%S'),
                'compte_icon': '🏦',
                'compte_nom': 'Epargne',
                'type': 'Dépôt' if op.type_operation == 'depot' else 'Retrait',
                'signe': '+' if op.type_operation == 'depot' else '-',
                'montant': f"{op.montant:,.0f}",
                'couleur': '#10b981' if op.type_operation == 'depot' else '#ef4444',
                'description': op.description or '-'
            })
    
    # Trier par date
    operations.sort(key=lambda x: x['date'], reverse=True)
    
    # Calculer les totaux
    total_encaissements = sum(int(op['montant'].replace(',', '').replace(' ', '')) for op in operations if op['signe'] == '+')
    total_decaissements = sum(int(op['montant'].replace(',', '').replace(' ', '')) for op in operations if op['signe'] == '-')
    
    # Pagination
    paginator = Paginator(operations, per_page)
    try:
        page_obj = paginator.page(page)
    except:
        page_obj = paginator.page(1)
    
    return JsonResponse({
        'success': True,
        'operations': list(page_obj),
        'total': paginator.count,
        'page': page_obj.number,
        'total_pages': paginator.num_pages,
        'has_next': page_obj.has_next(),
        'has_previous': page_obj.has_previous(),
        'per_page': per_page,
        'total_encaissements': f"{total_encaissements:,.0f}",
        'total_decaissements': f"{total_decaissements:,.0f}"
    })

@login_required
def api_totaux_operations(request):
    """API pour recuperer les totaux d'encaissements et decaissements"""
    from django.db.models import Sum
    from django.http import JsonResponse
    
    # Totaux caisse
    total_encaissements_caisse = OperationCaisse.objects.filter(
        user=request.user, 
        type_operation='encaissement'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    total_decaissements_caisse = OperationCaisse.objects.filter(
        user=request.user, 
        type_operation='decaissement'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Totaux UV
    total_ajouts_uv = OperationUv.objects.filter(
        user=request.user, 
        type_operation='ajout'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    total_retraits_uv = OperationUv.objects.filter(
        user=request.user, 
        type_operation='retrait'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Totaux Epargne
    compte_epargne = CompteEpargneAdmin.objects.filter(user=request.user).first()
    total_depots_epargne = 0
    total_retraits_epargne = 0
    
    if compte_epargne:
        total_depots_epargne = OperationEpargne.objects.filter(
            compte=compte_epargne, 
            type_operation='depot'
        ).aggregate(total=Sum('montant'))['total'] or 0
        
        total_retraits_epargne = OperationEpargne.objects.filter(
            compte=compte_epargne, 
            type_operation='retrait'
        ).aggregate(total=Sum('montant'))['total'] or 0
    
    # Totaux generaux
    total_encaissements = int(total_encaissements_caisse + total_ajouts_uv + total_depots_epargne)
    total_decaissements = int(total_decaissements_caisse + total_retraits_uv + total_retraits_epargne)
    
    return JsonResponse({
        'total_encaissements': total_encaissements,
        'total_decaissements': total_decaissements
    })

@login_required
def generer_rapport_admin(request):
    """
    RAPPORT ADMINISTRATEUR - Version Finale
    =======================================
    Contient :
    1. Onglet "RECAP ADMIN" - Totaux généraux + Stats Admin + Soldes Admin + Liste des Agents + Liste des Assistants
    2. Onglet "TRANSACTIONS" - Toutes les transactions
    3. Onglet "DEMANDES" - Toutes les demandes
    4. Onglets individuels pour chaque Agent
    5. Onglets individuels pour chaque Assistant
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.utils.timezone import now
    from datetime import datetime, timedelta
    from django.db.models import Sum, Q
    from transactions.models import Transaction, Agent, Assistant, Admin, Caisse, DemandeApprovisionnement
    from django.contrib.auth.models import User
    
    # ==================== PARAMETRES ====================
    format_type = request.GET.get('format', 'excel')
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    today = now().date()
    
    if date_debut_str and date_fin_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            date_debut = today - timedelta(days=30)
            date_fin = today
    else:
        date_debut = today - timedelta(days=30)
        date_fin = today
    
    # ==================== STYLES EXCEL ====================
    COLOR_PRIMARY = "1a73e8"
    COLOR_DARK = "202124"
    COLOR_SUCCESS = "28a745"
    COLOR_DANGER = "dc3545"
    
    font_title = Font(bold=True, size=16, color="FFFFFF")
    font_subtitle = Font(bold=True, size=12, color=COLOR_PRIMARY)
    font_header = Font(bold=True, size=10, color="FFFFFF")
    font_bold = Font(bold=True)
    font_money = Font(bold=True, size=11, color=COLOR_SUCCESS)
    font_entree = Font(bold=True, size=14, color=COLOR_SUCCESS)
    font_sortie = Font(bold=True, size=14, color=COLOR_DANGER)
    
    fill_title = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
    fill_header = PatternFill(start_color=COLOR_DARK, end_color=COLOR_DARK, fill_type="solid")
    fill_success = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    fill_warning = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    fill_info = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
    fill_entree = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    fill_sortie = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    fill_admin = PatternFill(start_color="e2f0d9", end_color="e2f0d9", fill_type="solid")
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    # ==================== COLLECTE DES DONNEES AVEC FILTRES DATES ====================
    # IMPORTANT: Toutes les transactions sont filtrées par la période
    all_transactions = Transaction.objects.filter(
        date__date__gte=date_debut,
        date__date__lte=date_fin
    ).select_related('user')
    
    # Totaux généraux (déjà filtrés par date)
    total_entree = int(all_transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
    total_sortie = int(all_transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
    total_commission = int(all_transactions.aggregate(Sum('commission'))['commission__sum'] or 0)
    total_transactions = all_transactions.count()
    
    # ==================== STATISTIQUES ADMIN ====================
    admin_stats = {
        'nom': "ADMINISTRATEUR",
        'solde_cash': 0,
        'solde_uv': 0,
        'solde_wave': 0,
        'solde_cash_hier': 0,
        'solde_uv_hier': 0,
        'solde_wave_hier': 0,
        'total_entree': 0,
        'total_sortie': 0,
        'commission': 0,
        'total_transactions': 0,
    }
    
    try:
        admin_user = User.objects.filter(is_superuser=True).first()
        if admin_user:
            caisse_admin = Caisse.objects.get(user=admin_user)
            admin_stats['solde_cash'] = int(caisse_admin.solde_cash or 0)
            admin_stats['solde_uv'] = int(caisse_admin.solde_uv or 0)
            admin_stats['solde_wave'] = int(caisse_admin.solde_wave or 0)
            
            # Transactions de l'admin sur la période
            admin_transactions = all_transactions.filter(user=admin_user)
            admin_stats['total_entree'] = int(admin_transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
            admin_stats['total_sortie'] = int(admin_transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
            admin_stats['commission'] = int(admin_transactions.aggregate(Sum('commission'))['commission__sum'] or 0)
            admin_stats['total_transactions'] = admin_transactions.count()
            
            # Calcul des soldes d'hier (basé sur les transactions d'aujourd'hui)
            transactions_auj = Transaction.objects.filter(date__date=today)
            cash_depot = int(transactions_auj.filter(type_transaction='depot', user=admin_user).aggregate(Sum('montant'))['montant__sum'] or 0)
            cash_retrait = int(transactions_auj.filter(type_transaction='retrait', user=admin_user).aggregate(Sum('montant'))['montant__sum'] or 0)
            admin_stats['solde_cash_hier'] = admin_stats['solde_cash'] - (cash_depot - cash_retrait)
            
            uv_depot = int(transactions_auj.filter(operateur__in=['orange','malitel','telecel'], type_transaction='depot', user=admin_user).aggregate(Sum('montant'))['montant__sum'] or 0)
            uv_retrait = int(transactions_auj.filter(operateur__in=['orange','malitel','telecel'], type_transaction='retrait', user=admin_user).aggregate(Sum('montant'))['montant__sum'] or 0)
            admin_stats['solde_uv_hier'] = admin_stats['solde_uv'] - (uv_retrait - uv_depot)
            
            wave_depot = int(transactions_auj.filter(operateur='wave', type_transaction='depot', user=admin_user).aggregate(Sum('montant'))['montant__sum'] or 0)
            wave_retrait = int(transactions_auj.filter(operateur='wave', type_transaction='retrait', user=admin_user).aggregate(Sum('montant'))['montant__sum'] or 0)
            admin_stats['solde_wave_hier'] = admin_stats['solde_wave'] - (wave_retrait - wave_depot)
    except Exception as e:
        print(f"Erreur admin stats: {e}")
    
    # Statistiques par opérateur (filtrées par date)
    stats = {
        'orange': {
            'depot': int(all_transactions.filter(operateur='orange', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
            'retrait': int(all_transactions.filter(operateur='orange', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
            'credit': int(all_transactions.filter(operateur='orange', type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0),
        },
        'malitel': {
            'depot': int(all_transactions.filter(operateur='malitel', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
            'retrait': int(all_transactions.filter(operateur='malitel', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
            'credit': int(all_transactions.filter(operateur='malitel', type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0),
        },
        'telecel': {
            'depot': int(all_transactions.filter(operateur='telecel', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
            'retrait': int(all_transactions.filter(operateur='telecel', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
            'credit': int(all_transactions.filter(operateur='telecel', type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0),
        },
        'wave': {
            'depot': int(all_transactions.filter(operateur='wave', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0),
            'retrait': int(all_transactions.filter(operateur='wave', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0),
        },
    }
    
    # Demandes filtrées par date
    all_demandes = DemandeApprovisionnement.objects.filter(
        date_demande__date__gte=date_debut,
        date_demande__date__lte=date_fin
    ).order_by('-date_demande')
    
    # ==================== COLLECTE DES AGENTS (avec filtres date) ====================
    agents_data = []
    
    for agent in Agent.objects.all():
        # Transactions de l'agent sur la période UNIQUEMENT
        agent_transactions = all_transactions.filter(user=agent.user)
        
        try:
            caisse = Caisse.objects.get(user=agent.user)
            solde_cash = int(caisse.solde_cash or 0)
            solde_uv = int(caisse.solde_uv or 0)
            solde_wave = int(caisse.solde_wave or 0)
        except Caisse.DoesNotExist:
            solde_cash = solde_uv = solde_wave = 0
        
        # Soldes d'hier (basé sur les transactions d'aujourd'hui seulement)
        transactions_auj = Transaction.objects.filter(date__date=today, user=agent.user)
        cash_depot = int(transactions_auj.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        cash_retrait = int(transactions_auj.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        solde_cash_hier = solde_cash - (cash_depot - cash_retrait)
        
        uv_depot = int(transactions_auj.filter(operateur__in=['orange','malitel','telecel'], type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        uv_retrait = int(transactions_auj.filter(operateur__in=['orange','malitel','telecel'], type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        uv_credit = int(transactions_auj.filter(operateur__in=['orange','malitel','telecel'], type_transaction='credit').aggregate(Sum('montant'))['montant__sum'] or 0)
        solde_uv_hier = solde_uv - (uv_retrait - uv_depot - uv_credit)
        
        wave_depot = int(transactions_auj.filter(operateur='wave', type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        wave_retrait = int(transactions_auj.filter(operateur='wave', type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        solde_wave_hier = solde_wave - (wave_retrait - wave_depot)
        
        # Stats sur la période
        total_entree_agent = int(agent_transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        total_sortie_agent = int(agent_transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        commission_agent = int(agent_transactions.aggregate(Sum('commission'))['commission__sum'] or 0)
        
        agents_data.append({
            'nom': agent.nom,
            'telephone': agent.telephone or '-',
            'email': agent.user.email if agent.user else '-',
            'solde_cash': solde_cash,
            'solde_cash_hier': solde_cash_hier,
            'solde_uv': solde_uv,
            'solde_uv_hier': solde_uv_hier,
            'solde_wave': solde_wave,
            'solde_wave_hier': solde_wave_hier,
            'total_entree': total_entree_agent,
            'total_sortie': total_sortie_agent,
            'commission': commission_agent,
            'total_transactions': agent_transactions.count(),
            'transactions': agent_transactions,
            'demandes': all_demandes.filter(agent=agent),
        })
    
    # ==================== COLLECTE DES ASSISTANTS (avec filtres date) ====================
    assistants_data = []
    
    for assistant in Assistant.objects.all():
        # Transactions de l'assistant sur la période UNIQUEMENT
        assistant_transactions = all_transactions.filter(user=assistant.user)
        
        total_entree_assistant = int(assistant_transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        total_sortie_assistant = int(assistant_transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        commission_assistant = int(assistant_transactions.aggregate(Sum('commission'))['commission__sum'] or 0)
        
        assistants_data.append({
            'nom': assistant.nom,
            'telephone': assistant.telephone or '-',
            'email': assistant.user.email if assistant.user else '-',
            'total_entree': total_entree_assistant,
            'total_sortie': total_sortie_assistant,
            'commission': commission_assistant,
            'total_transactions': assistant_transactions.count(),
            'transactions': assistant_transactions,
            'demandes': all_demandes.filter(assistant_destinataire=assistant),
        })
    
    # ==================== EXPORT EXCEL ====================
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="rapport_admin_{date_debut}_{date_fin}.xlsx"'
    
    wb = Workbook()
    
    # ==================== ONGLET 1: RECAP ADMIN ====================
    ws = wb.active
    ws.title = "1. RECAP ADMIN"
    
    # Titre
    ws.merge_cells('A1:H1')
    ws['A1'] = "📊 RAPPORT ADMINISTRATEUR"
    ws['A1'].font = font_title
    ws['A1'].fill = fill_title
    ws['A1'].alignment = align_center
    
    ws['A2'] = f"📅 Periode: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    ws['A2'].font = font_bold
    ws['A3'] = f"⏰ Date d'export: {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}"
    
    current_row = 5
    
    # ========== SECTION 1: TOTAUX GENERAUX ==========
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "💰 TOTAUX GENERAUX"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_info
    current_row += 1
    
    ws[f'A{current_row}'] = "📈 TOTAL ENTREES"
    ws[f'A{current_row}'].font = font_entree
    ws[f'B{current_row}'] = f"{total_entree:,.0f} FCFA"
    ws[f'B{current_row}'].font = font_entree
    ws[f'B{current_row}'].fill = fill_entree
    ws[f'D{current_row}'] = "📉 TOTAL SORTIES"
    ws[f'D{current_row}'].font = font_sortie
    ws[f'E{current_row}'] = f"{total_sortie:,.0f} FCFA"
    ws[f'E{current_row}'].font = font_sortie
    ws[f'E{current_row}'].fill = fill_sortie
    current_row += 1
    
    ws[f'A{current_row}'] = "🎯 TOTAL COMMISSION"
    ws[f'A{current_row}'].font = font_bold
    ws[f'B{current_row}'] = f"{total_commission:,.0f} FCFA"
    ws[f'B{current_row}'].font = font_money
    ws[f'D{current_row}'] = "📋 NOMBRE DE TRANSACTIONS"
    ws[f'D{current_row}'].font = font_bold
    ws[f'E{current_row}'] = total_transactions
    current_row += 2
    
    # ========== SECTION 2: SOLDES ET ACTIVITE DE L'ADMIN ==========
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "👨‍💼 ADMINISTRATEUR"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_admin
    current_row += 1
    
    # Sous-section: Soldes
    ws[f'A{current_row}'] = "💰 SOLDES"
    ws[f'A{current_row}'].font = font_bold
    current_row += 1
    
    ws[f'A{current_row}'] = "Compte"
    ws[f'B{current_row}'] = "Solde Aujourd'hui"
    ws[f'C{current_row}'] = "Solde Hier"
    ws[f'D{current_row}'] = "Variation"
    for col in range(1, 5):
        ws.cell(row=current_row, column=col).font = font_header
        ws.cell(row=current_row, column=col).fill = fill_header
    current_row += 1
    
    ws[f'A{current_row}'] = "💵 Cash (Espèces)"
    ws[f'B{current_row}'] = f"{admin_stats['solde_cash']:,.0f} FCFA"
    ws[f'C{current_row}'] = f"{admin_stats['solde_cash_hier']:,.0f} FCFA"
    ws[f'D{current_row}'] = f"{admin_stats['solde_cash'] - admin_stats['solde_cash_hier']:+,.0f} FCFA"
    current_row += 1
    
    ws[f'A{current_row}'] = "📱 UV Touchpiont"
    ws[f'B{current_row}'] = f"{admin_stats['solde_uv']:,.0f} FCFA"
    ws[f'C{current_row}'] = f"{admin_stats['solde_uv_hier']:,.0f} FCFA"
    ws[f'D{current_row}'] = f"{admin_stats['solde_uv'] - admin_stats['solde_uv_hier']:+,.0f} FCFA"
    current_row += 1
    
    ws[f'A{current_row}'] = "🌊 UV Wave"
    ws[f'B{current_row}'] = f"{admin_stats['solde_wave']:,.0f} FCFA"
    ws[f'C{current_row}'] = f"{admin_stats['solde_wave_hier']:,.0f} FCFA"
    ws[f'D{current_row}'] = f"{admin_stats['solde_wave'] - admin_stats['solde_wave_hier']:+,.0f} FCFA"
    current_row += 2
    
    # Sous-section: Activité sur la période
    ws[f'A{current_row}'] = "📊 ACTIVITE SUR LA PERIODE"
    ws[f'A{current_row}'].font = font_bold
    current_row += 1
    
    ws[f'A{current_row}'] = f"💰 Entrées: {admin_stats['total_entree']:,.0f} FCFA"
    ws[f'C{current_row}'] = f"💸 Sorties: {admin_stats['total_sortie']:,.0f} FCFA"
    ws[f'E{current_row}'] = f"🎯 Commission: {admin_stats['commission']:,.0f} FCFA"
    ws[f'G{current_row}'] = f"📋 Nb Ops: {admin_stats['total_transactions']}"
    current_row += 2
    
    # ========== SECTION 3: STATISTIQUES PAR OPERATEUR ==========
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "📱 STATISTIQUES PAR OPERATEUR"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_info
    current_row += 1
    
    headers = ['Opérateur', 'Dépôts', 'Retraits', 'Crédits', 'Total', 'Commission']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    current_row += 1
    
    ops_data = [('Orange', 'orange'), ('Malitel', 'malitel'), ('Telecel', 'telecel'), ('Wave', 'wave')]
    for op_name, op_key in ops_data:
        ws.cell(row=current_row, column=1, value=op_name)
        ws.cell(row=current_row, column=2, value=f"{stats[op_key]['depot']:,.0f} FCFA")
        ws.cell(row=current_row, column=3, value=f"{stats[op_key]['retrait']:,.0f} FCFA")
        ws.cell(row=current_row, column=4, value=f"{stats[op_key]['credit']:,.0f} FCFA" if op_key != 'wave' else "-")
        total = stats[op_key]['depot'] + stats[op_key]['retrait'] + (stats[op_key].get('credit', 0) if op_key != 'wave' else 0)
        ws.cell(row=current_row, column=5, value=f"{total:,.0f} FCFA")
        ws.cell(row=current_row, column=6, value=f"{int(total * 0.01):,.0f} FCFA")
        current_row += 1
    current_row += 2
    
    # ========== SECTION 4: LISTE DES AGENTS ==========
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "👥 LISTE DES AGENTS"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_info
    current_row += 1
    
    for agent in agents_data:
        # Nom de l'agent
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = f"👤 {agent['nom']}"
        ws[f'A{current_row}'].font = font_bold
        ws[f'A{current_row}'].fill = fill_info
        current_row += 1
        
        # Téléphone et email
        ws[f'A{current_row}'] = f"📞 {agent['telephone']}  |  ✉️ {agent['email']}"
        current_row += 1
        
        # En-tête du tableau des soldes
        ws[f'A{current_row}'] = "Compte"
        ws[f'B{current_row}'] = "Solde Aujourd'hui"
        ws[f'C{current_row}'] = "Solde Hier"
        ws[f'D{current_row}'] = "Variation"
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).font = font_header
            ws.cell(row=current_row, column=col).fill = fill_header
        current_row += 1
        
        # Cash
        ws[f'A{current_row}'] = "💵 Cash (Espèces)"
        ws[f'B{current_row}'] = f"{agent['solde_cash']:,.0f} FCFA"
        ws[f'C{current_row}'] = f"{agent['solde_cash_hier']:,.0f} FCFA"
        var_cash = agent['solde_cash'] - agent['solde_cash_hier']
        ws[f'D{current_row}'] = f"{var_cash:+,.0f} FCFA"
        ws[f'D{current_row}'].fill = fill_success if var_cash >= 0 else fill_warning
        current_row += 1
        
        # UV
        ws[f'A{current_row}'] = "📱 UV Touchpiont"
        ws[f'B{current_row}'] = f"{agent['solde_uv']:,.0f} FCFA"
        ws[f'C{current_row}'] = f"{agent['solde_uv_hier']:,.0f} FCFA"
        ws[f'D{current_row}'] = f"{agent['solde_uv'] - agent['solde_uv_hier']:+,.0f} FCFA"
        current_row += 1
        
        # Wave
        ws[f'A{current_row}'] = "🌊 UV Wave"
        ws[f'B{current_row}'] = f"{agent['solde_wave']:,.0f} FCFA"
        ws[f'C{current_row}'] = f"{agent['solde_wave_hier']:,.0f} FCFA"
        ws[f'D{current_row}'] = f"{agent['solde_wave'] - agent['solde_wave_hier']:+,.0f} FCFA"
        current_row += 1
        
        # Résumé période
        ws[f'A{current_row}'] = "📊 RÉSUMÉ PÉRIODE"
        ws[f'B{current_row}'] = f"Entrées: {agent['total_entree']:,.0f} FCFA | Sorties: {agent['total_sortie']:,.0f} FCFA | Commission: {agent['commission']:,.0f} FCFA | Nb Ops: {agent['total_transactions']}"
        current_row += 2
        
        # Séparateur
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "-" * 60
        current_row += 1
    
    # ========== SECTION 5: LISTE DES ASSISTANTS (AJOUTÉE) ==========
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "👥 LISTE DES ASSISTANTS"
    ws[f'A{current_row}'].font = font_subtitle
    ws[f'A{current_row}'].fill = fill_info
    current_row += 1
    
    headers_assistants = ['Assistant', 'Téléphone', 'Email', 'Entrées', 'Sorties', 'Commission', 'Nb Ops']
    for col, header in enumerate(headers_assistants, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    current_row += 1
    
    for assistant in assistants_data:
        ws.cell(row=current_row, column=1, value=assistant['nom'])
        ws.cell(row=current_row, column=2, value=assistant['telephone'])
        ws.cell(row=current_row, column=3, value=assistant['email'])
        ws.cell(row=current_row, column=4, value=f"{assistant['total_entree']:,.0f} FCFA")
        ws.cell(row=current_row, column=5, value=f"{assistant['total_sortie']:,.0f} FCFA")
        ws.cell(row=current_row, column=6, value=f"{assistant['commission']:,.0f} FCFA")
        ws.cell(row=current_row, column=7, value=assistant['total_transactions'])
        current_row += 1
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12
    
    # ==================== ONGLET 2: TRANSACTIONS ====================
    ws2 = wb.create_sheet("2. TRANSACTIONS")
    ws2.merge_cells('A1:H1')
    ws2['A1'] = "DETAIL DES TRANSACTIONS"
    ws2['A1'].font = font_title
    ws2['A1'].fill = fill_title
    ws2['A1'].alignment = align_center
    ws2['A2'] = f"📅 Periode: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    
    row = 4
    headers_trans = ['Date', 'Référence', 'Type', 'Opérateur', 'Client', 'Montant', 'Commission', 'Utilisateur']
    for col, header in enumerate(headers_trans, 1):
        cell = ws2.cell(row=row, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    row += 1
    
    for t in all_transactions.order_by('-date'):
        ws2.cell(row=row, column=1, value=t.date.strftime('%d/%m/%Y %H:%M'))
        ws2.cell(row=row, column=2, value=t.reference)
        ws2.cell(row=row, column=3, value=t.get_type_transaction_display())
        ws2.cell(row=row, column=4, value=t.get_operateur_display())
        ws2.cell(row=row, column=5, value=t.numero_client)
        ws2.cell(row=row, column=6, value=f"{int(t.montant):,.0f}")
        ws2.cell(row=row, column=6).font = font_money
        ws2.cell(row=row, column=7, value=f"{int(t.commission):,.0f}")
        ws2.cell(row=row, column=8, value=t.user.username if t.user else "-")
        row += 1
    
    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 16
    
    # ==================== ONGLET 3: DEMANDES ====================
    if all_demandes.exists():
        ws3 = wb.create_sheet("3. DEMANDES")
        ws3.merge_cells('A1:G1')
        ws3['A1'] = "DEMANDES D'APPROVISIONNEMENT"
        ws3['A1'].font = font_title
        ws3['A1'].fill = fill_title
        ws3['A1'].alignment = align_center
        ws3['A2'] = f"📅 Periode: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        
        row = 4
        headers_dem = ['Date', 'Type', 'Montant', 'Statut', 'Motif', 'Agent', 'Assistant']
        for col, header in enumerate(headers_dem, 1):
            cell = ws3.cell(row=row, column=col, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        row += 1
        
        for d in all_demandes:
            ws3.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y %H:%M'))
            ws3.cell(row=row, column=2, value=d.get_type_echange_display())
            ws3.cell(row=row, column=3, value=f"{int(d.montant):,.0f} FCFA")
            ws3.cell(row=row, column=4, value=d.get_statut_display())
            ws3.cell(row=row, column=5, value=d.motif or "-")
            ws3.cell(row=row, column=6, value=d.agent.nom if d.agent else "-")
            ws3.cell(row=row, column=7, value=d.assistant_destinataire.nom if d.assistant_destinataire else "-")
            row += 1
        
        for col in range(1, 8):
            ws3.column_dimensions[get_column_letter(col)].width = 18
    
    # ==================== ONGLETS INDIVIDUELS POUR CHAQUE AGENT ====================
    for agent in agents_data:
        nom_feuille = f"Agent_{agent['nom'][:20]}".replace(' ', '_').replace('-', '_')
        ws_agent = wb.create_sheet(nom_feuille[:25])
        
        ws_agent.merge_cells('A1:D1')
        ws_agent['A1'] = f"AGENT : {agent['nom']}"
        ws_agent['A1'].font = font_title
        ws_agent['A1'].fill = fill_title
        ws_agent['A1'].alignment = align_center
        
        ws_agent['A2'] = f"📞 {agent['telephone']}  |  ✉️ {agent['email']}"
        ws_agent['A3'] = f"📅 Periode: {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        
        row = 5
        
        # SOLDES
        ws_agent.merge_cells(f'A{row}:D{row}')
        ws_agent[f'A{row}'] = "💰 SOLDES"
        ws_agent[f'A{row}'].font = font_subtitle
        ws_agent[f'A{row}'].fill = fill_info
        row += 1
        
        ws_agent[f'A{row}'] = "Compte"
        ws_agent[f'B{row}'] = "Aujourd'hui"
        ws_agent[f'C{row}'] = "Hier"
        ws_agent[f'D{row}'] = "Variation"
        for col in range(1, 5):
            ws_agent.cell(row=row, column=col).font = font_header
            ws_agent.cell(row=row, column=col).fill = fill_header
        row += 1
        
        ws_agent[f'A{row}'] = "💵 Cash"
        ws_agent[f'B{row}'] = f"{agent['solde_cash']:,.0f} FCFA"
        ws_agent[f'C{row}'] = f"{agent['solde_cash_hier']:,.0f} FCFA"
        ws_agent[f'D{row}'] = f"{agent['solde_cash'] - agent['solde_cash_hier']:+,.0f} FCFA"
        row += 1
        
        ws_agent[f'A{row}'] = "📱 UV Touchpiont"
        ws_agent[f'B{row}'] = f"{agent['solde_uv']:,.0f} FCFA"
        ws_agent[f'C{row}'] = f"{agent['solde_uv_hier']:,.0f} FCFA"
        ws_agent[f'D{row}'] = f"{agent['solde_uv'] - agent['solde_uv_hier']:+,.0f} FCFA"
        row += 1
        
        ws_agent[f'A{row}'] = "🌊 UV Wave"
        ws_agent[f'B{row}'] = f"{agent['solde_wave']:,.0f} FCFA"
        ws_agent[f'C{row}'] = f"{agent['solde_wave_hier']:,.0f} FCFA"
        ws_agent[f'D{row}'] = f"{agent['solde_wave'] - agent['solde_wave_hier']:+,.0f} FCFA"
        row += 2
        
        ws_agent[f'A{row}'] = "📊 RÉSUMÉ PÉRIODE"
        ws_agent[f'A{row}'].font = font_bold
        ws_agent[f'B{row}'] = f"Entrées: {agent['total_entree']:,.0f} FCFA"
        ws_agent[f'C{row}'] = f"Sorties: {agent['total_sortie']:,.0f} FCFA"
        ws_agent[f'D{row}'] = f"Commission: {agent['commission']:,.0f} FCFA"
        row += 2
        
        # DEMANDES
        if agent['demandes'].exists():
            ws_agent.merge_cells(f'A{row}:D{row}')
            ws_agent[f'A{row}'] = "📨 DEMANDES"
            ws_agent[f'A{row}'].font = font_subtitle
            ws_agent[f'A{row}'].fill = fill_info
            row += 1
            
            headers_dem = ['Date', 'Type', 'Montant', 'Statut']
            for col, header in enumerate(headers_dem, 1):
                cell = ws_agent.cell(row=row, column=col, value=header)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
            row += 1
            
            for d in agent['demandes']:
                ws_agent.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y'))
                ws_agent.cell(row=row, column=2, value=d.get_type_echange_display())
                ws_agent.cell(row=row, column=3, value=f"{int(d.montant):,.0f} FCFA")
                ws_agent.cell(row=row, column=4, value=d.get_statut_display())
                row += 1
        
        # TRANSACTIONS (colonne F à K)
        col_trans = 6
        row_trans = 5
        
        ws_agent.merge_cells(start_row=row_trans, start_column=col_trans, end_row=row_trans, end_column=col_trans+5)
        ws_agent.cell(row=row_trans, column=col_trans, value="📊 TRANSACTIONS SUR LA PÉRIODE")
        ws_agent.cell(row=row_trans, column=col_trans).font = font_subtitle
        ws_agent.cell(row=row_trans, column=col_trans).fill = fill_info
        ws_agent.cell(row=row_trans, column=col_trans).alignment = align_center
        row_trans += 1
        
        headers_trans_pers = ['Date', 'Type', 'Opérateur', 'Client', 'Montant', 'Commission']
        for col, header in enumerate(headers_trans_pers, col_trans):
            cell = ws_agent.cell(row=row_trans, column=col, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        row_trans += 1
        
        for t in agent['transactions'].order_by('-date'):
            ws_agent.cell(row=row_trans, column=col_trans, value=t.date.strftime('%d/%m/%Y'))
            ws_agent.cell(row=row_trans, column=col_trans+1, value=t.get_type_transaction_display())
            ws_agent.cell(row=row_trans, column=col_trans+2, value=t.get_operateur_display())
            ws_agent.cell(row=row_trans, column=col_trans+3, value=t.numero_client)
            ws_agent.cell(row=row_trans, column=col_trans+4, value=f"{int(t.montant):,.0f}")
            ws_agent.cell(row=row_trans, column=col_trans+5, value=f"{int(t.commission):,.0f}")
            row_trans += 1
        
        if agent['transactions'].count() == 0:
            ws_agent.cell(row=row_trans, column=col_trans, value="Aucune transaction sur cette période")
        
        # Ajuster largeurs
        for col in range(1, 5):
            ws_agent.column_dimensions[get_column_letter(col)].width = 20
        for col in range(6, 12):
            ws_agent.column_dimensions[get_column_letter(col)].width = 16
    
    # ==================== ONGLETS INDIVIDUELS POUR CHAQUE ASSISTANT ====================
    for assistant in assistants_data:
        nom_feuille = f"Assistant_{assistant['nom'][:20]}".replace(' ', '_').replace('-', '_')
        ws_assistant = wb.create_sheet(nom_feuille[:25])
        
        ws_assistant.merge_cells('A1:D1')
        ws_assistant['A1'] = f"ASSISTANT : {assistant['nom']}"
        ws_assistant['A1'].font = font_title
        ws_assistant['A1'].fill = fill_title
        ws_assistant['A1'].alignment = align_center
        
        ws_assistant['A2'] = f"📞 {assistant['telephone']}  |  ✉️ {assistant['email']}"
        ws_assistant['A3'] = f"📅 Periode: {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        ws_assistant['A4'] = "ℹ️ Les assistants partagent le solde de l'administrateur"
        ws_assistant['A4'].fill = fill_info
        
        row = 6
        
        # RÉSUMÉ
        ws_assistant.merge_cells(f'A{row}:D{row}')
        ws_assistant[f'A{row}'] = "📊 RÉSUMÉ DES OPÉRATIONS"
        ws_assistant[f'A{row}'].font = font_subtitle
        ws_assistant[f'A{row}'].fill = fill_info
        row += 1
        
        ws_assistant[f'A{row}'] = "💰 Total Entrées"
        ws_assistant[f'B{row}'] = f"{assistant['total_entree']:,.0f} FCFA"
        ws_assistant[f'C{row}'] = "💸 Total Sorties"
        ws_assistant[f'D{row}'] = f"{assistant['total_sortie']:,.0f} FCFA"
        row += 1
        
        ws_assistant[f'A{row}'] = "🎯 Commission totale"
        ws_assistant[f'B{row}'] = f"{assistant['commission']:,.0f} FCFA"
        ws_assistant[f'C{row}'] = "📋 Nombre de transactions"
        ws_assistant[f'D{row}'] = assistant['total_transactions']
        row += 2
        
        # DEMANDES
        if assistant['demandes'].exists():
            ws_assistant.merge_cells(f'A{row}:D{row}')
            ws_assistant[f'A{row}'] = "📨 DEMANDES TRAITÉES"
            ws_assistant[f'A{row}'].font = font_subtitle
            ws_assistant[f'A{row}'].fill = fill_info
            row += 1
            
            headers_dem = ['Date', 'Type', 'Montant', 'Statut']
            for col, header in enumerate(headers_dem, 1):
                cell = ws_assistant.cell(row=row, column=col, value=header)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
            row += 1
            
            for d in assistant['demandes']:
                ws_assistant.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y'))
                ws_assistant.cell(row=row, column=2, value=d.get_type_echange_display())
                ws_assistant.cell(row=row, column=3, value=f"{int(d.montant):,.0f} FCFA")
                ws_assistant.cell(row=row, column=4, value=d.get_statut_display())
                row += 1
        
        # TRANSACTIONS
        col_trans = 6
        row_trans = 6
        
        ws_assistant.merge_cells(start_row=row_trans, start_column=col_trans, end_row=row_trans, end_column=col_trans+5)
        ws_assistant.cell(row=row_trans, column=col_trans, value="📊 TRANSACTIONS SUR LA PÉRIODE")
        ws_assistant.cell(row=row_trans, column=col_trans).font = font_subtitle
        ws_assistant.cell(row=row_trans, column=col_trans).fill = fill_info
        ws_assistant.cell(row=row_trans, column=col_trans).alignment = align_center
        row_trans += 1
        
        headers_trans_pers = ['Date', 'Type', 'Opérateur', 'Client', 'Montant', 'Commission']
        for col, header in enumerate(headers_trans_pers, col_trans):
            cell = ws_assistant.cell(row=row_trans, column=col, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        row_trans += 1
        
        for t in assistant['transactions'].order_by('-date'):
            ws_assistant.cell(row=row_trans, column=col_trans, value=t.date.strftime('%d/%m/%Y'))
            ws_assistant.cell(row=row_trans, column=col_trans+1, value=t.get_type_transaction_display())
            ws_assistant.cell(row=row_trans, column=col_trans+2, value=t.get_operateur_display())
            ws_assistant.cell(row=row_trans, column=col_trans+3, value=t.numero_client)
            ws_assistant.cell(row=row_trans, column=col_trans+4, value=f"{int(t.montant):,.0f}")
            ws_assistant.cell(row=row_trans, column=col_trans+5, value=f"{int(t.commission):,.0f}")
            row_trans += 1
        
        if assistant['transactions'].count() == 0:
            ws_assistant.cell(row=row_trans, column=col_trans, value="Aucune transaction sur cette période")
        
        # Ajuster largeurs
        for col in range(1, 5):
            ws_assistant.column_dimensions[get_column_letter(col)].width = 20
        for col in range(6, 12):
            ws_assistant.column_dimensions[get_column_letter(col)].width = 16
    
    wb.save(response)
    return response

def export_complete_report(transactions, nom, user_type, caisse, total_entree, total_sortie, total_commission, demandes, format_type, date_debut, date_fin, stats_orange, stats_malitel, stats_telecel, stats_wave):
    """Export complet avec recap, demandes et transactions"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import csv
    from django.http import HttpResponse
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    today = datetime.now().date()
    
    # Styles pour Excel
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="0f766e", end_color="0f766e", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Calcul des soldes hier si caisse disponible
    solde_cash_hier = 0
    solde_uv_hier = 0
    solde_wave_hier = 0
    
    if caisse and nom != "Tous":
        transactions_today = transactions.filter(date__date=today)
        
        cash_depot_today = int(transactions_today.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0)
        cash_retrait_today = int(transactions_today.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0)
        variation_cash_today = cash_depot_today - cash_retrait_today
        
        uv_depot_today = int(transactions_today.filter(
            operateur__in=['orange', 'malitel', 'telecel'],
            type_transaction='depot'
        ).aggregate(Sum('montant'))['montant__sum'] or 0)
        uv_retrait_today = int(transactions_today.filter(
            operateur__in=['orange', 'malitel', 'telecel'],
            type_transaction='retrait'
        ).aggregate(Sum('montant'))['montant__sum'] or 0)
        uv_credit_today = int(transactions_today.filter(
            operateur__in=['orange', 'malitel', 'telecel'],
            type_transaction='credit'
        ).aggregate(Sum('montant'))['montant__sum'] or 0)
        variation_uv_today = uv_retrait_today - uv_depot_today - uv_credit_today
        
        wave_depot_today = int(transactions_today.filter(
            operateur='wave',
            type_transaction='depot'
        ).aggregate(Sum('montant'))['montant__sum'] or 0)
        wave_retrait_today = int(transactions_today.filter(
            operateur='wave',
            type_transaction='retrait'
        ).aggregate(Sum('montant'))['montant__sum'] or 0)
        variation_wave_today = wave_retrait_today - wave_depot_today
        
        solde_cash_hier = int(caisse.solde_cash - variation_cash_today)
        solde_uv_hier = int(caisse.solde_uv - variation_uv_today)
        solde_wave_hier = int(caisse.solde_wave - variation_wave_today)
    
    type_label = ""
    if user_type == "agent":
        type_label = "Agent"
    elif user_type == "assistant":
        type_label = "Assistant"
    elif user_type == "admin":
        type_label = "Admin"
    
    # ==================== EXPORT CSV ====================
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        if nom != "Tous":
            response['Content-Disposition'] = f'attachment; filename="rapport_{nom}_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
        else:
            response['Content-Disposition'] = f'attachment; filename="rapport_global_{date_debut}_{date_fin}.csv"'
        response.write('\ufeff')
        writer = csv.writer(response)
        
        # En-tete
        if nom != "Tous":
            writer.writerow([f"RAPPORT DETAILLE - {nom} ({type_label})"])
        else:
            writer.writerow(["RAPPORT FINANCIER GLOBAL"])
        writer.writerow([f"Periode: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"])
        writer.writerow([f"Date export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        writer.writerow([])
        
        # SOLDES
        if caisse and nom != "Tous":
            writer.writerow(["=== SOLDES ==="])
            writer.writerow(["Compte", "Solde actuel", "Solde hier", "Variation"])
            writer.writerow(["Argent Cash", f"{caisse.solde_cash:,.0f} FCFA", f"{solde_cash_hier:,.0f} FCFA", f"{caisse.solde_cash - solde_cash_hier:+,.0f} FCFA"])
            writer.writerow(["UV Touspiont", f"{caisse.solde_uv:,.0f} FCFA", f"{solde_uv_hier:,.0f} FCFA", f"{caisse.solde_uv - solde_uv_hier:+,.0f} FCFA"])
            writer.writerow(["UV Wave", f"{caisse.solde_wave:,.0f} FCFA", f"{solde_wave_hier:,.0f} FCFA", f"{caisse.solde_wave - solde_wave_hier:+,.0f} FCFA"])
            writer.writerow([])
        
        # TOTAUX TRANSACTIONS
        writer.writerow(["=== TOTAUX DES TRANSACTIONS ==="])
        writer.writerow(["Total Entrees", f"{total_entree:,.0f} FCFA"])
        writer.writerow(["Total Sorties", f"{total_sortie:,.0f} FCFA"])
        writer.writerow(["Total Commission", f"{total_commission:,.0f} FCFA"])
        writer.writerow(["Nombre de transactions", transactions.count()])
        writer.writerow([])
        
        # STATISTIQUES PAR OPERATEUR
        writer.writerow(["=== STATISTIQUES PAR OPERATEUR ==="])
        writer.writerow(["Operateur", "Depots", "Retraits", "Credits", "Total"])
        writer.writerow(["Orange", f"{stats_orange['depot']:,.0f}", f"{stats_orange['retrait']:,.0f}", f"{stats_orange['credit']:,.0f}", f"{stats_orange['depot'] + stats_orange['retrait'] + stats_orange['credit']:,.0f}"])
        writer.writerow(["Malitel", f"{stats_malitel['depot']:,.0f}", f"{stats_malitel['retrait']:,.0f}", f"{stats_malitel['credit']:,.0f}", f"{stats_malitel['depot'] + stats_malitel['retrait'] + stats_malitel['credit']:,.0f}"])
        writer.writerow(["Telecel", f"{stats_telecel['depot']:,.0f}", f"{stats_telecel['retrait']:,.0f}", f"{stats_telecel['credit']:,.0f}", f"{stats_telecel['depot'] + stats_telecel['retrait'] + stats_telecel['credit']:,.0f}"])
        writer.writerow(["Wave", f"{stats_wave['depot']:,.0f}", f"{stats_wave['retrait']:,.0f}", "0", f"{stats_wave['depot'] + stats_wave['retrait']:,.0f}"])
        writer.writerow([])
        
        # DEMANDES
        if demandes:
            writer.writerow(["=== DEMANDES D'APPROVISIONNEMENT ==="])
            writer.writerow(["Date", "Type", "Montant", "Statut", "Motif", "Agent", "Assistant"])
            for d in demandes:
                writer.writerow([
                    d.date_demande.strftime('%d/%m/%Y %H:%M'),
                    d.get_type_echange_display(),
                    f"{int(d.montant):,.0f} FCFA",
                    d.get_statut_display(),
                    d.motif or "",
                    d.agent.nom if d.agent else "-",
                    d.assistant_destinataire.nom if d.assistant_destinataire else "-"
                ])
            writer.writerow([])
        
        # DETAIL DES TRANSACTIONS
        writer.writerow(["=== DETAIL DES TRANSACTIONS ==="])
        writer.writerow(['Reference', 'Type', 'Operateur', 'Client', 'Montant (FCFA)', 'Commission (FCFA)', 'Date'])
        for t in transactions:
            writer.writerow([
                t.reference,
                t.get_type_transaction_display(),
                t.get_operateur_display(),
                t.numero_client,
                f"{int(t.montant):,}",
                f"{int(t.commission):,}",
                t.date.strftime('%d/%m/%Y %H:%M:%S')
            ])
        
        return response
    
    # ==================== EXPORT EXCEL ====================
    else:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        if nom != "Tous":
            response['Content-Disposition'] = f'attachment; filename="rapport_{nom}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        else:
            response['Content-Disposition'] = f'attachment; filename="rapport_global_{date_debut}_{date_fin}.xlsx"'
        
        wb = Workbook()
        
        # ========== FEUILLE 1: RECAPITULATIF ==========
        ws_summary = wb.active
        ws_summary.title = "Recapitulatif"
        
        # Titre principal
        ws_summary.merge_cells('A1:E1')
        if nom != "Tous":
            ws_summary['A1'] = f"RAPPORT - {nom} ({type_label})"
        else:
            ws_summary['A1'] = "RAPPORT FINANCIER GLOBAL"
        ws_summary['A1'].font = title_font
        ws_summary['A1'].fill = title_fill
        ws_summary['A1'].alignment = center_align
        
        ws_summary['A2'] = f"Periode: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        ws_summary['A3'] = f"Date export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        row = 5
        
        # SECTION SOLDES
        if caisse and nom != "Tous":
            ws_summary[f'A{row}'] = "SOLDES"
            ws_summary[f'A{row}'].font = subheader_font
            row += 1
            
            for col, header in enumerate(['Compte', 'Solde actuel', 'Solde hier', 'Variation'], 1):
                cell = ws_summary.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            
            soldes_data = [
                ["Argent Cash", f"{caisse.solde_cash:,.0f} FCFA", f"{solde_cash_hier:,.0f} FCFA", f"{caisse.solde_cash - solde_cash_hier:+,.0f} FCFA"],
                ["UV Touspiont", f"{caisse.solde_uv:,.0f} FCFA", f"{solde_uv_hier:,.0f} FCFA", f"{caisse.solde_uv - solde_uv_hier:+,.0f} FCFA"],
                ["UV Wave", f"{caisse.solde_wave:,.0f} FCFA", f"{solde_wave_hier:,.0f} FCFA", f"{caisse.solde_wave - solde_wave_hier:+,.0f} FCFA"],
            ]
            for data in soldes_data:
                for col, val in enumerate(data, 1):
                    ws_summary.cell(row=row, column=col, value=val)
                row += 1
            row += 1
        
        # SECTION TOTAUX TRANSACTIONS
        ws_summary[f'A{row}'] = "TOTAUX DES TRANSACTIONS"
        ws_summary[f'A{row}'].font = subheader_font
        row += 1
        
        totals_data = [
            ["Total Entrees", f"{total_entree:,.0f} FCFA"],
            ["Total Sorties", f"{total_sortie:,.0f} FCFA"],
            ["Total Commission", f"{total_commission:,.0f} FCFA"],
            ["Nombre de transactions", transactions.count()],
        ]
        for data in totals_data:
            ws_summary[f'A{row}'] = data[0]
            ws_summary[f'B{row}'] = data[1]
            row += 1
        row += 1
        
        # SECTION STATISTIQUES PAR OPERATEUR
        ws_summary[f'A{row}'] = "STATISTIQUES PAR OPERATEUR"
        ws_summary[f'A{row}'].font = subheader_font
        row += 1
        
        for col, header in enumerate(['Operateur', 'Depots', 'Retraits', 'Credits', 'Total'], 1):
            cell = ws_summary.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1
        
        stats_data = [
            ["Orange", stats_orange['depot'], stats_orange['retrait'], stats_orange['credit'], stats_orange['depot'] + stats_orange['retrait'] + stats_orange['credit']],
            ["Malitel", stats_malitel['depot'], stats_malitel['retrait'], stats_malitel['credit'], stats_malitel['depot'] + stats_malitel['retrait'] + stats_malitel['credit']],
            ["Telecel", stats_telecel['depot'], stats_telecel['retrait'], stats_telecel['credit'], stats_telecel['depot'] + stats_telecel['retrait'] + stats_telecel['credit']],
            ["Wave", stats_wave['depot'], stats_wave['retrait'], 0, stats_wave['depot'] + stats_wave['retrait']],
        ]
        for data in stats_data:
            ws_summary[f'A{row}'] = data[0]
            ws_summary[f'B{row}'] = f"{data[1]:,.0f} FCFA"
            ws_summary[f'C{row}'] = f"{data[2]:,.0f} FCFA"
            ws_summary[f'D{row}'] = f"{data[3]:,.0f} FCFA" if data[3] > 0 else "-"
            ws_summary[f'E{row}'] = f"{data[4]:,.0f} FCFA"
            row += 1
        
        # Ajustement largeurs colonnes
        for col in range(1, 6):
            ws_summary.column_dimensions[get_column_letter(col)].width = 22
        
        # ========== FEUILLE 2: DEMANDES ==========
        if demandes:
            ws_demandes = wb.create_sheet("Demandes")
            
            headers_demandes = ['Date', 'Type', 'Montant', 'Statut', 'Motif', 'Agent', 'Assistant']
            for col, header in enumerate(headers_demandes, 1):
                cell = ws_demandes.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            
            for row, d in enumerate(demandes, 2):
                ws_demandes.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y %H:%M'))
                ws_demandes.cell(row=row, column=2, value=d.get_type_echange_display())
                ws_demandes.cell(row=row, column=3, value=f"{int(d.montant):,.0f} FCFA")
                ws_demandes.cell(row=row, column=4, value=d.get_statut_display())
                ws_demandes.cell(row=row, column=5, value=d.motif or "")
                ws_demandes.cell(row=row, column=6, value=d.agent.nom if d.agent else "-")
                ws_demandes.cell(row=row, column=7, value=d.assistant_destinataire.nom if d.assistant_destinataire else "-")
            
            for col in range(1, 8):
                ws_demandes.column_dimensions[get_column_letter(col)].width = 20
        
        # ========== FEUILLE 3: TRANSACTIONS ==========
        ws_trans = wb.create_sheet("Transactions")
        
        headers_trans = ['Reference', 'Type', 'Operateur', 'Client', 'Montant (FCFA)', 'Commission (FCFA)', 'Date']
        for col, header in enumerate(headers_trans, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        for row, t in enumerate(transactions, 2):
            ws_trans.cell(row=row, column=1, value=t.reference)
            ws_trans.cell(row=row, column=2, value=t.get_type_transaction_display())
            ws_trans.cell(row=row, column=3, value=t.get_operateur_display())
            ws_trans.cell(row=row, column=4, value=t.numero_client)
            ws_trans.cell(row=row, column=5, value=int(t.montant))
            ws_trans.cell(row=row, column=6, value=int(t.commission))
            ws_trans.cell(row=row, column=7, value=t.date.strftime('%d/%m/%Y %H:%M:%S'))
        
        for col in range(1, 8):
            ws_trans.column_dimensions[get_column_letter(col)].width = 18
        
        wb.save(response)
        return response
