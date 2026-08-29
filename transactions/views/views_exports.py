# Auto-split depuis transactions/views.py (refactor)
# imports partages (niveau module d'origine)

import base64

from pathlib import Path

import django

from django.shortcuts import render, redirect, get_object_or_404

from django.contrib.auth.decorators import login_required

from django.contrib import messages

from django.http import JsonResponse

from django.views.decorators.http import require_POST

from django.db.models import Sum, Q

from datetime import datetime, timedelta

from decimal import Decimal

import json

from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator

from django.contrib.auth.models import User

from django.contrib.auth.hashers import make_password

from django.contrib.auth import logout

from django.views.decorators.http import require_http_methods

from django.utils import timezone

import csv

from ..models import Admin, Agent, Assistant, Caisse, CompteEpargneAdmin, OperationCaisse, OperationEpargne, OperationUv, Transaction, DemandeApprovisionnement, ApprovisionnementDirect

from openpyxl import Workbook

from openpyxl.styles import Font, Alignment, PatternFill

from reportlab.lib.pagesizes import A4

from reportlab.lib.units import cm

from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from django.http import HttpResponse

from django.template.loader import get_template

from xhtml2pdf import pisa

import io

from ..models import (
    Admin, Agent, Assistant, Caisse, Transaction, DemandeApprovisionnement, 
    ApprovisionnementDirect, Facture, PaiementFacture, Dette, RemboursementDette,
    CompteEpargne, OperationCompte
)

from ..forms import (OrangeTransactionForm, WaveTransactionForm, 
                   MalitelTransactionForm, TelecelTransactionForm)

from django.shortcuts import get_object_or_404

from django.http import JsonResponse

from django.contrib.auth.decorators import login_required

from django.template.loader import render_to_string

from django.utils import timezone

from decimal import Decimal

from django.contrib.auth.decorators import login_required

from django.http import JsonResponse

from django.shortcuts import redirect, render, get_object_or_404

from django.contrib import messages

from django.db.models import Sum

from transactions.models import Agent, Assistant, Admin, Caisse, HistoriqueAgent

from django.core.paginator import Paginator

from datetime import datetime

from django.db.models import Sum, Q

from transactions.models import HistoriqueAgent, Agent

import json

import csv

from datetime import datetime, timedelta

from django.db.models import Sum

from django.core.paginator import Paginator

from django.http import JsonResponse, HttpResponse

from django.shortcuts import render, redirect, get_object_or_404

from django.contrib.auth.decorators import login_required

from django.contrib import messages

from django.utils import timezone

from openpyxl import Workbook

from openpyxl.styles import Font, Alignment, PatternFill

from ..models import (
    Transaction, Agent, Admin, Caisse, DemandeApprovisionnement,
    Facture, PaiementFacture, Dette, RemboursementDette,
    CompteEpargne, OperationCompte
)

from django.contrib.auth.decorators import login_required

from django.shortcuts import render, redirect, get_object_or_404

from django.contrib import messages

from django.db.models import Sum

from django.http import JsonResponse, FileResponse

from django.utils import timezone

from django.views.decorators.csrf import csrf_exempt

from django.views.decorators.http import require_http_methods

import json

from datetime import datetime

from reportlab.lib import colors

from reportlab.lib.pagesizes import A4

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from reportlab.lib.units import cm

from io import BytesIO

from ..models import Dette, RemboursementDette, Debiteur

from reportlab.lib.pagesizes import A4

from reportlab.lib.units import cm

from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from django.http import HttpResponse

from datetime import datetime

from reportlab.lib.units import cm

from reportlab.lib import colors

from reportlab.lib.pagesizes import A4

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from django.http import HttpResponse

from datetime import datetime

from django.db.models import Sum, Count, Avg

from django.utils import timezone

from datetime import timedelta, datetime

from django.http import JsonResponse

from django.contrib.auth.decorators import login_required

from decimal import Decimal

def export_transactions(transactions, agent, caisse, total_entree, total_sortie, total_commission, demandes, format_type):
    """
    Exporte les transactions, demandes et soldes au format CSV ou Excel
    """
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    
    # Calcul des soldes d'hier
    transactions_today = Transaction.objects.filter(user=agent.user, date__date=today)
    
    cash_depot_today = transactions_today.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    cash_retrait_today = transactions_today.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    variation_cash_today = cash_depot_today - cash_retrait_today
    
    uv_depot_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_retrait_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_credit_today = transactions_today.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='credit'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_uv_today = uv_retrait_today - uv_depot_today - uv_credit_today
    
    wave_depot_today = transactions_today.filter(
        operateur='wave',
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    wave_retrait_today = transactions_today.filter(
        operateur='wave',
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_wave_today = wave_retrait_today - wave_depot_today
    
    solde_cash_hier = caisse.solde_cash - variation_cash_today
    solde_uv_hier = caisse.solde_uv - variation_uv_today
    solde_wave_hier = caisse.solde_wave - variation_wave_today
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="rapport_{agent.nom}_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
        
        writer = csv.writer(response)
        
        # En-tête principal
        writer.writerow([f"RAPPORT DETAILLE - {agent.nom}"])
        writer.writerow([f"Date d'export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        writer.writerow([])
        
        # SOLDES
        writer.writerow(["=== SOLDES ==="])
        writer.writerow(["Compte", "Solde actuel", "Solde hier", "Variation"])
        writer.writerow(["Argent Cash", f"{caisse.solde_cash:,.0f} FCFA", f"{solde_cash_hier:,.0f} FCFA", f"{caisse.solde_cash - solde_cash_hier:+,.0f} FCFA"])
        writer.writerow(["UV Touspiont", f"{caisse.solde_uv:,.0f} FCFA", f"{solde_uv_hier:,.0f} FCFA", f"{caisse.solde_uv - solde_uv_hier:+,.0f} FCFA"])
        writer.writerow(["UV Wave", f"{caisse.solde_wave:,.0f} FCFA", f"{solde_wave_hier:,.0f} FCFA", f"{caisse.solde_wave - solde_wave_hier:+,.0f} FCFA"])
        writer.writerow([])
        
        # TOTAUX TRANSACTIONS
        writer.writerow(["=== TOTAUX DES TRANSACTIONS ==="])
        writer.writerow(["Total Entrées", f"{total_entree:,.0f} FCFA"])
        writer.writerow(["Total Sorties", f"{total_sortie:,.0f} FCFA"])
        writer.writerow(["Total Commission", f"{total_commission:,.0f} FCFA"])
        writer.writerow(["Nombre de transactions", transactions.count()])
        writer.writerow([])
        
        # DEMANDES
        writer.writerow(["=== DEMANDES D'APPROVISIONNEMENT ==="])
        writer.writerow(["Date", "Type", "Montant", "Statut", "Motif"])
        for d in demandes:
            writer.writerow([
                d.date_demande.strftime('%d/%m/%Y %H:%M'),
                d.get_type_echange_display(),
                f"{d.montant:,.0f} FCFA",
                d.get_statut_display(),
                d.motif or ""
            ])
        writer.writerow([])
        
        # DETAIL DES TRANSACTIONS
        writer.writerow(["=== DETAIL DES TRANSACTIONS ==="])
        writer.writerow(['Référence', 'Type', 'Opérateur', 'Client', 'Montant (FCFA)', 'Commission (FCFA)', 'Date'])
        
        for t in transactions:
            writer.writerow([
                t.reference,
                t.get_type_transaction_display(),
                t.get_operateur_display(),
                t.numero_client,
                f"{t.montant:,.0f}",
                f"{t.commission:,.0f}",
                t.date.strftime('%d/%m/%Y %H:%M:%S')
            ])
        
        return response
    
    elif format_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="rapport_{agent.nom}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        wb = Workbook()
        
        # Styles
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center')
        
        # ========== FEUILLE 1: RÉCAPITULATIF ==========
        ws_summary = wb.active
        ws_summary.title = "Récapitulatif"
        
        ws_summary['A1'] = f"RAPPORT - {agent.nom}"
        ws_summary['A1'].font = title_font
        ws_summary['A2'] = f"Date d'export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        # Soldes
        ws_summary['A4'] = "SOLDES"
        ws_summary['A4'].font = header_font
        ws_summary['A5'] = "Compte"
        ws_summary['B5'] = "Solde actuel"
        ws_summary['C5'] = "Solde hier"
        ws_summary['D5'] = "Variation"
        
        for col in range(1, 5):
            ws_summary.cell(row=5, column=col).font = header_font
        
        soldes_data = [
            ["Argent Cash", f"{caisse.solde_cash:,.0f} FCFA", f"{solde_cash_hier:,.0f} FCFA", f"{caisse.solde_cash - solde_cash_hier:+,.0f} FCFA"],
            ["UV Touspiont", f"{caisse.solde_uv:,.0f} FCFA", f"{solde_uv_hier:,.0f} FCFA", f"{caisse.solde_uv - solde_uv_hier:+,.0f} FCFA"],
            ["UV Wave", f"{caisse.solde_wave:,.0f} FCFA", f"{solde_wave_hier:,.0f} FCFA", f"{caisse.solde_wave - solde_wave_hier:+,.0f} FCFA"],
        ]
        
        for row, data in enumerate(soldes_data, 6):
            for col, val in enumerate(data, 1):
                ws_summary.cell(row=row, column=col, value=val)
        
        # Totaux transactions
        ws_summary['A10'] = "TOTAUX DES TRANSACTIONS"
        ws_summary['A10'].font = header_font
        ws_summary['A11'] = "Total Entrées"
        ws_summary['B11'] = f"{total_entree:,.0f} FCFA"
        ws_summary['A12'] = "Total Sorties"
        ws_summary['B12'] = f"{total_sortie:,.0f} FCFA"
        ws_summary['A13'] = "Total Commission"
        ws_summary['B13'] = f"{total_commission:,.0f} FCFA"
        ws_summary['A14'] = "Nombre de transactions"
        ws_summary['B14'] = transactions.count()
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 25
        ws_summary.column_dimensions['C'].width = 25
        ws_summary.column_dimensions['D'].width = 20
        
        # ========== FEUILLE 2: DEMANDES ==========
        ws_demandes = wb.create_sheet("Demandes")
        
        headers_demandes = ['Date', 'Type', 'Montant', 'Statut', 'Motif']
        for col, header in enumerate(headers_demandes, 1):
            cell = ws_demandes.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
        
        for row, d in enumerate(demandes, 2):
            ws_demandes.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y %H:%M'))
            ws_demandes.cell(row=row, column=2, value=d.get_type_echange_display())
            ws_demandes.cell(row=row, column=3, value=f"{d.montant:,.0f} FCFA")
            ws_demandes.cell(row=row, column=4, value=d.get_statut_display())
            ws_demandes.cell(row=row, column=5, value=d.motif or "")
        
        for col in range(1, 6):
            ws_demandes.column_dimensions[chr(64 + col)].width = 20
        
        # ========== FEUILLE 3: TRANSACTIONS ==========
        ws_trans = wb.create_sheet("Transactions")
        
        headers_trans = ['Référence', 'Type', 'Opérateur', 'Client', 'Montant (FCFA)', 'Commission (FCFA)', 'Date']
        for col, header in enumerate(headers_trans, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
        
        for row, t in enumerate(transactions, 2):
            ws_trans.cell(row=row, column=1, value=t.reference)
            ws_trans.cell(row=row, column=2, value=t.get_type_transaction_display())
            ws_trans.cell(row=row, column=3, value=t.get_operateur_display())
            ws_trans.cell(row=row, column=4, value=t.numero_client)
            ws_trans.cell(row=row, column=5, value=float(t.montant))
            ws_trans.cell(row=row, column=6, value=float(t.commission))
            ws_trans.cell(row=row, column=7, value=t.date.strftime('%d/%m/%Y %H:%M:%S'))
        
        for col in range(1, 8):
            ws_trans.column_dimensions[chr(64 + col)].width = 18
        
        wb.save(response)
        return response
    
    return None

@login_required
def exporter_historique_agent(request, format_type):
    """
    Exporte les transactions de l'agent avec les filtres appliqués
    Inclut les soldes, variations et demandes
    format_type: 'csv' ou 'excel'
    """
    try:
        agent = Agent.objects.get(user=request.user)
    except Agent.DoesNotExist:
        messages.error(request, 'Vous n\'êtes pas autorisé.')
        return redirect('login')
    
    # Récupérer les dates du filtre
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    # Définir la période à analyser
    if date_debut_str and date_fin_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            date_debut = timezone.now().date()
            date_fin = timezone.now().date()
    else:
        # Par défaut: aujourd'hui
        date_debut = timezone.now().date()
        date_fin = timezone.now().date()
    
    # Vérifier que date_debut <= date_fin
    if date_debut > date_fin:
        date_debut, date_fin = date_fin, date_debut
    
    # Date du jour pour le nom du fichier
    today = timezone.now().date()
    
    # Date de la veille pour calculer solde hier
    date_hier = date_debut - timedelta(days=1)
    
    # Récupérer la caisse de l'agent
    caisse = agent.user.caisse
    
    # Récupérer les transactions pour la période (date__date__range)
    transactions = Transaction.objects.filter(
        user=request.user,
        date__date__range=[date_debut, date_fin]
    ).order_by('-date')
    
    # Récupérer les demandes pour la période
    demandes = DemandeApprovisionnement.objects.filter(
        agent=agent,
        date_demande__date__range=[date_debut, date_fin]
    ).order_by('-date_demande')
    
    # Récupérer les transactions avant la période (jusqu'à la veille inclus)
    transactions_avant = Transaction.objects.filter(
        user=request.user,
        date__date__lte=date_hier
    )
    
    # ========== CALCUL DES TOTAUX POUR LA PÉRIODE ==========
    total_entree = transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    total_sortie = transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    
    # ========== CALCUL DES SOLDES ==========
    
            # 1. Calculer les soldes à HIER (avant la période)
    cash_depot_avant = transactions_avant.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    cash_retrait_avant = transactions_avant.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    variation_cash_avant = cash_depot_avant - cash_retrait_avant
    
    uv_depot_avant = transactions_avant.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_retrait_avant = transactions_avant.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_credit_avant = transactions_avant.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='credit'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_uv_avant = uv_retrait_avant - uv_depot_avant - uv_credit_avant
    
    wave_depot_avant = transactions_avant.filter(
        operateur='wave',
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    wave_retrait_avant = transactions_avant.filter(
        operateur='wave',
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_wave_avant = wave_retrait_avant - wave_depot_avant
    
    # Solde à HIER (avant la période)
    solde_cash_hier = caisse.solde_cash - variation_cash_avant
    solde_uv_hier = caisse.solde_uv - variation_uv_avant
    solde_wave_hier = caisse.solde_wave - variation_wave_avant
    
    # 2. Calculer les variations PENDANT la période
    cash_depot_periode = transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    cash_retrait_periode = transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    variation_cash_periode = cash_depot_periode - cash_retrait_periode
    
    uv_depot_periode = transactions.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_retrait_periode = transactions.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_credit_periode = transactions.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='credit'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_uv_periode = uv_retrait_periode - uv_depot_periode - uv_credit_periode
    
    wave_depot_periode = transactions.filter(
        operateur='wave',
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    wave_retrait_periode = transactions.filter(
        operateur='wave',
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_wave_periode = wave_retrait_periode - wave_depot_periode
    
    # Solde à la FIN de la période
    solde_cash_fin = solde_cash_hier + variation_cash_periode
    solde_uv_fin = solde_uv_hier + variation_uv_periode
    solde_wave_fin = solde_wave_hier + variation_wave_periode
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="historique_{today.strftime("%Y%m%d")}.csv"'
        
        response.write('\ufeff')
        writer = csv.writer(response)
        
        # En-tête principal
        writer.writerow([f"HISTORIQUE COMPLET - {agent.nom}"])
        writer.writerow([f"Période: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"])
        writer.writerow([f"Date d'export: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        writer.writerow([])
        
        # SOLDES
        writer.writerow(["=== SOLDES ==="])
        writer.writerow(["Compte", f"Solde au {date_hier.strftime('%d/%m/%Y')}", f"Solde au {date_fin.strftime('%d/%m/%Y')}", "Variation"])
        writer.writerow(["Argent Cash", f"{solde_cash_hier:,.0f} FCFA", f"{solde_cash_fin:,.0f} FCFA", f"{variation_cash_periode:+,.0f} FCFA"])
        writer.writerow(["UV Touspiont", f"{solde_uv_hier:,.0f} FCFA", f"{solde_uv_fin:,.0f} FCFA", f"{variation_uv_periode:+,.0f} FCFA"])
        writer.writerow(["UV Wave", f"{solde_wave_hier:,.0f} FCFA", f"{solde_wave_fin:,.0f} FCFA", f"{variation_wave_periode:+,.0f} FCFA"])
        writer.writerow([])
        
        # TOTAUX TRANSACTIONS
        writer.writerow(["=== TOTAUX DES TRANSACTIONS ==="])
        writer.writerow([f"Total Entrées (Dépôts)", f"{total_entree:,.0f} FCFA"])
        writer.writerow([f"Total Sorties (Retraits)", f"{total_sortie:,.0f} FCFA"])
        writer.writerow(["Nombre de transactions", transactions.count()])
        writer.writerow([])
        
        # DEMANDES
        writer.writerow(["=== DEMANDES D'APPROVISIONNEMENT ==="])
        writer.writerow(["Date", "Type", "Montant", "Statut", "Motif"])
        for d in demandes:
            writer.writerow([
                d.date_demande.strftime('%d/%m/%Y %H:%M'),
                d.get_type_echange_display(),
                f"{d.montant:,.0f} FCFA",
                d.get_statut_display(),
                d.motif or ""
            ])
        writer.writerow([])
        
        # DETAIL DES TRANSACTIONS
        writer.writerow(["=== DETAIL DES TRANSACTIONS ==="])
        writer.writerow(['Référence', 'Type', 'Opérateur', 'Client', 'Montant (FCFA)', 'Date'])
        
        for t in transactions:
            writer.writerow([
                t.reference,
                t.get_type_transaction_display(),
                t.get_operateur_display(),
                t.numero_client,
                f"{t.montant:,.0f}",
                t.date.strftime('%d/%m/%Y %H:%M:%S')
            ])
        
        return response
    
    elif format_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="historique_{today.strftime("%Y%m%d")}.xlsx"'
        
        wb = Workbook()
        
        # Styles
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center')
        
        # ========== FEUILLE 1: RÉCAPITULATIF ==========
        ws_summary = wb.active
        ws_summary.title = "Récapitulatif"
        
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = f"HISTORIQUE COMPLET - {agent.nom}"
        ws_summary['A1'].font = title_font
        ws_summary['A1'].alignment = center_align
        
        ws_summary['A2'] = f"Période: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        ws_summary['A3'] = f"Date d'export: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        # Soldes
        ws_summary['A5'] = "SOLDES"
        ws_summary['A5'].font = header_font
        ws_summary['A6'] = "Compte"
        ws_summary['B6'] = f"Solde au {date_hier.strftime('%d/%m/%Y')}"
        ws_summary['C6'] = f"Solde au {date_fin.strftime('%d/%m/%Y')}"
        ws_summary['D6'] = "Variation"
        
        for col in range(1, 5):
            ws_summary.cell(row=6, column=col).font = header_font
            ws_summary.cell(row=6, column=col).alignment = center_align
        
        soldes_data = [
            ["Argent Cash", f"{solde_cash_hier:,.0f} FCFA", f"{solde_cash_fin:,.0f} FCFA", f"{variation_cash_periode:+,.0f} FCFA"],
            ["UV Touspiont", f"{solde_uv_hier:,.0f} FCFA", f"{solde_uv_fin:,.0f} FCFA", f"{variation_uv_periode:+,.0f} FCFA"],
            ["UV Wave", f"{solde_wave_hier:,.0f} FCFA", f"{solde_wave_fin:,.0f} FCFA", f"{variation_wave_periode:+,.0f} FCFA"],
        ]
        
        for row, data in enumerate(soldes_data, 7):
            for col, val in enumerate(data, 1):
                ws_summary.cell(row=row, column=col, value=val)
        
        # Totaux transactions
        ws_summary['A11'] = "TOTAUX DES TRANSACTIONS"
        ws_summary['A11'].font = header_font
        ws_summary['A12'] = "Total Entrées (Dépôts)"
        ws_summary['B12'] = f"{total_entree:,.0f} FCFA"
        ws_summary['A13'] = "Total Sorties (Retraits)"
        ws_summary['B13'] = f"{total_sortie:,.0f} FCFA"
        ws_summary['A14'] = "Nombre de transactions"
        ws_summary['B14'] = transactions.count()
        
        # Stats des demandes
        ws_summary['A16'] = "STATISTIQUES DES DEMANDES"
        ws_summary['A16'].font = header_font
        ws_summary['A17'] = "En attente"
        ws_summary['B17'] = demandes.filter(statut='attente').count()
        ws_summary['A18'] = "Validées"
        ws_summary['B18'] = demandes.filter(statut='valide').count()
        ws_summary['A19'] = "Refusées"
        ws_summary['B19'] = demandes.filter(statut='refuse').count()
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 25
        ws_summary.column_dimensions['C'].width = 25
        ws_summary.column_dimensions['D'].width = 20
        
        # ========== FEUILLE 2: DEMANDES ==========
        ws_demandes = wb.create_sheet("Demandes")
        
        headers_demandes = ['Date', 'Type', 'Montant', 'Statut', 'Motif']
        for col, header in enumerate(headers_demandes, 1):
            cell = ws_demandes.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
        
        for row, d in enumerate(demandes, 2):
            ws_demandes.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y %H:%M'))
            ws_demandes.cell(row=row, column=2, value=d.get_type_echange_display())
            ws_demandes.cell(row=row, column=3, value=f"{d.montant:,.0f} FCFA")
            ws_demandes.cell(row=row, column=4, value=d.get_statut_display())
            ws_demandes.cell(row=row, column=5, value=d.motif or "")
        
        for col in range(1, 6):
            ws_demandes.column_dimensions[chr(64 + col)].width = 20
        
        # ========== FEUILLE 3: TRANSACTIONS ==========
        ws_trans = wb.create_sheet("Transactions")
        
        headers_trans = ['Référence', 'Type', 'Opérateur', 'Client', 'Montant (FCFA)', 'Date']
        for col, header in enumerate(headers_trans, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
        
        for row, t in enumerate(transactions, 2):
            ws_trans.cell(row=row, column=1, value=t.reference)
            ws_trans.cell(row=row, column=2, value=t.get_type_transaction_display())
            ws_trans.cell(row=row, column=3, value=t.get_operateur_display())
            ws_trans.cell(row=row, column=4, value=t.numero_client)
            ws_trans.cell(row=row, column=5, value=float(t.montant))
            ws_trans.cell(row=row, column=6, value=t.date.strftime('%d/%m/%Y %H:%M:%S'))
        
        # Format des nombres
        for row in range(2, transactions.count() + 2):
            ws_trans.cell(row=row, column=5).number_format = '#,##0'
        
        for col in range(1, 7):
            ws_trans.column_dimensions[chr(64 + col)].width = 18
        
        wb.save(response)
        return response
    
    return None

@login_required
def exporter_rapport_complet_agent(request, format_type):
    """
    Exporte un rapport complet: soldes, transactions, demandes
    Pour agent ou admin
    format_type: 'csv', 'excel' ou 'pdf'
    Prend en compte les filtres de date pour calculer les soldes
    """
    # Récupérer les dates du filtre (si présentes)
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    # Définir la période à analyser
    if date_debut_str and date_fin_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            date_debut = timezone.now().date()
            date_fin = timezone.now().date()
    else:
        # Par défaut: aujourd'hui
        date_debut = timezone.now().date()
        date_fin = timezone.now().date()
    
    # Date du jour pour le nom du fichier
    today = timezone.now().date()
    
    # Date de la veille pour calculer solde hier
    date_hier = date_debut - timedelta(days=1)
    
    # Vérifier si l'utilisateur est un agent ou un admin
    if hasattr(request.user, 'agent_profile'):
        # C'est un agent
        agent = request.user.agent_profile
        caisse = agent.user.caisse
        user_type = "Agent"
        user_name = agent.nom
        
        # Récupérer les transactions de l'agent pour la période
        transactions = Transaction.objects.filter(
            user=request.user,
            date__date__gte=date_debut,
            date__date__lte=date_fin
        ).order_by('-date')
        
        # Récupérer les demandes de l'agent UNIQUEMENT pour la période
        demandes = DemandeApprovisionnement.objects.filter(
            agent=agent,
            date_demande__date__gte=date_debut,
            date_demande__date__lte=date_fin
        ).order_by('-date_demande')
        
        # Calculer les soldes à la date de début
        transactions_avant = Transaction.objects.filter(
            user=request.user,
            date__date__lt=date_debut
        )
        
    elif hasattr(request.user, 'admin_profile'):
        # C'est un admin
        admin = request.user.admin_profile
        caisse = request.user.caisse
        user_type = "Administrateur"
        user_name = admin.nom
        
        # Récupérer toutes les transactions pour la période
        transactions = Transaction.objects.filter(
            date__date__gte=date_debut,
            date__date__lte=date_fin
        ).order_by('-date')
        
        # Récupérer toutes les demandes pour la période
        demandes = DemandeApprovisionnement.objects.filter(
            date_demande__date__gte=date_debut,
            date_demande__date__lte=date_fin
        ).order_by('-date_demande')
        
        # Calculer les soldes à la date de début
        transactions_avant = Transaction.objects.filter(
            date__date__lt=date_debut
        )
        
    elif hasattr(request.user, 'assistant_profile'):
        # C'est un assistant
        assistant = request.user.assistant_profile
        caisse = assistant.get_caisse
        user_type = "Assistant"
        user_name = assistant.nom
        
        # Récupérer les transactions de l'assistant pour la période
        transactions = Transaction.objects.filter(
            user=request.user,
            date__date__gte=date_debut,
            date__date__lte=date_fin
        ).order_by('-date')
        
        # Récupérer les demandes
        demandes = DemandeApprovisionnement.objects.filter(
            date_demande__date__gte=date_debut,
            date_demande__date__lte=date_fin
        ).order_by('-date_demande')
        
        # Calculer les soldes à la date de début
        transactions_avant = Transaction.objects.filter(
            user=request.user,
            date__date__lt=date_debut
        )
        
    else:
        # Superutilisateur ou autre
        try:
            caisse = Caisse.objects.get(user=request.user)
            user_type = "Utilisateur"
            user_name = request.user.username
            transactions = Transaction.objects.filter(
                user=request.user,
                date__date__gte=date_debut,
                date__date__lte=date_fin
            ).order_by('-date')
            demandes = DemandeApprovisionnement.objects.filter(
                agent__user=request.user,
                date_demande__date__gte=date_debut,
                date_demande__date__lte=date_fin
            ).order_by('-date_demande')
            transactions_avant = Transaction.objects.filter(
                user=request.user,
                date__date__lt=date_debut
            )
        except:
            return HttpResponse("Impossible de générer le rapport. Données manquantes.", status=400)
    
    # ========== CALCUL DES TOTAUX POUR LA PÉRIODE ==========
    total_entree = transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    total_sortie = transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    
    # ========== CALCUL DES SOLDES ==========
    
    # 1. Calculer les soldes à HIER (avant la période)
    cash_depot_avant = transactions_avant.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    cash_retrait_avant = transactions_avant.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    variation_cash_avant = cash_depot_avant - cash_retrait_avant
    
    uv_depot_avant = transactions_avant.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_retrait_avant = transactions_avant.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_credit_avant = transactions_avant.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='credit'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_uv_avant = uv_retrait_avant - uv_depot_avant - uv_credit_avant
    
    wave_depot_avant = transactions_avant.filter(
        operateur='wave',
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    wave_retrait_avant = transactions_avant.filter(
        operateur='wave',
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_wave_avant = wave_retrait_avant - wave_depot_avant
    
    # Solde à HIER (avant la période)
    solde_cash_hier = caisse.solde_cash - variation_cash_avant
    solde_uv_hier = caisse.solde_uv - variation_uv_avant
    solde_wave_hier = caisse.solde_wave - variation_wave_avant
    
    # 2. Calculer les variations PENDANT la période
    cash_depot_periode = transactions.filter(type_transaction='depot').aggregate(Sum('montant'))['montant__sum'] or 0
    cash_retrait_periode = transactions.filter(type_transaction='retrait').aggregate(Sum('montant'))['montant__sum'] or 0
    variation_cash_periode = cash_depot_periode - cash_retrait_periode
    
    uv_depot_periode = transactions.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_retrait_periode = transactions.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    uv_credit_periode = transactions.filter(
        operateur__in=['orange', 'malitel', 'telecel'],
        type_transaction='credit'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_uv_periode = uv_retrait_periode - uv_depot_periode - uv_credit_periode
    
    wave_depot_periode = transactions.filter(
        operateur='wave',
        type_transaction='depot'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    wave_retrait_periode = transactions.filter(
        operateur='wave',
        type_transaction='retrait'
    ).aggregate(Sum('montant'))['montant__sum'] or 0
    variation_wave_periode = wave_retrait_periode - wave_depot_periode
    
    # Solde à la FIN de la période
    solde_cash_fin = solde_cash_hier + variation_cash_periode
    solde_uv_fin = solde_uv_hier + variation_uv_periode
    solde_wave_fin = solde_wave_hier + variation_wave_periode
    
    # ========== EXPORT PDF ==========
    if format_type == 'pdf':
        # Créer le contexte pour le template PDF
        context = {
            'user_name': user_name,
            'user_type': user_type,
            'date_export': timezone.now(),
            'date_debut': date_debut,
            'date_fin': date_fin,
            'date_hier': date_hier,
            'solde_cash_hier': solde_cash_hier,
            'solde_uv_hier': solde_uv_hier,
            'solde_wave_hier': solde_wave_hier,
            'solde_cash_fin': solde_cash_fin,
            'solde_uv_fin': solde_uv_fin,
            'solde_wave_fin': solde_wave_fin,
            'variation_cash_periode': variation_cash_periode,
            'variation_uv_periode': variation_uv_periode,
            'variation_wave_periode': variation_wave_periode,
            'total_entree': total_entree,
            'total_sortie': total_sortie,
            'transactions': transactions,
            'demandes': demandes,
        }
        
        # Rendre le template HTML
        template = get_template('transactions/rapport_pdf.html')
        html = template.render(context)
        
        # Créer le PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_{today.strftime("%Y%m%d")}.pdf"'
        
        # Convertir HTML en PDF
        pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('UTF-8')), dest=response)
        
        if pisa_status.err:
            return HttpResponse('Erreur lors de la génération du PDF', status=400)
        
        return response
    
    # ========== EXPORT CSV ==========
    elif format_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="rapport_{today.strftime("%Y%m%d")}.csv"'
        
        response.write('\ufeff')
        writer = csv.writer(response)
        
        writer.writerow([f"RAPPORT COMPLET - {user_name} ({user_type})"])
        writer.writerow([f"Date d'export: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        writer.writerow([])
        
        writer.writerow(["=== SOLDES ==="])
        writer.writerow(["Compte", f"Solde au {date_hier.strftime('%d/%m/%Y')}", f"Solde au {date_fin.strftime('%d/%m/%Y')}", "Variation"])
        writer.writerow(["Argent Cash", f"{solde_cash_hier:,.0f} FCFA", f"{solde_cash_fin:,.0f} FCFA", f"{variation_cash_periode:+,.0f} FCFA"])
        writer.writerow(["UV Touspiont", f"{solde_uv_hier:,.0f} FCFA", f"{solde_uv_fin:,.0f} FCFA", f"{variation_uv_periode:+,.0f} FCFA"])
        writer.writerow(["UV Wave", f"{solde_wave_hier:,.0f} FCFA", f"{solde_wave_fin:,.0f} FCFA", f"{variation_wave_periode:+,.0f} FCFA"])
        writer.writerow([])
        
        writer.writerow(["=== TOTAUX DES TRANSACTIONS ==="])
        writer.writerow([f"Total Entrées (Dépôts)", f"{total_entree:,.0f} FCFA"])
        writer.writerow([f"Total Sorties (Retraits)", f"{total_sortie:,.0f} FCFA"])
        writer.writerow(["Nombre de transactions", transactions.count()])
        writer.writerow([])
        
        writer.writerow(["=== DEMANDES D'APPROVISIONNEMENT ==="])
        writer.writerow(["Date", "Type", "Montant", "Statut", "Motif"])
        for d in demandes:
            writer.writerow([
                d.date_demande.strftime('%d/%m/%Y %H:%M'),
                d.get_type_echange_display(),
                f"{d.montant:,.0f} FCFA",
                d.get_statut_display(),
                d.motif or ""
            ])
        writer.writerow([])
        
        writer.writerow(["=== DETAIL DES TRANSACTIONS ==="])
        writer.writerow(['Référence', 'Type', 'Opérateur', 'Client', 'Montant (FCFA)', 'Date'])
        
        for t in transactions:
            writer.writerow([
                t.reference,
                t.get_type_transaction_display(),
                t.get_operateur_display(),
                t.numero_client,
                f"{t.montant:,.0f}",
                t.date.strftime('%d/%m/%Y %H:%M:%S')
            ])
        
        return response
    
    # ========== EXPORT EXCEL ==========
    elif format_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="rapport_{today.strftime("%Y%m%d")}.xlsx"'
        
        wb = Workbook()
        
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center')
        
        ws_summary = wb.active
        ws_summary.title = "Récapitulatif"
        
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = f"RAPPORT COMPLET - {user_name} ({user_type})"
        ws_summary['A1'].font = title_font
        ws_summary['A1'].alignment = center_align
        
        ws_summary['A2'] = f"Date d'export: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        ws_summary['A4'] = "SOLDES"
        ws_summary['A4'].font = header_font
        ws_summary['A5'] = "Compte"
        ws_summary['B5'] = f"Solde au {date_hier.strftime('%d/%m/%Y')}"
        ws_summary['C5'] = f"Solde au {date_fin.strftime('%d/%m/%Y')}"
        ws_summary['D5'] = "Variation"
        
        for col in range(1, 5):
            ws_summary.cell(row=5, column=col).font = header_font
            ws_summary.cell(row=5, column=col).alignment = center_align
        
        soldes_data = [
            ["Argent Cash", f"{solde_cash_hier:,.0f} FCFA", f"{solde_cash_fin:,.0f} FCFA", f"{variation_cash_periode:+,.0f} FCFA"],
            ["UV Touspiont", f"{solde_uv_hier:,.0f} FCFA", f"{solde_uv_fin:,.0f} FCFA", f"{variation_uv_periode:+,.0f} FCFA"],
            ["UV Wave", f"{solde_wave_hier:,.0f} FCFA", f"{solde_wave_fin:,.0f} FCFA", f"{variation_wave_periode:+,.0f} FCFA"],
        ]
        
        for row, data in enumerate(soldes_data, 6):
            for col, val in enumerate(data, 1):
                ws_summary.cell(row=row, column=col, value=val)
        
        ws_summary['A10'] = "TOTAUX DES TRANSACTIONS"
        ws_summary['A10'].font = header_font
        ws_summary['A11'] = "Total Entrées (Dépôts)"
        ws_summary['B11'] = f"{total_entree:,.0f} FCFA"
        ws_summary['A12'] = "Total Sorties (Retraits)"
        ws_summary['B12'] = f"{total_sortie:,.0f} FCFA"
        ws_summary['A13'] = "Nombre de transactions"
        ws_summary['B13'] = transactions.count()
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 25
        ws_summary.column_dimensions['C'].width = 25
        ws_summary.column_dimensions['D'].width = 20
        
        ws_demandes = wb.create_sheet("Demandes")
        
        headers_demandes = ['Date', 'Type', 'Montant', 'Statut', 'Motif']
        for col, header in enumerate(headers_demandes, 1):
            cell = ws_demandes.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
        
        for row, d in enumerate(demandes, 2):
            ws_demandes.cell(row=row, column=1, value=d.date_demande.strftime('%d/%m/%Y %H:%M'))
            ws_demandes.cell(row=row, column=2, value=d.get_type_echange_display())
            ws_demandes.cell(row=row, column=3, value=f"{d.montant:,.0f} FCFA")
            ws_demandes.cell(row=row, column=4, value=d.get_statut_display())
            ws_demandes.cell(row=row, column=5, value=d.motif or "")
        
        for col in range(1, 6):
            ws_demandes.column_dimensions[chr(64 + col)].width = 20
        
        ws_trans = wb.create_sheet("Transactions")
        
        headers_trans = ['Référence', 'Type', 'Opérateur', 'Client', 'Montant (FCFA)', 'Date']
        for col, header in enumerate(headers_trans, 1):
            cell = ws_trans.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
        
        for row, t in enumerate(transactions, 2):
            ws_trans.cell(row=row, column=1, value=t.reference)
            ws_trans.cell(row=row, column=2, value=t.get_type_transaction_display())
            ws_trans.cell(row=row, column=3, value=t.get_operateur_display())
            ws_trans.cell(row=row, column=4, value=t.numero_client)
            ws_trans.cell(row=row, column=5, value=float(t.montant))
            ws_trans.cell(row=row, column=6, value=t.date.strftime('%d/%m/%Y %H:%M:%S'))
        
        for row in range(2, transactions.count() + 2):
            ws_trans.cell(row=row, column=5).number_format = '#,##0'
        
        for col in range(1, 7):
            ws_trans.column_dimensions[chr(64 + col)].width = 18
        
        wb.save(response)
        return response
    
    return None
